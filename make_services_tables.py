#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_services_tables.py
=======================

Turn the raw ITC (International Trade Centre) services Excel downloads
into styled Table 1-4 + Figure 1 workbooks that the services report
generator (``generate_services_report.py``) expects.

Raw source files expected in ``--excel-dir`` (located by keywords in the
filename, case-insensitive; all Excel formats accepted: .xls, .xlsx, .xlsm,
.xlsb, .csv):

    List_of_exported_services_for_the_selected_service_*.xls   -> global service exports by category
    List_of_exporters_for_the_selected_service_*.xls           -> global top exporters by country
    List_of_imported_services_for_the_selected_service_*.xls   -> global service imports by category
    List_of_importers_for_the_selected_service_*.xls           -> global top importers by country
    List_of_services_exported_by_Kenya_*.xls                   -> Kenya service exports by category
    Data manipulation:
    * Source values are in USD Thousand.
    * Global tables (1, 2): divide by 1,000,000 -> "Value in USD Billion".
    * Kenya tables (3, 4):  divide by 1,000     -> "Value in USD Million".
    * Figure 1 (balance):   divide by 1,000     -> "Value in USD Million".
"""

import argparse
import collections
import csv
import gzip
import os
import re
import shutil
import unicodedata
import sys
import zipfile

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

import openpyxl
from lxml import etree
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from charts import draw_share_pie, new_fig, finish
from country_names import display_name, short_product_name
from xlsx_compat import HTMLTableParser, convert_to_xlsx, is_spreadsheet

# ---------------------------------------------------------------------------
# Constants / styles (shared with make_tables.py)
# ---------------------------------------------------------------------------
# Professional color palette (muted, modern tones)
FONT = "Calibri"
FONT_SIZE = 10
HEADER_FONT_SIZE = 10

TOP_EXPORTERS = 25

# Header styles - deep navy with white text
HDR_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
HDR_FONT_COLOR = "FFFFFF"

# Alternating row fills (very subtle gray tones)
ROW_FILL_EVEN = PatternFill(fill_type="solid", fgColor="F2F2F2")
ROW_FILL_ODD = None  # No fill for odd rows (white)

# Kenya highlight (soft gold)
KENYA_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")

# Red highlight for the focus country (Kenya) in the development-status tables
# (always shown, even off top-N).
DEV_PIN_FILL = PatternFill(fill_type="solid", fgColor="C00000")
DEV_PIN_TEXT = "FFC7CE"

# Focus country that must always appear in the dev-status tables (below top-N
# if needed), normalized via _norm_country().
PINNED_DEV_COUNTRIES = {"kenya"}

# Optional UNCTADstat *annual* service-by-category file fetched from the
# US.TradeServCatTotal/cur Facts OData endpoint (see fetch_unctad_tradeserv.py).
# The all-economy file (``UNCTAD_*_tradeserv_annual_all.csv.gz``) is the single
# canonical annual source: it supplies Kenya's per-category detail, economy-level
# total services + the true published World total, and per-category world sums.
UNCTAD_ANNUAL_KEYWORDS = ("tradeserv",)

# UNCTADstat data exported from the R helper (Services: Trade by category -
# Annual).  The R download writes ``unctad_services_<years>.csv/.xlsx`` with a
# long format carrying every economy (no numeric ``Economy/Code`` column), the
# full category list incl. the granular aggregates (``S``, ``SPX1``, ``SOX``,
# ...) and Category_Type TOTAL/DETAIL.  When supplied alongside the ITC files it
# is the preferred UNCTAD source (bridges Kenya's 2024 category detail and the
# World/Country total-services ``S`` totals).
UNCTAD_R_KEYWORDS = ("unctad_services",)

# UNCTADstat EBOPS service code -> ITC Trade Map service code for the
# categories that map cleanly between the two sources.
UNCTAD_CATEGORY_TO_ITC_CODE = {
    "SA": "1",   # Manufacturing services on physical inputs owned by others (goods-related)
    "SB": "2",   # Maintenance and repair services n.i.e.
    "SC": "3",   # Transport
    "SD": "4",   # Travel
    "SE": "5",   # Construction
    "SF": "6",   # Insurance and pension services
    "SG": "7",   # Financial services
    "SH": "8",   # Charges for the use of intellectual property n.i.e.
    "SI": "9",   # Telecommunications, computer, and information services
    "SJ": "10",  # Other business services
    "SK": "11",  # Personal, cultural, and recreational services
    "SL": "12",  # Government goods and services n.i.e.
}


def _norm_country(name):
    """Normalise a country name for stable comparison (handles 'ü' -> 'u')."""
    return (name or "").strip().lower().replace("ü", "u").replace("Ü", "u")

# Band/summary row (light blue-gray)
BAND_FILL = PatternFill(fill_type="solid", fgColor="D6DCE4")

# Development status fills
DEV_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE")      # Developed - soft blue
DEVEL_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")    # Developing - soft green
LDC_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")      # LDC - soft orange

# Region fills (muted, professional)
REGION_FILLS = {
    "Africa": PatternFill(fill_type="solid", fgColor="D9E2F3"),
    "Asia": PatternFill(fill_type="solid", fgColor="E2EFDA"),
    "Americas": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Pacific": PatternFill(fill_type="solid", fgColor="DAEEF3"),
    "Europe": PatternFill(fill_type="solid", fgColor="EDEDED"),
    "Other": PatternFill(fill_type="solid", fgColor="F2F2F2"),
}

# Border styles - thin bottom only for headers, no side borders
THIN = Side(style="thin", color="B4C6E7")  # Light blue-gray
MEDIUM = Side(style="medium", color="1F3864")  # Navy accent
HEADER_BORDER = Border(bottom=MEDIUM)
DATA_BORDER = Border(bottom=Side(style="thin", color="D6DCE4"))
TOTAL_BORDER = Border(top=Side(style="thin", color="808080"),
                      bottom=Side(style="double", color="808080"))

# Legacy border (kept for compatibility)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Number formats
FMT_VALUE = "#,##0.0"
FMT_VALUE_INT = "#,##0"
FMT_SHARE = "0.0%"
FMT_GROWTH = "0.0%"
FMT_CAGR = "0.0%"
FMT_RCA = "0.00"

BALANCE_WIDTHS = {"A": 16, "B": 30, "C": 12, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12, "J": 12, "K": 12}

# ---------------------------------------------------------------------------
# Development status classification (HDI-based / World Bank / UN LDC)
# ---------------------------------------------------------------------------
CLASSIFICATION_PATH = os.path.join(os.path.dirname(__file__),
                                   "services", "CLASS_2026_07_15.xlsx")
HDI_PATH = os.path.join(os.path.dirname(__file__),
                        "services", "devcountries-2026.csv")

# HDI thresholds (UNDP standard)
HDI_DEVELOPED_THRESHOLD = 0.800

# UN Least Developed Countries (December 2024)
LDC_COUNTRIES = {
    "angola", "benin", "burkina faso", "burundi", "central african republic",
    "chad", "comoros", "congo, democratic republic of the", "djibouti",
    "eritrea", "ethiopia", "gambia", "guinea", "guinea-bissau", "lesotho",
    "liberia", "madagascar", "malawi", "mali", "mauritania", "mozambique",
    "niger", "rwanda", "senegal", "sierra leone", "somalia", "south sudan",
    "sudan", "togo", "uganda", "united republic of tanzania", "zambia",
    "afghanistan", "bangladesh", "cambodia", "lao people's democratic republic",
    "myanmar", "nepal", "timor-leste", "yemen",
    "haiti",
    "kiribati", "solomon islands", "tuvalu",
}

# LDCs that have graduated to developing status
LDC_GRADUATED = {
    "botswana", "cabo verde", "maldives", "samoa",
    "equatorial guinea", "vanuatu", "bhutan", "sao tome and principe",
}

# Countries with high HDI that are still considered developing economies
DEVELOPING_OVERRIDE = {
    "turkey", "türkiye",
}

# ITC name -> HDI / World Bank name aliases (for mismatches between data sources)
NAME_ALIASES = {
    "united states of america": "united states",
    "czech republic": "czechia",
    "hong kong, china": "hong kong",
    "macao, china": "macao sar, china",
    "iran, islamic republic of": "iran",
    "tanzania, united republic of": "tanzania",
    "macedonia, north": "north macedonia",
    "moldova, republic of": "moldova",
    "bolivia, plurinational state of": "bolivia",
    "palestine, state of": "palestine",
    "venezuela, bolivarian republic of": "venezuela",
    "libya, state of": "libya",
    "congo, democratic republic of the": "dr congo",
    "congo": "republic of the congo",
    "korea, republic of": "south korea",
    "russian federation": "russia",
    "türkiye": "turkey",
    "viet nam": "vietnam",
    "lao people's democratic republic": "lao pdr",
    "anguilla": None,
    "montserrat": None,
    "cook islands": None,
    "turks and caicos islands": None,
    "sint maarten (dutch part)": None,
    "cayman islands": None,
    "bermuda": None,
    "curaçao": None,
    "aruba": None,
    "french polynesia": None,
    "new caledonia": None,
    "faroe islands": None,
    "marshall islands": None,
}

# Non-country aggregation entries to exclude from dev-status ranking
NON_COUNTRY_ENTRIES = {
    "european union nes",
    "chinese taipei",
}

# ITC Regional classifications (based on tradebriefs.intracen.org)
ITC_REGIONS = {
    "Africa": {
        "algeria", "angola", "benin", "botswana", "burkina faso", "burundi",
        "cameroon", "cape verde", "central african republic", "chad", "comoros",
        "congo", "congo, democratic republic of the", "cote d'ivoire",
        "djibouti", "egypt", "equatorial guinea", "eritrea", "eswatini",
        "ethiopia", "gabon", "gambia", "ghana", "guinea", "guinea-bissau",
        "kenya", "lesotho", "liberia", "libya", "madagascar", "malawi",
        "mali", "mauritania", "mauritius", "morocco", "mozambique",
        "namibia", "niger", "nigeria", "rwanda", "sao tome and principe",
        "senegal", "seychelles", "sierra leone", "somalia", "south africa",
        "south sudan", "sudan", "tanzania", "togo", "tunisia", "uganda",
        "zambia", "zimbabwe",
    },
    "Asia": {
        "afghanistan", "armenia", "azerbaijan", "bahrain", "bangladesh",
        "bhutan", "brunei", "cambodia", "china", "cyprus", "georgia",
        "hong kong", "india", "indonesia", "iran", "iraq", "israel",
        "japan", "jordan", "kazakhstan", "korea", "kuwait", "kyrgyzstan",
        "lao people's democratic republic", "lebanon", "macao",
        "malaysia", "maldives", "mongolia", "myanmar", "nepal",
        "oman", "pakistan", "palestine", "philippines", "qatar",
        "russia", "saudi arabia", "singapore", "sri lanka", "syria",
        "taiwan", "tajikistan", "thailand", "timor-leste", "turkey",
        "türkiye", "turkmenistan", "united arab emirates", "uzbekistan",
        "viet nam", "vietnam", "yemen",
    },
    "Americas": {
        "antigua and barbuda", "argentina", "bahamas", "barbados",
        "belize", "bolivia", "brazil", "canada", "chile", "colombia",
        "costa rica", "cuba", "dominica", "dominican republic",
        "ecuador", "el salvador", "grenada", "guatemala", "guyana",
        "haiti", "honduras", "jamaica", "mexico", "nicaragua",
        "panama", "paraguay", "peru", "saint kitts and nevis",
        "saint lucia", "saint vincent and the grenadines",
        "suriname", "trinidad and tobago", "united states",
        "united states of america", "uruguay", "venezuela",
    },
    "Pacific": {
        "australia", "fiji", "kiribati", "marshall islands",
        "micronesia", "nauru", "new zealand", "papua new guinea",
        "samoa", "solomon islands", "tonga", "tuvalu", "vanuatu",
    },
    "Europe": {
        "albania", "belarus", "belgium", "bosnia and herzegovina",
        "bulgaria", "croatia", "czech republic", "czechia", "denmark",
        "estonia", "finland", "france", "germany", "greece", "hungary",
        "iceland", "ireland", "italy", "latvia", "lithuania",
        "luxembourg", "malta", "montenegro", "netherlands",
        "north macedonia", "norway", "poland", "portugal", "romania",
        "serbia", "slovakia", "slovenia", "spain", "sweden",
        "switzerland", "ukraine", "united kingdom",
    },
}


def load_hdi_data(path=None):
    """Load HDI data from devcountries CSV.

    Returns dict mapping country name (lowercase) -> HDI score.
    """
    import csv
    path = path or HDI_PATH
    if not os.path.exists(path):
        return {}
    result = {}
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("country", "").strip()
            hdi_str = row.get("HumanDevelopmentIndex_2023", "").strip()
            if name and hdi_str:
                try:
                    result[name.lower()] = float(hdi_str)
                except ValueError:
                    pass
    return result


def load_development_classification(path=None):
    """Load World Bank income group classification.

    Returns dict mapping country name (lowercase) -> income group.
    """
    path = path or CLASSIFICATION_PATH
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["List of economies"]
    result = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        income = ws.cell(r, 4).value
        if name and income:
            result[name.strip().lower()] = income.strip()
    wb.close()

    # Add common name variations for better matching
    aliases = {
        "united states of america": "High income",
        "usa": "High income",
        "u.s.a.": "High income",
        "u.s.": "High income",
        "uk": "High income",
        "u.k.": "High income",
        "uae": "High income",
        "u.a.e.": "High income",
        "south korea": "High income",
        "korea, republic of": "High income",
        "russia": "Upper middle income",
        "russian federation": "Upper middle income",
        "turkey": "Upper middle income",
        "türkiye": "Upper middle income",
        "democratic republic of the congo": "Low income",
        "congo, democratic republic of the": "Low income",
        "ivory coast": "Lower middle income",
        "côte d'ivoire": "Lower middle income",
        "egypt": "Lower middle income",
        "egypt, arab rep.": "Lower middle income",
        "iran": "Upper middle income",
        "iran, islamic rep.": "Upper middle income",
        "venezuela": "Upper middle income",
        "venezuela, rb": "Upper middle income",
        "viet nam": "Lower middle income",
        "vietnam": "Lower middle income",
        "bolivia": "Lower middle income",
        "bolivia (plurinational state of)": "Lower middle income",
        "tanzania": "Lower middle income",
        "united republic of tanzania": "Lower middle income",
    }
    for alias, income in aliases.items():
        if alias not in result:
            result[alias] = income

    return result


def get_development_status(name, classification, hdi_data=None):
    """Return 'Developed', 'Developing', or 'LDC' for a country name.

    Classification priority:
    1. UN LDC list (highest priority)
    2. HDI score (primary): >= 0.800 = Developed, < 0.800 = Developing
    3. World Bank income group (fallback)

    LDC classification:
    - UN LDC list (not graduated) -> LDC
    - HDI < 0.550 -> LDC (very low development)
    """
    name_lower = name.strip().lower()
    # Resolve ITC name -> HDI / World Bank name
    lookup = NAME_ALIASES.get(name_lower, name_lower)

    # Check LDC first (overrides other classifications)
    if name_lower in LDC_COUNTRIES and name_lower not in LDC_GRADUATED:
        return "LDC"

    # Manual override for countries with high HDI but still developing economies
    if name_lower in DEVELOPING_OVERRIDE:
        return "Developing"

    # Try HDI data first (most accurate)
    if hdi_data and lookup in hdi_data:
        hdi = hdi_data[lookup]
        if hdi >= HDI_DEVELOPED_THRESHOLD:
            return "Developed"
        elif hdi < 0.550:
            return "LDC"
        else:
            return "Developing"

    # Fallback to World Bank income group
    ig = classification.get(lookup, "")
    if ig == "High income":
        return "Developed"
    return "Developing"


def get_region(name):
    """Return ITC region for a country name.

    Returns one of: Africa, Asia, Americas, Pacific, Europe, or 'Other'.
    """
    name_lower = name.strip().lower()
    lookup = NAME_ALIASES.get(name_lower, name_lower)
    for region, countries in ITC_REGIONS.items():
        if lookup in countries or name_lower in countries:
            return region
    return "Other"


def _dev_fill(status):
    if status == "Developed":
        return DEV_FILL
    elif status == "LDC":
        return PatternFill(fill_type="solid", fgColor="FCE4D6")
    return DEVEL_FILL


# ---------------------------------------------------------------------------
# HTML table parser (for ITC .xls files that are actually HTML)
# ---------------------------------------------------------------------------
# _HTMLTableParser and _read_html moved to xlsx_compat.py


# ---------------------------------------------------------------------------
# Formula caching (from make_tables.py)
# ---------------------------------------------------------------------------
def _sheet_order(path):
    with zipfile.ZipFile(path) as zin:
        xml = zin.read("xl/workbook.xml").decode("utf-8", "ignore")
    return re.findall(r'<sheet[^>]*name="([^"]+)"', xml)


def _inject_cached_values(path, cache):
    """Add cached <v> results next to <f> formulas so data_only=True works."""
    if isinstance(cache, dict):
        per_sheet = {k: {r: v for r, v in m.items() if v is not None}
                     for k, m in cache.items()}
    else:
        per_sheet = {"Sheet1": {r: v for r, v in cache if v is not None}}
    if not any(per_sheet.values()):
        return
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        blobs = {n: zin.read(n) for n in names}
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    order = _sheet_order(path)
    patched = 0
    for i, sheet in enumerate(order):
        refs = per_sheet.get(sheet)
        if not refs:
            continue
        fname = f"xl/worksheets/sheet{i + 1}.xml"
        if fname not in blobs:
            continue
        root = etree.fromstring(blobs[fname])
        for cell in root.iter("{%s}c" % ns):
            ref = cell.get("r")
            if ref not in refs:
                continue
            if cell.find("{%s}f" % ns) is None:
                continue
            v = cell.find("{%s}v" % ns)
            if v is None:
                v = etree.SubElement(cell, "{%s}v" % ns)
            v.text = repr(refs[ref])
        blobs[fname] = etree.tostring(root, xml_declaration=True,
                                      encoding="UTF-8")
        patched += 1
    if not patched:
        return
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, blobs[n])


def _save_workbook(wb, path, cache):
    """Save the workbook, patch in cached formula values, and fall back to
    plain values if anything goes wrong so the pipeline never breaks."""
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(path)
    if not cache:
        return
    try:
        _inject_cached_values(path, {wb.sheetnames[0]: {r: v for r, v in cache
                                                          if v is not None}})
        wb2 = openpyxl.load_workbook(path, data_only=True)
        ws2 = wb2.active
        bad = []
        for ref, val in cache:
            if val is not None and ws2[ref].value is None:
                bad.append(ref)
        if bad:
            raise RuntimeError(f"cached values missing for {bad}")
    except Exception:
        wb3 = openpyxl.load_workbook(path)
        ws3 = wb3.active
        for ref, val in cache:
            if val is not None:
                ws3[ref] = val
        wb3.save(path)


def _finalize(ws, path, cache):
    """Convenience wrapper for _save_workbook with a single worksheet."""
    _save_workbook(ws.parent, path, cache)


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def to_float(v):
    """Convert a raw cell value to float; None for blanks / '...' / garbage."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", "")
        if v in ("", "...", "..", "-"):
            return None
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def _parse_years_from_header(header_row):
    """Extract year integers from a header row like
    ['Code', 'Service label', 'Exported value in 2020', ...]."""
    years = []
    for cell in header_row:
        m = re.search(r"(\d{4})", str(cell))
        if m:
            y = int(m.group(1))
            if 2000 <= y <= 2100:
                years.append(y)
    return years


def _header_year_indices(header_row):
    """Return list of (col_index, year) for year columns in the header."""
    out = []
    for i, cell in enumerate(header_row):
        m = re.search(r"(\d{4})", str(cell))
        if m:
            y = int(m.group(1))
            if 2000 <= y <= 2100:
                out.append((i, y))
    return out


# ---------------------------------------------------------------------------
# Country name canonicalisation (ITC ↔ UNCTAD matching)
# ---------------------------------------------------------------------------
# UNCTAD economy labels → canonical short names.  Keys are lowercase.
_UNCTAD_CANONICAL = {
    "united states": "USA",
    "united states of america": "USA",
    "czech republic": "Czechia",
    "czechia": "Czechia",
    "bolivia, plurinational state of": "Bolivia",
    "bolivia (plurinational state of)": "Bolivia",
    "hong kong, china": "Hong Kong",
    "china, hong kong sar": "Hong Kong",
    "chinese taipei": "Taiwan",
    "china, taiwan province of": "Taiwan",
    "taiwan, province of china": "Taiwan",
    "korea, republic of": "South Korea",
    "republic of korea": "South Korea",
    "iran, islamic republic of": "Iran",
    "iran (islamic republic of)": "Iran",
    "congo, democratic republic of the": "DR Congo",
    "dem. rep. of the congo": "DR Congo",
    "democratic republic of the congo": "DR Congo",
    "lao people's democratic republic": "Laos",
    "lao people's dem. rep.": "Laos",
    "macedonia, north": "North Macedonia",
    "north macedonia": "North Macedonia",
    "netherlands (kingdom of the)": "Netherlands",
    "netherlands antilles": "Netherlands Antilles",
    "kyrgyzstan": "Kyrgyzstan",
    "cote d'ivoire": "Cote d'Ivoire",
    "côte d'ivoire": "Cote d'Ivoire",
    "ivory coast": "Cote d'Ivoire",
    "myanmar": "Myanmar",
    "myanmar (burma)": "Myanmar",
    "swaziland": "Eswatini",
    "eswatini": "Eswatini",
    "east timor": "Timor-Leste",
    "timor-leste": "Timor-Leste",
    "congo": "Congo",
    "republic of the congo": "Congo",
    "democratic republic of germany": "Germany",
    "federal republic of germany": "Germany",
    "kosovo": "Kosovo",
}


def _canonical_country(name):
    """Map a country name to a canonical short form for ITC ↔ UNCTAD matching."""
    s = name.lower().strip()
    nfkd = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    if s in _UNCTAD_CANONICAL:
        return _UNCTAD_CANONICAL[s].lower()
    from country_names import SHORT_NAMES
    if s in SHORT_NAMES:
        return SHORT_NAMES[s].lower()
    return s


# ---------------------------------------------------------------------------
# File detection
# ---------------------------------------------------------------------------
def find_service_files(excel_dir):
    """Locate the six required raw service source files by filename keywords.

    Returns ``{role: [path, ...]}`` mapping each known role to *every* file
    in *excel_dir* whose name matches that role's keywords (several candidates
    can exist when both the raw ``.xls`` download and a converted ``.xlsx``
    copy of the same query are present, or when an upload was renamed with a
    ``__1`` suffix).  Unrecognised files are simply ignored.

    The ``kenya_commercialized`` file is optional; when present its data is
    loaded but the balance is always computed from export − import totals.
    The UNCTADstat *annual* services file is optional; when present its data
    fills Kenya's missing 2024 category detail, extends Tables 1/2 & the world
    pie to the latest UNCTAD year using the true World total, and supplies the
    per-category world sums for the RCA/concentration/diversification tables.
    """
    found = collections.defaultdict(list)
    for fname in sorted(os.listdir(excel_dir)):
        low = fname.lower()
        if any(kw in low for kw in UNCTAD_R_KEYWORDS):
            found["unctad_r"].append(os.path.join(excel_dir, fname))
            continue
        if any(kw in low for kw in UNCTAD_ANNUAL_KEYWORDS):
            found["unctad_annual"].append(os.path.join(excel_dir, fname))
            continue
        if not is_spreadsheet(fname):
            continue
        if "exported_services_for" in low or "list_of_exported_services" in low:
            role = "exported_services"
        elif "exporters_for" in low or "list_of_exporters_for" in low:
            role = "exporters"
        elif "imported_services_for" in low or "list_of_imported_services" in low:
            role = "imported_services"
        elif "importers_for" in low or "list_of_importers_for" in low:
            role = "importers"
        elif "services_exported_by" in low:
            role = "kenya_exports"
        elif "services_imported_by" in low:
            role = "kenya_imports"
        elif "services_commercialized_by" in low:
            role = "kenya_commercialized"
        else:
            continue
        found[role].append(os.path.join(excel_dir, fname))

    required = ("exported_services", "exporters",
                 "imported_services", "importers",
                 "kenya_exports", "kenya_imports")
    missing = [k for k in required if not found[k]]
    if missing:
        expected = {
            "exported_services": "List_of_exported_services_for_the_selected_service",
            "exporters": "List_of_exporters_for_the_selected_service",
            "imported_services": "List_of_imported_services_for_the_selected_service",
            "importers": "List_of_importers_for_the_selected_service",
            "kenya_exports": "List_of_services_exported_by_Kenya",
            "kenya_imports": "List_of_services_imported_by_Kenya",
        }
        missing_detail = []
        for k in missing:
            missing_detail.append(f"  - {k}: expected filename containing '{expected[k]}'")
        found_summary = ", ".join(
            f"{k}=[{', '.join(os.path.basename(p) for p in v)}]"
            for k, v in sorted(found.items()))
        sys.exit(
            "[ERROR] Missing required service source files in '%s'.\n"
            "Found %d candidate file(s): %s\n"
            "Missing %d required file(s):\n%s"
            % (excel_dir, sum(len(v) for v in found.values()),
               found_summary or "(none)",
               len(missing), "\n".join(missing_detail)))
    return dict(found)


def _load_first(found, role, loader):
    """Try each candidate file for *role* until one parses successfully.

    Returns the loader's result, ignoring any unusable candidates for the
    same role, or ``None`` when the role has no candidates (optional roles).

    Raises the last parsing error when *role* has candidates but none of them
    could be read (required roles) — the caller decides whether a failure is
    fatal.
    """
    candidates = found.get(role) or []
    last_err = None
    for path in candidates:
        try:
            return loader(path)
        except Exception as exc:  # noqa: BLE001 - resilience by design
            last_err = exc
            print(f"Warning: skipping unreadable '{os.path.basename(path)}' "
                  f"for role '{role}': {exc}")
    if candidates and last_err is not None:
        raise last_err
    return None


# ---------------------------------------------------------------------------
# Source parsing
# ---------------------------------------------------------------------------
def parse_service_html_table(path):
    """Parse an ITC services file.

    Handles .xlsx, .xlsm, .xls (HTML-wrapped or binary), .xlsb, and .csv.

    Returns (data_table, years) where data_table is the data
    as a list of row-lists, and years is the list of year ints
    found in the header.
    """
    wb, used_path = convert_to_xlsx(path)
    ws = wb.active
    data_table = []
    for row in ws.iter_rows(values_only=True):
        data_table.append([str(c) if c is not None else "" for c in row])
    wb.close()

    if len(data_table) < 2:
        raise ValueError(f"Data table too small in {path}")
    header = data_table[0]
    years = _parse_years_from_header(header)
    return data_table, years


def parse_exported_services(path):
    """Global service exports by category. Returns (total, items).

    total = {'label': ..., 'vals': [...]}  (the 'All services' row)
    items = [{'code': ..., 'label': ..., 'vals': [...]}, ...]
    """
    table, years = parse_service_html_table(path)
    ycols = _header_year_indices(table[0])
    total = None
    items = []
    for row in table[1:]:
        if len(row) < 3:
            continue
        code = str(row[0]).strip()
        label = str(row[1]).strip()
        vals = [to_float(row[ci]) if ci < len(row) else None for ci, _ in ycols]
        if code.upper() == "S" or label.lower() == "all services":
            total = {"code": code, "label": label, "vals": vals}
        else:
            items.append({"code": code, "label": label, "vals": vals})
    return total, items, years


def parse_exporters(path):
    """Global exporters by country. Returns (total_sum, items).

    total_sum = {'label': 'World', 'vals': [...]}  (computed sum)
    items = [{'label': ..., 'vals': [...]}, ...]
    """
    table, years = parse_service_html_table(path)
    ycols = _header_year_indices(table[0])
    items = []
    for row in table[1:]:
        if len(row) < 2:
            continue
        label = str(row[0]).strip()
        vals = [to_float(row[ci]) if ci < len(row) else None for ci, _ in ycols]
        if not label:
            continue
        items.append({"label": label, "vals": vals})
    # Compute world total
    n_years = len(ycols)
    world_vals = []
    for k in range(n_years):
        s = 0.0
        for it in items:
            v = it["vals"][k] if k < len(it["vals"]) else None
            s += v if v is not None else 0.0
        world_vals.append(s)
    total = {"label": "World", "vals": world_vals}
    return total, items, years


def parse_importers(path):
    """Global importers by country. Returns (total_sum, items)."""
    return parse_exporters(path)  # identical structure


def parse_imported_services(path):
    """Global service imports by category. Returns (total, items, years)."""
    return parse_exported_services(path)  # identical structure


def parse_kenya_services(path):
    """Kenya services exports or imports by category.

    Returns (total, items, years).
    total = {'code': ..., 'label': ..., 'vals': [...]}
    items = [{'code': ..., 'label': ..., 'vals': [...]}, ...]
    """
    table, years = parse_service_html_table(path)
    ycols = _header_year_indices(table[0])
    total = None
    items = []
    for row in table[1:]:
        if len(row) < 3:
            continue
        code = str(row[0]).strip()
        label = str(row[1]).strip()
        vals = [to_float(row[ci]) if ci < len(row) else None for ci, _ in ycols]
        if code.upper() == "S" or label.lower() == "all services":
            total = {"code": code, "label": label, "vals": vals}
        else:
            items.append({"code": code, "label": label, "vals": vals})
    return total, items, years


def parse_kenya_commercialized(path):
    """Kenya services balance by category.

    Returns (total, items, years) where items have 'balance', 'export_val',
    'import_val' keys alongside 'vals' (balance values).
    """
    table, years = parse_service_html_table(path)
    # Header: Code, Service label, Balance in value 2020, ..., Exported Value 2024, Imported Value 2024
    header = table[0]
    ycols = _header_year_indices(header)
    # Find export/import 2024 columns
    export_col = None
    import_col = None
    for i, cell in enumerate(header):
        s = str(cell).lower()
        if "exported value" in s:
            export_col = i
        elif "imported value" in s:
            import_col = i

    total = None
    items = []
    for row in table[1:]:
        if len(row) < 3:
            continue
        code = str(row[0]).strip()
        label = str(row[1]).strip()
        vals = [to_float(row[ci]) if ci < len(row) else None for ci, _ in ycols]
        export_val = to_float(row[export_col]) if export_col and export_col < len(row) else None
        import_val = to_float(row[import_col]) if import_col and import_col < len(row) else None
        entry = {"code": code, "label": label, "vals": vals,
                 "export_val": export_val, "import_val": import_val}
        if code.upper() == "S" or label.lower() == "all services":
            total = entry
        else:
            items.append(entry)
    return total, items, years


def _truncate_to_category_year(items, total, years):
    """Drop trailing year columns that have no real per-category data.

    Unpublished-or-partial recent years would leave only the all-services
    total populated (with every category cell blank), which reads as a column
    with no breakdown.  To present an honest, internally-consistent table we
    trim the axis back to the latest year where at least two category cells
    carry a value, so every shown column has both a total and a category
    breakdown.

    Mutates ``items``/``total`` in place (their ``vals`` lists are trimmed).
    Returns the trimmed ``years`` list.
    """
    if not items:
        return years
    keep = len(years)
    while keep > 0:
        k = keep - 1
        n_cells = sum(
            1 for it in items
            if it.get("vals") and k < len(it["vals"]) and it["vals"][k] is not None
        )
        if n_cells >= 2:
            break
        keep -= 1
    keep = max(keep, 1)
    if keep < len(years):
        for it in items:
            if it.get("vals"):
                it["vals"] = it["vals"][:keep]
        if total and total.get("vals"):
            total["vals"] = total["vals"][:keep]
        years = list(years[:keep])
    return years


def parse_unctad_annual_kenya(path):
    """Parse a fetched UNCTADstat annual services-by-category CSV (Kenya).

    Generated by ``fetch_unctad_tradeserv.py`` (see its ``--out`` / the file
    dropped into *excel_dir*).  Columns are:

        Economy_Label, Flow_Label, Category_Code, Category_Label, Year,
        Millions_of_US_at_current_prices_Value,
        US_at_current_prices_Footnote, US_at_current_prices_MissingValue

    Per the user directive we never estimate a category breakdown: rows whose
    value cell is blank or flagged ``Not publishable`` are dropped, so only
    genuinely published values enter the result.

    Returns ``{"Exports": {code: {year: millions}}, "Imports": {...}}`` where
    ``code`` is the UNCTAD EBOPS category code (``S`` total, ``SC`` Transport,
    ``SD`` Travel, …) and values are in millions of US$.
    """
    f = gzip.open(path, "rt", encoding="utf-8") if str(path).lower().endswith(".gz") \
        else open(path, newline="", encoding="utf-8")
    out = {"Exports": {}, "Imports": {}}
    with f as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            flow = (row.get("Flow_Label") or "").strip()
            if flow not in out:
                continue
            code = (row.get("Category_Code") or "").strip()
            year_s = (row.get("Year") or "").strip()
            if not code or not year_s.isdigit():
                continue
            year = int(year_s)
            raw = (row.get("Millions_of_US_at_current_prices_Value") or "").strip()
            missing = (row.get("US_at_current_prices_MissingValue") or "").strip()
            if not raw or missing:
                continue
            try:
                val = float(raw.replace(",", ""))
            except ValueError:
                continue
            out[flow].setdefault(code, {})[year] = val
    return out


def parse_unctad_annual(path):
    """Parse the fetched UNCTADstat *annual* **all-economy** services file.

    Produced by ``fetch_unctad_tradeserv.py --all-economies`` (→ an
    ``UNCTAD_*_tradeserv_annual_all.csv.gz`` in *excel_dir*).  It carries every
    economy (real 3-digit countries plus the 4-digit aggregate/region rows) by
    EBOPS category and year, in millions of USD.  Columns are:

        Economy_Label, Economy_Code, Flow_Label, Category_Code, Category_Label,
        Year, Millions_of_US_at_current_prices_Value,
        US_at_current_prices_Footnote, US_at_current_prices_MissingValue

    A real economy has a 3-digit ``Economy_Code`` (e.g. Kenya = ``404``);
    aggregate rows carry 4-digit codes (``0000`` = World, regions, development
    groupings).  Rows with a blank value or a ``Not publishable`` note are
    dropped (we never estimate a category).

    Returns::

        {
          "kenya":     {"Exports": {cat: {year: Mn}}, "Imports": {...}},  # Kenya
          "world":     {"Exports": {cat: {year: Mn}}, "Imports": {...}},  # sum of 3-digit economies per category
          "country_s": {"Exports": {year: {eco_lower: Mn}}, "Imports": {...}},  # S total, 3-digit only
          "world_s":   {"Exports": {year: Mn}, "Imports": {...}},         # TRUE published World (0000) total
        }

    All values are in millions of USD.
    """
    f = gzip.open(path, "rt", encoding="utf-8") if str(path).lower().endswith(".gz") \
        else open(path, newline="", encoding="utf-8")

    def _mk():
        return {"Exports": collections.defaultdict(
            lambda: collections.defaultdict(dict)),
            "Imports": collections.defaultdict(
                lambda: collections.defaultdict(dict))}

    by_flow = _mk()          # flow -> cat -> {year -> {eco -> value}}
    world_s = {"Exports": {}, "Imports": {}}
    country_s = {"Exports": {}, "Imports": {}}
    kenya = _mk()
    eco_is_country = {}

    with f as fh:
        for row in csv.DictReader(fh):
            flow = (row.get("Flow_Label") or "").strip()
            if flow not in ("Exports", "Imports"):
                continue
            code = (row.get("Category_Code") or "").strip()
            year_s = (row.get("Year") or "").strip()
            if not code or not year_s.isdigit():
                continue
            year = int(year_s)
            raw = (row.get("Millions_of_US_at_current_prices_Value") or "").strip()
            missing = (row.get("US_at_current_prices_MissingValue") or "").strip()
            if not raw or missing:
                continue
            try:
                val = float(raw.replace(",", ""))
            except ValueError:
                continue
            eco = (row.get("Economy_Label") or "").strip()
            eco_code = (row.get("Economy_Code") or "").strip()
            if not eco:
                continue
            is_country = eco_code.isdigit() and len(eco_code) == 3
            eco_is_country[eco.lower()] = is_country
            by_flow[flow][code][year].setdefault(eco, 0.0)
            by_flow[flow][code][year][eco] += val
            if eco.lower() == "kenya":
                kenya[flow][code].setdefault(year, 0.0)
                kenya[flow][code][year] += val

    for flow in ("Exports", "Imports"):
        # TRUE published World total (the 0000 aggregate row labelled 'World').
        if "S" in by_flow[flow]:
            for year, byeco in by_flow[flow]["S"].items():
                w = byeco.get("World") or byeco.get("world")
                if w:
                    world_s[flow][year] = w
        # country_s: per-economy S total for 3-digit economies only.
        if "S" in by_flow[flow]:
            for year, byeco in by_flow[flow]["S"].items():
                for eco, val in byeco.items():
                    if eco_is_country.get(eco.lower(), False):
                        country_s[flow].setdefault(year, {})[eco.lower()] = val

    # world: per-category sum over 3-digit economies only.
    world_out = {"Exports": {}, "Imports": {}}
    for flow in ("Exports", "Imports"):
        for code, byyear in by_flow[flow].items():
            wc = {}
            for year, byeco in byyear.items():
                s = sum(v for eco, v in byeco.items()
                        if eco_is_country.get(eco.lower(), False))
                if s:
                    wc[year] = s
            if wc:
                world_out[flow][code] = wc

    kenya_out = {"Exports": {c: dict(yy) for c, yy in kenya["Exports"].items()},
                 "Imports": {c: dict(yy) for c, yy in kenya["Imports"].items()}}
    return {
        "kenya": kenya_out,
        "world": world_out,
        "country_s": country_s,
        "world_s": world_s,
    }


# UNCTAD economy labels in the R-script output that are REGIONAL/AGGREGATE
# groupings rather than real countries.  Every other label in that file is a
# genuine economy (Kenya, United States, Turkiye, ...) whose per-category (S)
# total is a true published figure.  These are excluded from ``country_s`` and
# from the per-category ``world`` sums to avoid double counting; the real
# published World aggregate is read separately from the ``World`` label.
_UNCTAD_R_AGGREGATES = frozenset({
    "world", "africa", "americas", "asia", "europe", "oceania",
    "individual economies", "developing economies", "developed economies",
    "ldcs", "land locked developing countries",
    "small island developing states",
    "sids", "other territories n.e.c.", "memo",
})


def parse_unctad_services_r(path):
    """Parse the R-script UNCTAD CSV (``unctad_services_<years>.csv``).

    The R download produces a long-format CSV with columns::

        Flow_Label, Category_Label, Economy_Label, Year,
        Millions_of_US_at_current_prices_Value, Category_Type, Category_Code

    Unlike the ``parse_unctad_annual`` input there is **no** numeric
    ``Economy/Code`` column, so real economies are recognised by *label*: any
    economy not in ``_UNCTAD_R_AGGREGATES`` is treated as a country.  The extra
    aggregate categories (``SPX1`` Other services, ``SPX4`` Goods-related,
    ``SOX`` Commercial services, ``Category_Type`` TOTAL vs DETAIL) are kept but
    only ``S`` contributes to ``country_s``/``world_s``.  ``NA`` values are
    treated as missing (never estimated).

    Returns::

        {
          "kenya":     {"Exports": {cat: {year: Mn}}, "Imports": {...}},
          "world":     {"Exports": {cat: {year: Mn}}, "Imports": {...}},
          "country_s": {"Exports": {year: {eco_lower: Mn}}, "Imports": {...}},
          "world_s":   {"Exports": {year: Mn}, "Imports": {...}},
        }

    All values are already in millions of USD.
    """
    def _mk():
        flow = lambda: collections.defaultdict(
            lambda: collections.defaultdict(dict))
        return {"Exports": flow(), "Imports": flow()}

    by_flow = _mk()
    world_s = {"Exports": {}, "Imports": {}}
    country_s = {"Exports": {}, "Imports": {}}
    kenya = _mk()
    is_country = {}

    opener = gzip.open(path, "rt", encoding="utf-8") \
        if str(path).lower().endswith(".gz") \
        else open(path, newline="", encoding="utf-8")

    with opener as fh:
        for row in csv.DictReader(fh):
            flow = (row.get("Flow_Label") or "").strip()
            if flow not in ("Exports", "Imports"):
                continue
            code = (row.get("Category_Code") or "").strip()
            if not code:
                continue
            year_s = (row.get("Year") or "").strip()
            if not year_s.isdigit():
                continue
            year = int(year_s)
            raw = (row.get("Millions_of_US_at_current_prices_Value")
                   or "").strip()
            # "NA" / blank / annotation => missing, never an estimate.
            if not raw or raw.upper() == "NA" or raw.upper() in ("..", "-", "NULL"):
                continue
            try:
                val = float(raw.replace(",", ""))
            except ValueError:
                continue
            eco = (row.get("Economy_Label") or "").strip()
            if not eco:
                continue
            eco_key = eco.lower()
            is_country[eco_key] = eco_key not in _UNCTAD_R_AGGREGATES
            by_flow[flow][code][year].setdefault(eco, 0.0)
            by_flow[flow][code][year][eco] += val
            if eco_key == "kenya":
                kenya[flow][code].setdefault(year, 0.0)
                kenya[flow][code][year] += val

    for flow in ("Exports", "Imports"):
        if "S" in by_flow[flow]:
            for year, byeco in by_flow[flow]["S"].items():
                world_val = byeco.get("World") or byeco.get("world")
                if world_val:
                    world_s[flow][year] = world_val
        if "S" in by_flow[flow]:
            for year, byeco in by_flow[flow]["S"].items():
                for eco, val in byeco.items():
                    if is_country.get(eco.lower(), False):
                        country_s[flow].setdefault(year, {})[eco.lower()] = val

    world_out = {"Exports": {}, "Imports": {}}
    for flow in ("Exports", "Imports"):
        for code, byyear in by_flow[flow].items():
            wc = {}
            for year, byeco in byyear.items():
                s = sum(v for eco, v in byeco.items()
                        if is_country.get(eco.lower(), False))
                if s:
                    wc[year] = s
            if wc:
                world_out[flow][code] = wc

    kenya_out = {"Exports": {c: dict(yy) for c, yy in kenya["Exports"].items()},
                 "Imports": {c: dict(yy) for c, yy in kenya["Imports"].items()}}
    return {
        "kenya": kenya_out,
        "world": world_out,
        "country_s": country_s,
        "world_s": world_s,
    }


def _gage_kenya_from_annual_gz(files, un_annual):
    """Fill Kenya per-category gaps in ``un_annual`` from the all-economy gz.

    The R-export CSV (``parse_unctad_services_r``) is the *primary* UNCTAD source
    and is preferred wherever it has a value, but its category list omits some
    EBOPS codes (notably ``SA`` Manufacturing and ``SB`` Maintenance/repair) that
    the fetched all-economy file (``parse_unctad_annual``) does carry.  For
    Kenya's Tables 3/4 the same country-specific detail can legitimately be taken
    from the gz, so we merge in any category/year that the R-file is missing -
    a generalized gap-fill, capped to the same notion of "complete real data"
    (no ``Not publishable`` rows, no estimates).

    ``un_annual`` is mutated in place (its ``kenya`` map gets any missing
    category cells filled) and returned.
    """
    if un_annual is None or not un_annual.get("kenya"):
        return un_annual
    # Only merge when a usable all-economy gz is also available.
    cands = [c for c in files.get("unctad_annual", [])
             if os.path.basename(c).lower().endswith("_all.csv.gz")] \
        or list(files.get("unctad_annual", []))
    gz_out = None
    for cand in sorted(cands):
        try:
            gz_out = parse_unctad_annual(cand)
            if gz_out and gz_out.get("kenya"):
                break
        except Exception:  # noqa: BLE001 - try next candidate
            gz_out = None
    if not gz_out or not gz_out.get("kenya"):
        return un_annual

    for flow in ("Exports", "Imports"):
        gz_flow = gz_out["kenya"].get(flow, {})
        if not gz_flow:
            continue
        r_flow = un_annual["kenya"].setdefault(flow, {})
        for cat, byyear in gz_flow.items():
            for year, val in byyear.items():
                # R-file wins wherever it already has the cell.
                if val is None:
                    continue
                r_cat = r_flow.setdefault(cat, {})
                if year not in r_cat or r_cat[year] is None:
                    r_cat[year] = val
    return un_annual


def _copy_unctad_source_files(files, out_dir):
    """Copy the upstream UNCTAD R-export file(s) into *out_dir*.

    The services deliverable should contain both the ITC tables *and* the raw
    UNCTAD R-export CSV that was analysed, so a zip of that folder bundles it
    alongside the tables.  Only the ``*.csv`` is mirrored: the analyst keeps the
    companion ``*.xlsx`` locally (it is the source and need not be shipped).
    """
    try:
        for src in files.get("unctad_r", []):
            base = os.path.basename(src)
            if base.lower().endswith(".csv"):
                shutil.copy2(src, os.path.join(out_dir, base))
    except Exception as exc:  # noqa: BLE001 - additive, never break the build
        print(f"Warning: could not copy UNCTAD source files into output: {exc}")


def extend_kenya_categories_from_unctad(items, total, years, un_kan, flow,
                                        max_year=None):
    """Extend ITC Kenya category tables (3/4) with UNCTAD annual category detail.

    The ITC Kenya files only break category values out through 2023; the final
    column (e.g. 2024) carries the aggregate (``S``) total with every category
    cell blank.  The fetched UNCTAD *annual* dataset has complete per-category
    detail for that year, so we fill the blank category cells (and keep the
    total in sync with UNCTAD) using genuine values - never an estimated share.
    2025 is excluded: it exists only as an estimated aggregate with no category
    breakdown.

    ``items``/``total``/``years`` are the ITC Kenya series (Unit: USD millions,
    after the caller's ÷1e3 conversion; years already include the latest column
    such as 2024).  ``un_kan`` is the ``parse_unctad_annual_kenya`` output for
    one flow.  ``max_year`` caps which columns may be filled.  ITC category
    values for earlier, fully-reported years are never overwritten; the
    aggregate-only later column is filled from UNCTAD and its grand total is set
    to the matching UNCTAD total so the column reconciles internally.

    Returns ``(items, total, years, note)``.
    """
    if not un_kan or "S" not in un_kan:
        return items, total, years, None
    items = [dict(it) for it in items]
    total = dict(total)

    # Fill blank category cells within the existing year axis using real UNCTAD
    # annual values (capped at max_year).  Existing ITC values for earlier years
    # (e.g. 2020-2023) already have category detail and are never overwritten;
    # only the later aggregate-only column (e.g. 2024) gets filled.
    # Then set each column that received UNCTAD categories to the UNCTAD grand
    # total too, so the category breakdown and the total reconcile within that
    # year instead of mixing two sources.
    filled = 0
    for idx, y in enumerate(years):
        if max_year is not None and y > max_year:
            continue
        year_filled = 0
        for it in items:
            itc_code = str(it.get("code"))
            un_cat = next((c for c, code in UNCTAD_CATEGORY_TO_ITC_CODE.items()
                           if code == itc_code), None)
            if not un_cat or un_cat not in un_kan:
                continue
            if it["vals"][idx] is None and y in un_kan[un_cat]:
                it["vals"][idx] = un_kan[un_cat][y]
                filled += 1
                year_filled += 1
        if year_filled and total and y in un_kan["S"]:
            total["vals"][idx] = un_kan["S"][y]

    # Append any later complete category years (real data only), capped to
    # max_year, so genuinely published newer years can still be added.
    un_years = [y for y in sorted(un_kan.get("S", {}))
                if y > (years[-1] if years else 0)]
    if max_year is not None:
        un_years = [y for y in un_years if y <= max_year]
    if un_years:
        new_years = list(years) + un_years

        def _pad(vals, n):
            vals = list(vals)
            return (vals + [None] * (n - len(vals)))[:n]

        n = len(new_years)
        year_idx = {y: i for i, y in enumerate(new_years)}
        for it in items:
            it["vals"] = _pad(it.get("vals", []), n)
        if total:
            total["vals"] = _pad(total.get("vals", []), n)
        exists = set(years)
        for it in items:
            itc_code = str(it.get("code"))
            un_cat = next((c for c, code in UNCTAD_CATEGORY_TO_ITC_CODE.items()
                           if code == itc_code), None)
            if not un_cat or un_cat not in un_kan:
                continue
            for y in un_years:
                idx = year_idx.get(y)
                if idx is None or y in exists:
                    continue
                if it["vals"][idx] is None and y in un_kan[un_cat]:
                    it["vals"][idx] = un_kan[un_cat][y]
                    filled += 1
        if total:
            for y in un_years:
                idx = year_idx.get(y)
                if idx is not None and y not in exists:
                    if total["vals"][idx] is None and y in un_kan["S"]:
                        total["vals"][idx] = un_kan["S"][y]
        years = new_years

    note = ("UNCTAD annual integration: extended Kenya %s categories with %d "
            "value(s)." % (flow, filled)) if filled else None
    return items, total, years, note


def extend_world_categories_from_unctad(world_items, world_total, itc_years,
                                        un_world, flow, max_year=None):
    """Extend ITC global service-category series with UNCTAD world sums.

    Appends UNCTAD years (e.g. 2025) to the year axis and fills the world
    series for the ITC codes that map cleanly to UNCTAD EBOPS categories so
    that Kenya's RCA can be computed for the extended years.

    ``max_year``: if given, only years ≤ this value are appended (used to
    keep the world axis aligned with Kenya's coverage).

    Returns ``(world_items, world_total, years, note)``.
    """
    if "S" not in un_world:
        return world_items, world_total, itc_years, None
    max_itc = itc_years[-1] if itc_years else 0
    un_years = [y for y in sorted(un_world["S"]) if y > max_itc]
    if max_year is not None:
        un_years = [y for y in un_years if y <= max_year]
    if not un_years:
        return world_items, world_total, itc_years, None
    new_years = list(itc_years) + un_years

    world_items = [dict(it) for it in world_items]
    world_total = dict(world_total)

    def _pad(vals, n):
        vals = list(vals)
        return (vals + [None] * (n - len(vals)))[:n]

    n = len(new_years)
    year_idx = {y: i for i, y in enumerate(new_years)}
    for it in world_items:
        it["vals"] = _pad(it.get("vals", []), n)
    world_total["vals"] = _pad(world_total.get("vals", []), n)

    itc_set = set(itc_years)
    filled = 0
    for it in world_items:
        itc_code = str(it.get("code"))
        un_cat = next((c for c, code in UNCTAD_CATEGORY_TO_ITC_CODE.items()
                       if code == itc_code), None)
        if not un_cat or un_cat not in un_world:
            continue
        for y in un_years:
            idx = year_idx.get(y)
            if idx is None or y in itc_set:
                continue
            if it["vals"][idx] is None and y in un_world[un_cat]:
                it["vals"][idx] = un_world[un_cat][y] / 1e3
                filled += 1
    for y in un_years:
        idx = year_idx.get(y)
        if idx is not None and world_total["vals"][idx] is None and y in un_world["S"]:
            world_total["vals"][idx] = un_world["S"][y] / 1e3

    note = ("UNCTAD integration: extended world %s categories with %d value(s) "
            "for %s." % (flow, filled, ", ".join(str(y) for y in un_years)))
    return world_items, world_total, new_years, note


def extend_country_rankings_from_unctad(items, total, itc_years,
                                         un_country_s, flow, max_year=None,
                                         world_s=None):
    """Extend ITC country-ranking series (exporters or importers) with UNCTAD data.

    ``items``: list of dicts with ``label`` and ``vals`` (already in USD
    billions, after the caller's /1e6 conversion).
    ``total``: world total dict with ``label`` and ``vals``.
    ``itc_years``: list of ITC year ints (e.g. [2020, …, 2024]).
    ``un_country_s``: UNCTAD ``country_s[flow]`` = {year: {eco_lower: millions}}.
    ``flow``: ``"Exports"`` or ``"Imports"`` (for UNCTAD lookup).
    ``max_year``: if given, only years ≤ this value are appended (used to keep
    the country axis aligned with Kenya's coverage).
    ``world_s``: optional ``{year: millions}`` of the *published* UNCTAD World
    total, used for the World total row when available (preferred over a
    subset sum of matched countries).

    Appends UNCTAD years (beyond the ITC max year) as new columns.  Values
    are converted from UNCTAD millions to USD billions (÷1e3).  Existing ITC
    values are never modified.

    Returns ``(items, total, years, note)``.
    """
    if not un_country_s or not itc_years:
        return items, total, itc_years, None
    max_itc = max(itc_years)
    un_years = sorted(y for y in un_country_s if y > max_itc)
    if max_year is not None:
        un_years = [y for y in un_years if y <= max_year]
    if not un_years:
        return items, total, itc_years, None
    new_years = list(itc_years) + un_years

    def _pad(vals, n):
        vals = list(vals)
        return (vals + [None] * (n - len(vals)))[:n]

    n = len(new_years)
    items = [dict(it) for it in items]
    for it in items:
        it["vals"] = _pad(it.get("vals", []), n)
    total = dict(total)
    total["vals"] = _pad(total.get("vals", []), n)

    year_idx = {y: i for i, y in enumerate(new_years)}
    itc_set = set(itc_years)
    filled_items = 0
    filled_world = 0

    # Build lookup: canonical name -> UNCTAD millions by year
    un_lookup = {}
    for y in un_years:
        byeco = un_country_s.get(y, {})
        for eco_label, val in byeco.items():
            canon = _canonical_country(eco_label)
            un_lookup.setdefault(canon, {})[y] = val

    # Map ITC items to canonical names
    itc_canon = {}
    for it in items:
        itc_canon[id(it)] = _canonical_country(it.get("label", ""))

    for it in items:
        canon = itc_canon[id(it)]
        un_years_for_eco = un_lookup.get(canon, {})
        for y in un_years:
            idx = year_idx.get(y)
            if idx is None or y in itc_set:
                continue
            if it["vals"][idx] is None and y in un_years_for_eco:
                it["vals"][idx] = un_years_for_eco[y] / 1e3
                filled_items += 1

    # World total from UNCTAD.  The CSV carries aggregate/region rows
    # (e.g. "Europe", "Asia", "Northern America") in addition to real
    # countries, so we cannot simply sum every row (that would double-count).
    # Prefer the published "World" aggregate (``world_s``); otherwise fall back
    # to summing only the rows that matched a real ITC country.
    itc_ecos = set(itc_canon.values())
    for y in un_years:
        idx = year_idx.get(y)
        if idx is None or total["vals"][idx] is not None:
            continue
        world_val = world_s.get(y) if world_s else None
        if world_val is None:
            byeco = un_country_s.get(y, {})
            if not byeco:
                continue
            world_val = byeco.get("world")
        if world_val is None:
            byeco = un_country_s.get(y, {})
            world_val = sum(v for eco, v in byeco.items()
                            if _canonical_country(eco) in itc_ecos)
        if world_val:
            total["vals"][idx] = world_val / 1e3
            filled_world += 1

    if filled_items or filled_world:
        note = ("UNCTAD country rankings: extended %s series with %d country "
                "value(s) and %d world value(s) for %s."
                % (flow.lower(), filled_items, filled_world,
                   ", ".join(str(y) for y in un_years)))
    else:
        note = None
    return items, total, new_years, note


# ---------------------------------------------------------------------------
# Ranking / aggregation
# ---------------------------------------------------------------------------
def rank_and_top(items, top_n):
    """Sort by latest-year value (desc), assign ranks, return top N."""
    items.sort(
        key=lambda it: (it["vals"][-1] if it["vals"] and it["vals"][-1] is not None
                        else float("-inf")),
        reverse=True,
    )
    for i, it in enumerate(items, 1):
        it["rank"] = i
    return items[:top_n]


def calc_growth_rates(items, years):
    """Calculate year-over-year and latest annual growth rates for each item."""
    if len(years) < 2:
        for it in items:
            it["growth"] = None
        return
    n = len(years)
    for it in items:
        vals = it.get("vals", [])
        if vals and len(vals) >= 2 and vals[-1] is not None and vals[-2] is not None and vals[-2] != 0:
            it["growth"] = (vals[-1] - vals[-2]) / abs(vals[-2])
        else:
            it["growth"] = None
    # Compute 5-year CAGR
    for it in items:
        cagr = calc_cagr(it.get("vals", []), years)
        it["cagr"] = cagr


def calc_cagr(vals, years):
    """Compute Compound Annual Growth Rate from a value series.

    CAGR = (end/start)^(1/n) - 1 where n is the number of year intervals.
    Returns None if computation is not possible.
    """
    if not vals or not years or len(vals) < 2:
        return None
    # Find first and last non-None values
    start_val = None
    end_val = None
    for v in vals:
        if v is not None and v > 0:
            if start_val is None:
                start_val = v
            end_val = v
    if start_val is None or end_val is None or start_val == 0:
        return None
    n_years = years[-1] - years[0]
    if n_years <= 0:
        return None
    return (end_val / start_val) ** (1.0 / n_years) - 1


def cagr_header_label(years):
    """Return the CAGR column header accurately reflecting the actual period.

    ``calc_cagr`` exponentiates over the true span ``years[-1] - years[0]``,
    which is not always 5 years (e.g. Kenya tables capped at 2020-2023 span
    3).  Labelling every column "5Y CAGR" is then a misnomer, so the header is
    derived from the real span: "3Y CAGR %", "5Y CAGR %", etc.
    """
    if not years or len(years) < 2:
        return "CAGR %"
    span = years[-1] - years[0]
    return f"{span}Y CAGR %" if span >= 1 else "CAGR %"


def calc_growth_stability(vals):
    """Compute coefficient of variation of year-over-year growth rates.

    Lower value = more stable growth. Returns None if insufficient data.
    """
    if not vals or len(vals) < 3:
        return None
    growths = []
    for i in range(1, len(vals)):
        if vals[i] is not None and vals[i-1] is not None and vals[i-1] > 0:
            growths.append((vals[i] - vals[i-1]) / vals[i-1])
    if len(growths) < 2:
        return None
    mean_g = sum(growths) / len(growths)
    if mean_g == 0:
        return None
    var = sum((g - mean_g) ** 2 for g in growths) / len(growths)
    std = var ** 0.5
    return abs(std / mean_g) if mean_g != 0 else None


def rank_by_dev_status(items, years, classification, top_n=10):
    """Split items by development status, rank within each group, return top N per group.

    Returns (dev_items, devel_items, ldc_items) where each is a list sorted by latest-year value.
    """
    hdi_data = load_hdi_data()
    # Filter out non-country aggregation entries
    items = [it for it in items if it.get("label", "").strip().lower()
             not in NON_COUNTRY_ENTRIES]
    for it in items:
        it["dev_status"] = get_development_status(it.get("label", ""), classification, hdi_data)
        it["region"] = get_region(it.get("label", ""))

    calc_growth_rates(items, years)

    dev_items = [it for it in items if it["dev_status"] == "Developed"]
    devel_items = [it for it in items if it["dev_status"] == "Developing"]
    ldc_items = [it for it in items if it["dev_status"] == "LDC"]

    dev_items.sort(
        key=lambda it: (it["vals"][-1] if it["vals"] and it["vals"][-1] is not None
                        else float("-inf")),
        reverse=True,
    )
    for i, it in enumerate(dev_items, 1):
        it["dev_rank"] = i

    devel_items.sort(
        key=lambda it: (it["vals"][-1] if it["vals"] and it["vals"][-1] is not None
                        else float("-inf")),
        reverse=True,
    )
    for i, it in enumerate(devel_items, 1):
        it["dev_rank"] = i

    ldc_items.sort(
        key=lambda it: (it["vals"][-1] if it["vals"] and it["vals"][-1] is not None
                        else float("-inf")),
        reverse=True,
    )
    for i, it in enumerate(ldc_items, 1):
        it["dev_rank"] = i

    # Always keep the pinned focus country (Kenya) visible in its own
    # development-status group, even when it falls outside the top N.
    # Pinned items are appended AFTER the top N (not inserted at top_n-1)
    # so they do not displace the Nth natural entry.
    pinned_added = []
    for group in (dev_items, devel_items, ldc_items):
        group_pinned = False
        for name in PINNED_DEV_COUNTRIES:
            pinned = next((it for it in group if _norm_country(it.get("label", "")) == name), None)
            if pinned is not None and pinned not in group[:top_n]:
                group.insert(top_n, pinned)
                group_pinned = True
        pinned_added.append(group_pinned)

    return (
        dev_items[:top_n + (1 if pinned_added[0] else 0)],
        devel_items[:top_n + (1 if pinned_added[1] else 0)],
        ldc_items[:top_n + (1 if pinned_added[2] else 0)],
    )


def rank_by_region(items, years, top_n=5):
    """Group items by region, rank within each, return top N per region.

    Returns dict mapping region -> list of items.
    """
    calc_growth_rates(items, years)

    by_region = {}
    for it in items:
        region = it.get("region", "Other")
        if region not in by_region:
            by_region[region] = []
        by_region[region].append(it)

    for region, region_items in by_region.items():
        region_items.sort(
            key=lambda it: (it["vals"][-1] if it["vals"] and it["vals"][-1] is not None
                            else float("-inf")),
            reverse=True,
        )
        for i, it in enumerate(region_items, 1):
            it["region_rank"] = i

    return {r: items[:top_n] for r, items in by_region.items()}


def sum_shown(shown, n_years):
    out = []
    for k in range(n_years):
        s = 0.0
        for it in shown:
            v = it["vals"][k] if k < len(it["vals"]) else None
            s += v if v is not None else 0.0
        out.append(s)
    return out


def all_other_vals(total, shown, n_years):
    """Total minus the sum of the displayed rows, per year."""
    if total is None:
        return [None] * n_years
    sums = sum_shown(shown, n_years)
    out = []
    for k in range(n_years):
        t = total["vals"][k] if k < len(total["vals"]) else None
        out.append(t - sums[k] if t is not None else None)
    return out


def share(v, total_v):
    if v is None or total_v is None or total_v == 0:
        return None
    return v / total_v


# ---------------------------------------------------------------------------
# Worksheet styling helpers (shared with make_tables.py)
# ---------------------------------------------------------------------------
def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_cell(cell, bold=False, size=FONT_SIZE, fill=None, numfmt=None,
               wrap=False, align="center", border_style="data"):
    """Apply professional styling to a cell.
    
    border_style: 'header', 'data', 'total', or 'none'
    """
    cell.font = Font(name=FONT, size=size, bold=bold, color="000000" if not bold else "1F3864")
    
    # Apply appropriate border
    if border_style == "header":
        cell.border = HEADER_BORDER
    elif border_style == "total":
        cell.border = TOTAL_BORDER
    elif border_style == "none":
        cell.border = Border()
    else:
        cell.border = DATA_BORDER
    
    if fill is not None:
        cell.fill = fill
    if numfmt is not None:
        cell.number_format = numfmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap,
                               indent=1 if align == "left" else 0)


def style_header(cell, text, size=HEADER_FONT_SIZE, fill=HDR_FILL, 
                 font_color=HDR_FONT_COLOR, align="center", wrap=True):
    """Style a header cell with professional dark header."""
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=font_color)
    cell.fill = fill
    cell.border = HEADER_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def get_row_fill(row_index):
    """Get alternating row fill for data rows (0-indexed)."""
    return ROW_FILL_EVEN if row_index % 2 == 0 else ROW_FILL_ODD


def put(ws, row, col, value, **kw):
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, **kw)
    return cell


def _put_formula(ws, cache, row, col, formula, cached, bold=False, fill=None,
                 numfmt=None):
    """Write an Excel formula and register the numeric result for caching."""
    put(ws, row, col, formula, bold=bold, fill=fill, numfmt=numfmt)
    cache.append((_cell_ref(row, col), cached))


def put_text(ws, row, col, text, bold=False, size=FONT_SIZE, fill=None, wrap=False,
             align="left", use_row_fill=True):
    """Write text with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, text, bold=bold, size=size, fill=fill,
               wrap=wrap, align=align)


def put_val(ws, row, col, value, bold=False, fill=None, fmt=FMT_VALUE, use_row_fill=True):
    """Write numeric value with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, value, bold=bold, fill=fill, numfmt=fmt)


def put_share(ws, row, col, value, bold=False, fill=None, use_row_fill=True):
    """Write share percentage with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, value, bold=bold, fill=fill, numfmt=FMT_SHARE)


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


# ---------------------------------------------------------------------------
# Table writers
# ---------------------------------------------------------------------------
def write_market_table(ws, data, hdr_label, verb, unit_row, kenya_highlight,
                       row1_label=None, row1_title=None):
    """Market-style table: rank | country | year values | share.

    Used for Tables 1 (exporters) and 2 (importers).
    """
    cache = []
    years = data["years"]
    n = len(years)
    latest = years[-1]
    total = data["total"]
    shown = data["shown"]

    share_col = 3 + n
    grow_col = share_col + 1
    cagr_col = share_col + 2

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, cagr_col)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    # Header row with dark navy fill
    put_text(ws, h, 1, f"Rank in {latest}", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 1).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 1, h + 1, 1)
    
    put_text(ws, h, 2, hdr_label, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 2).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 3, h, 2 + n)
    put_text(ws, h, 3, verb, bold=True, size=HEADER_FONT_SIZE if unit_row else HEADER_FONT_SIZE,
             fill=None if unit_row else HDR_FILL, wrap=True, align="center", use_row_fill=False)
    if not unit_row:
        ws.cell(h, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, share_col, h + 1, share_col)
    put_text(ws, h, share_col, f"Share in {latest} %", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, share_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    put_text(ws, h, grow_col, "Annual Growth %", bold=True, size=HEADER_FONT_SIZE,
             fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, grow_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    put_text(ws, h, cagr_col, cagr_header_label(years), bold=True, size=HEADER_FONT_SIZE,
             fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, cagr_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    for k, y in enumerate(years):
        put_text(ws, h + 1, 3 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 1, 3 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    if unit_row:
        merge(ws, h + 2, 3, h + 2, 2 + n)
        put_text(ws, h + 2, 3, "Value in USD Billion", bold=True, size=HEADER_FONT_SIZE,
                 fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 2, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
        put_text(ws, h + 2, 2, "", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        row0 = h + 3
    else:
        row0 = h + 2

    r_first = row0
    r_last = row0 + len(shown) - 1
    r_ao = row0 + len(shown)
    r_world = r_ao + 1

    for i, it in enumerate(shown):
        r = row0 + i
        is_kenya = kenya_highlight and str(it["label"]).strip().lower() == "kenya"
        fill = KENYA_FILL if is_kenya else get_row_fill(i)
        put_text(ws, r, 1, it["rank"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, display_name(it["label"]), size=FONT_SIZE, fill=fill, wrap=True, use_row_fill=False)
        for k in range(n):
            put_val(ws, r, 3 + k, it["vals"][k] if k < len(it["vals"]) else None, fill=fill, use_row_fill=False)
        t_latest = total["vals"][-1] if total and len(total["vals"]) >= n else None
        val_latest = it["vals"][-1] if it["vals"] else None
        cached_share = share(val_latest, t_latest)
        formula = f"={_cell_ref(r, 3 + n - 1)}/{_cell_ref(r_world, 3 + n - 1)}"
        _put_formula(ws, cache, r, share_col, formula, cached_share, fill=fill)
        put_share(ws, r, grow_col, it.get("growth"), fill=fill, use_row_fill=False)
        put_share(ws, r, cagr_col, it.get("cagr"), fill=fill, use_row_fill=False)

    # All other countries
    r = r_ao
    ao_fill = BAND_FILL
    put_text(ws, r, 2, "All other countries", size=FONT_SIZE, fill=ao_fill, wrap=True, use_row_fill=False)
    ao_vals = data["all_other"]["vals"]
    for k in range(n):
        col_letter = get_column_letter(3 + k)
        formula = f"={col_letter}{r_world}-SUM({col_letter}{r_first}:{col_letter}{r_last})"
        cached_val = ao_vals[k] if k < len(ao_vals) else None
        _put_formula(ws, cache, r, 3 + k, formula, cached_val, fill=ao_fill)
    ao_share_val = data["all_other"]["share"]
    ao_formula = f"={_cell_ref(r, 3 + n - 1)}/{_cell_ref(r_world, 3 + n - 1)}"
    _put_formula(ws, cache, r, share_col, ao_formula, ao_share_val, fill=ao_fill)

    # World total
    r = r_world
    total_fill = BAND_FILL
    put_text(ws, r, 2, "World", bold=True, size=FONT_SIZE, fill=total_fill, use_row_fill=False)
    ws.cell(r, 2).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
    for k in range(n):
        val = total["vals"][k] if k < len(total["vals"]) else None
        put_val(ws, r, 3 + k, val, bold=True, fill=total_fill, use_row_fill=False)
    put_share(ws, r, share_col, total.get("share"), bold=True)
    ws.freeze_panes = f"A{row0}"


    return cache


def write_product_table(ws, data, hdr, unit_row,
                        row1_label=None, row1_title=None):
    """Product-style table: rank | code | label | year values | share.

    Used for Tables 3 (Kenya exports) and 4 (Kenya imports).
    """
    cache = []
    years = data["years"]
    n = len(years)
    latest = years[-1]
    total = data["total"]
    shown = data["shown"]

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 2 + n)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    # Header row with dark navy fill
    put_text(ws, h, 1, f"Rank in {latest}", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 1).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 2, h + 1, 2)
    put_text(ws, h, 2, "Code", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 2).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 3, h + 1, 3)
    put_text(ws, h, 3, "Service label", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 4, h, 3 + n)
    put_text(ws, h, 4, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, 4).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    share_col = 4 + n
    grow_col = share_col + 1
    cagr_col = share_col + 2
    merge(ws, h, share_col, h + 1, share_col)
    put_text(ws, h, share_col, f"Share in {latest} %", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, share_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    put_text(ws, h, grow_col, "Annual Growth %", bold=True, size=HEADER_FONT_SIZE,
             fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, grow_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    put_text(ws, h, cagr_col, cagr_header_label(years), bold=True, size=HEADER_FONT_SIZE,
             fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, cagr_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    for k, y in enumerate(years):
        put_text(ws, h + 1, 4 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 1, 4 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    if unit_row:
        merge(ws, h + 2, 4, h + 2, 3 + n)
        put_text(ws, h + 2, 4, unit_row, bold=True, size=HEADER_FONT_SIZE,
                 fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 2, 4).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
        row0 = h + 3
    else:
        row0 = h + 2

    r_first = row0
    r_last = row0 + len(shown) - 1
    r_ao = row0 + len(shown)
    r_total = r_ao + 1

    for i, it in enumerate(shown):
        r = row0 + i
        fill = get_row_fill(i)
        put_text(ws, r, 1, it["rank"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, it["code"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 3, it["label"], size=FONT_SIZE, fill=fill, wrap=True, use_row_fill=False)
        for k in range(n):
            put_val(ws, r, 4 + k, it["vals"][k] if k < len(it["vals"]) else None, fill=fill, use_row_fill=False)
        t_latest = total["vals"][-1] if total and len(total["vals"]) >= n else None
        val_latest = it["vals"][-1] if it["vals"] else None
        cached_share = share(val_latest, t_latest)
        formula = f"={_cell_ref(r, 4 + n - 1)}/{_cell_ref(r_total, 4 + n - 1)}"
        _put_formula(ws, cache, r, share_col, formula, cached_share, fill=fill)

        # Annual growth (latest year-over-year) as a live Excel formula so the
        # figure is independently verifiable in the spreadsheet.  Mirrors
        # calc_growth_rates: (latest - prev) / prev using the two rightmost
        # populated year columns.
        latest_col = 4 + n - 1
        prev_col = 4 + n - 2
        v_latest = it["vals"][-1] if it["vals"] and it["vals"][-1] is not None else None
        v_prev = it["vals"][-2] if it["vals"] and len(it["vals"]) >= 2 and it["vals"][-2] is not None else None
        if n >= 2 and v_latest is not None and v_prev is not None and v_prev != 0:
            g_formula = (f"=({_cell_ref(r, latest_col)}-{_cell_ref(r, prev_col)})"
                         f"/{_cell_ref(r, prev_col)}")
            _put_formula(ws, cache, r, grow_col, g_formula, it.get("growth"),
                         fill=fill, numfmt=FMT_SHARE)
        else:
            put_share(ws, r, grow_col, it.get("growth"), fill=fill, use_row_fill=False)

        # Period CAGR as a live Excel formula, mirroring calc_cagr:
        # (endvalue / startvalue)^(1/span) - 1, where start/end are the first
        # and last populated year columns and span = last_year - first_year.
        n_years = (years[-1] - years[0]) if years else 0
        start_idx = end_idx = None
        if it.get("vals"):
            for _i, _v in enumerate(it["vals"]):
                if _v is not None and _v > 0:
                    if start_idx is None:
                        start_idx = _i
                    end_idx = _i
        if start_idx is not None and end_idx is not None and n_years > 0:
            c_formula = (f"=({_cell_ref(r, 4 + end_idx)}/{_cell_ref(r, 4 + start_idx)})"
                         f"^(1/{n_years})-1")
            _put_formula(ws, cache, r, cagr_col, c_formula, it.get("cagr"),
                         fill=fill, numfmt=FMT_SHARE)
        else:
            put_share(ws, r, cagr_col, it.get("cagr"), fill=fill, use_row_fill=False)

    # All other
    r = r_ao
    ao_fill = BAND_FILL
    put_text(ws, r, 3, "All other services", size=FONT_SIZE, fill=ao_fill, wrap=True, use_row_fill=False)
    ao_vals = data["all_other"]["vals"]
    for k in range(n):
        col_letter = get_column_letter(4 + k)
        formula = f"={col_letter}{r_total}-SUM({col_letter}{r_first}:{col_letter}{r_last})"
        cached_val = ao_vals[k] if k < len(ao_vals) else None
        _put_formula(ws, cache, r, 4 + k, formula, cached_val, fill=ao_fill)
    ao_share_val = data["all_other"]["share"]
    ao_formula = f"={_cell_ref(r, 4 + n - 1)}/{_cell_ref(r_total, 4 + n - 1)}"
    _put_formula(ws, cache, r, share_col, ao_formula, ao_share_val, fill=ao_fill)

    # Total
    r = r_total
    total_fill = BAND_FILL
    if total and total.get("code"):
        put_text(ws, r, 2, str(total["code"]), bold=True, size=FONT_SIZE,
                 fill=total_fill, wrap=True, use_row_fill=False)
    put_text(ws, r, 3, "All services", bold=True, size=FONT_SIZE, fill=total_fill, wrap=True, use_row_fill=False)
    ws.cell(r, 3).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
    for k in range(n):
        val = total["vals"][k] if k < len(total["vals"]) else None
        put_val(ws, r, 4 + k, val, bold=True, fill=total_fill, use_row_fill=False)
    put_share(ws, r, share_col, total.get("share") if total else None, bold=True, fill=total_fill)
    ws.freeze_panes = f"A{row0}"


    return cache


def write_balance(ws, years, exports, imports, cache=None):
    """Figure 1: Kenya services balance of trade.

    exports = Kenya's service exports per year (USD Million)
    imports = Kenya's service imports per year (USD Million)
    """
    cache = [] if cache is None else cache
    hdr = HDR_FILL
    put_text(ws, 1, 1, "Figure 1", bold=True, size=12, align="center", use_row_fill=False)
    put_text(ws, 1, 2, "Kenya Services Balance of Trade", bold=True, size=12, align="center", use_row_fill=False)

    # Header row
    put_text(ws, 2, 2, "", bold=True, size=HEADER_FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    for k, y in enumerate(years):
        put_text(ws, 3, 2 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
        ws.cell(3, 2 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    r_exp = 4
    r_imp = 5
    r_bal = 6
    
    # Row labels with header fill
    put_text(ws, r_exp, 1, "Exports", bold=True, size=FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    ws.cell(r_exp, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    for k, v in enumerate(exports):
        put_val(ws, r_exp, 2 + k, v, fill=ROW_FILL_EVEN)

    put_text(ws, r_imp, 1, "Imports", bold=True, size=FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    ws.cell(r_imp, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    for k, v in enumerate(imports):
        put_val(ws, r_imp, 2 + k, v, fill=ROW_FILL_ODD)

    put_text(ws, r_bal, 1, "Balance of Trade", bold=True, size=FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    ws.cell(r_bal, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    for k, (e, i) in enumerate(zip(exports, imports)):
        c = 2 + k
        if e is not None and i is not None:
            cache.append((_cell_ref(r_bal, c), e - i))
            put(ws, r_bal, c, f"={_cell_ref(r_exp, c)}-{_cell_ref(r_imp, c)}",
                bold=False, numfmt=FMT_VALUE, fill=BAND_FILL)
        else:
            put_val(ws, r_bal, c, (e - i) if (e is not None and i is not None) else None, fill=BAND_FILL)



# ---------------------------------------------------------------------------
# Development status table writer
# ---------------------------------------------------------------------------
def write_dev_status_table(ws, dev_items, devel_items, ldc_items, years,
                           verb, unit_label, row1_label=None, row1_title=None):
    """Write a table split by development status.

    Columns: Dev Status | Rank | Country | Region | Value (latest year) | Growth %
    The Development Status label is written once per group and merged across
    all rows in that group.
    """
    latest = years[-1] if years else ""
    n = len(years)

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 6)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Development Status", f"Rank in {latest}", verb, "Region",
            f"Value in USD Billion ({latest})", "Annual Growth %", cagr_header_label(years)]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center",
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    row0 = h + 1
    r = row0
    for status, items in [("Developed Economies", dev_items),
                          ("Developing Economies", devel_items),
                          ("Least Developed Countries (LDCs)", ldc_items)]:
        if not items:
            continue
        group_start = r
        fill = _dev_fill(status.split()[0] if status.startswith("Developed") else
                         ("LDC" if "LDC" in status else "Developing"))
        for it in items:
            is_pinned = _norm_country(it.get("label", "")) in PINNED_DEV_COUNTRIES
            row_fill = DEV_PIN_FILL if is_pinned else fill
            put_text(ws, r, 2, it.get("dev_rank", ""), size=FONT_SIZE, fill=row_fill, use_row_fill=False)
            put_text(ws, r, 3, it.get("label", ""), size=FONT_SIZE, fill=row_fill, wrap=True,
                     align="left", use_row_fill=False)
            put_text(ws, r, 4, it.get("region", ""), size=FONT_SIZE, fill=row_fill, wrap=True,
                     align="left", use_row_fill=False)
            val = it["vals"][-1] if it.get("vals") and it["vals"][-1] is not None else None
            put_val(ws, r, 5, val, fill=row_fill, use_row_fill=False)
            growth = it.get("growth")
            if growth is not None:
                put_share(ws, r, 6, growth, fill=row_fill, use_row_fill=False)
            else:
                put_share(ws, r, 6, None, fill=row_fill, use_row_fill=False)
            put_share(ws, r, 7, it.get("cagr"), fill=row_fill, use_row_fill=False)
            if is_pinned:
                for c in range(1, 8):
                    cell = ws.cell(r, c)
                    if cell.font:
                        cell.font = Font(name=FONT, size=cell.font.size or FONT_SIZE,
                                         bold=True, color=DEV_PIN_TEXT)
            r += 1
        # Write status label in first row and merge down for the group
        put_text(ws, group_start, 1, status, bold=True, size=FONT_SIZE, fill=fill,
                 wrap=True, align="center", use_row_fill=False)
        ws.cell(group_start, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
        if r - group_start > 1:
            merge(ws, group_start, 1, r - 1, 1)

    set_widths(ws, {"A": 18, "B": 14, "C": 32, "D": 28, "E": 22, "F": 18, "G": 14})
    ws.freeze_panes = f"A{row0}"


def write_region_table(ws, by_region, years, verb, unit_label,
                       row1_label=None, row1_title=None):
    """Write a table grouped by region.

    Columns: Region | Rank | Country | Value (latest year) | Growth %
    """
    latest = years[-1] if years else ""

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 5)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Region", f"Rank in {latest}", verb,
            f"Value in USD Billion ({latest})", "Annual Growth %", cagr_header_label(years)]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center",
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    row0 = h + 1
    r = row0
    # Use the globally defined REGION_FILLS
    for region in ["Africa", "Asia", "Americas", "Pacific", "Europe", "Other"]:
        items = by_region.get(region, [])
        if not items:
            continue
        fill = REGION_FILLS.get(region, BAND_FILL)
        for it in items:
            put_text(ws, r, 1, region, size=FONT_SIZE, fill=fill, wrap=True, align="left", use_row_fill=False)
            put_text(ws, r, 2, it.get("region_rank", ""), size=FONT_SIZE, fill=fill, use_row_fill=False)
            put_text(ws, r, 3, it.get("label", ""), size=FONT_SIZE, fill=fill, wrap=True,
                     align="left", use_row_fill=False)
            val = it["vals"][-1] if it.get("vals") and it["vals"][-1] is not None else None
            put_val(ws, r, 4, val, fill=fill, use_row_fill=False)
            growth = it.get("growth")
            if growth is not None:
                put_share(ws, r, 5, growth, fill=fill, use_row_fill=False)
            else:
                put_share(ws, r, 5, None, fill=fill, use_row_fill=False)
            put_share(ws, r, 6, it.get("cagr"), fill=fill, use_row_fill=False)
            r += 1

    set_widths(ws, {"A": 28, "B": 14, "C": 32, "D": 22, "E": 18, "F": 14})
    ws.freeze_panes = f"A{row0}"


# ---------------------------------------------------------------------------
# RCA (Revealed Comparative Advantage) analysis
# ---------------------------------------------------------------------------
def compute_services_rca(kenya_items, kenya_total, world_items, world_total, years):
    """Compute Balassa-style RCA for each service category.

    RCA_i = (Kenya_cat_i / Kenya_total) / (World_cat_i / World_total)

    Uses the latest year where Kenya has per-category data (not None).
    Returns list of dicts sorted by RCA (desc), each with:
      code, label, kenya_val, world_val, kenya_share, world_share, rca, rca_values
    """
    # Index world items by code
    world_by_code = {}
    for it in world_items:
        world_by_code[it["code"]] = it

    # Find the latest year with actual Kenya category data
    best_year_idx = None
    for k in range(len(years) - 1, -1, -1):
        has_data = any(
            k < len(it["vals"]) and it["vals"][k] is not None
            for it in kenya_items
        )
        if has_data:
            best_year_idx = k
            break
    if best_year_idx is None:
        best_year_idx = len(years) - 1

    results = []
    for kit in kenya_items:
        code = kit["code"]
        wit = world_by_code.get(code)
        if wit is None:
            continue

        entry = {
            "code": code,
            "label": kit["label"],
            "rca_values": [],
        }

        for k in range(len(years)):
            kv = kit["vals"][k] if k < len(kit["vals"]) else None
            kt = kenya_total["vals"][k] if kenya_total and k < len(kenya_total["vals"]) else None
            wv = wit["vals"][k] if k < len(wit["vals"]) else None
            wt = world_total["vals"][k] if world_total and k < len(world_total["vals"]) else None

            if kv is not None and kt is not None and kt != 0 and wv is not None and wt is not None and wt != 0:
                k_share = kv / kt
                w_share = wv / wt
                rca = k_share / w_share if w_share != 0 else None
            else:
                k_share = None
                w_share = None
                rca = None

            if k == best_year_idx:
                entry["kenya_val"] = kv
                entry["world_val"] = wv
                entry["kenya_share"] = k_share
                entry["world_share"] = w_share
                entry["rca"] = rca
            entry["rca_values"].append(rca)

        entry["best_year"] = years[best_year_idx]
        results.append(entry)

    results.sort(key=lambda x: x["rca"] if x["rca"] is not None else -1, reverse=True)
    for i, it in enumerate(results, 1):
        it["rank"] = i
    return results, years[best_year_idx]


def write_rca_table(ws, rca_items, years, row1_label=None, row1_title=None):
    """Write an RCA analysis table.

    Columns: Rank | Code | Service Category | Kenya Share % | World Share % | RCA | Status
    """
    latest = years[-1] if years else ""

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 7)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Rank", "Code", f"Service Category",
            f"Kenya Share ({latest}) %", f"World Share ({latest}) %",
            f"RCA ({latest})", "Classification"]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center",
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    # RCA classification fills (professional muted tones)
    STRONG_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
    MODERATE_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
    DISADVANTAGE_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")

    row0 = h + 1
    for i, it in enumerate(rca_items):
        r = row0 + i
        rca_val = it.get("rca")
        if rca_val is not None and rca_val >= 2.5:
            fill = STRONG_FILL
            status = "Strong Advantage"
        elif rca_val is not None and rca_val >= 1.0:
            fill = MODERATE_FILL
            status = "Moderate Advantage"
        else:
            fill = DISADVANTAGE_FILL
            status = "Disadvantage"

        put_text(ws, r, 1, it.get("rank", ""), size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, str(it.get("code", "")), size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 3, it.get("label", ""), size=FONT_SIZE, fill=fill, wrap=True,
                 align="left", use_row_fill=False)
        put_share(ws, r, 4, it.get("kenya_share"), fill=fill, use_row_fill=False)
        put_share(ws, r, 5, it.get("world_share"), fill=fill, use_row_fill=False)
        if rca_val is not None:
            put(ws, r, 6, round(rca_val, 2), fill=fill, numfmt=FMT_RCA)
        else:
            put(ws, r, 6, "", fill=fill)
        put_text(ws, r, 7, status, size=FONT_SIZE, fill=fill, align="center", use_row_fill=False)

    # Legend row
    r = row0 + len(rca_items) + 1
    put(ws, r, 1, "RCA > 2.5: Strong comparative advantage", size=9)
    r += 1
    put(ws, r, 1, "RCA 1.0-2.5: Moderate comparative advantage", size=9)
    r += 1
    put(ws, r, 1, "RCA < 1.0: Comparative disadvantage", size=9)

    set_widths(ws, {"A": 8, "B": 8, "C": 48, "D": 18, "E": 18, "F": 14, "G": 22})
    ws.freeze_panes = f"A{row0}"


# ---------------------------------------------------------------------------
# Export Concentration Index (HHI) analysis
# ---------------------------------------------------------------------------
def compute_concentration_index(kenya_items, kenya_total, world_items, world_total, years):
    """Compute HHI and concentration ratios for Kenya and World service exports.

    HHI = sum of squared shares across categories.
    Lower HHI = more diversified.
    """
    latest_idx = -1

    def _shares(items, total):
        vals = []
        for it in items:
            v = it["vals"][latest_idx] if it.get("vals") and len(it["vals"]) > abs(latest_idx) else None
            if v is not None:
                vals.append(v)
        t = total["vals"][latest_idx] if total and total.get("vals") and len(total["vals"]) > abs(latest_idx) else None
        if t is None or t == 0:
            return []
        return [v / t for v in vals]

    ke_shares = _shares(kenya_items, kenya_total)
    w_shares = _shares(world_items, world_total)

    def _metrics(shares):
        if not shares:
            return {"hhi": 0, "top3": 0, "top5": 0, "eff_cats": 0}
        sorted_s = sorted(shares, reverse=True)
        hhi = sum(s * s for s in sorted_s)
        top3 = sum(sorted_s[:3])
        top5 = sum(sorted_s[:5])
        eff = 1.0 / hhi if hhi > 0 else 0
        return {"hhi": hhi, "top3": top3, "top5": top5, "eff_cats": eff}

    ke_m = _metrics(ke_shares)
    w_m = _metrics(w_shares)

    return {
        "kenya": ke_m,
        "world": w_m,
        "kenya_n_categories": len(ke_shares),
        "world_n_categories": len(w_shares),
    }


def write_concentration_table(ws, concentration, row1_label=None, row1_title=None):
    """Write export concentration index comparison table."""
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 4)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Metric", "Kenya", "World", "Assessment"]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", 
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    ke = concentration["kenya"]
    w = concentration["world"]

    def _assess(ke_val, w_val, lower_is_better=True):
        if lower_is_better:
            if ke_val < w_val * 0.8:
                return "More diversified"
            elif ke_val > w_val * 1.2:
                return "More concentrated"
        else:
            if ke_val > w_val * 1.2:
                return "More diversified"
            elif ke_val < w_val * 0.8:
                return "More concentrated"
        return "Similar"

    rows = [
        ("Herfindahl-Hirschman Index (HHI)", ke["hhi"], w["hhi"],
         _assess(ke["hhi"], w["hhi"]), "0.000"),
        ("Top-3 Category Share", ke["top3"], w["top3"],
         _assess(ke["top3"], w["top3"]), FMT_SHARE),
        ("Top-5 Category Share", ke["top5"], w["top5"],
         _assess(ke["top5"], w["top5"]), FMT_SHARE),
        ("Effective No. of Categories (1/HHI)", ke["eff_cats"], w["eff_cats"],
         _assess(ke["eff_cats"], w["eff_cats"], lower_is_better=False), "0.0"),
        ("Number of Service Categories", concentration["kenya_n_categories"],
         concentration["world_n_categories"], "", FMT_VALUE_INT),
    ]

    row0 = h + 1
    for i, (label, ke_val, w_val, assess, fmt) in enumerate(rows):
        r = row0 + i
        fill = ROW_FILL_EVEN if i % 2 == 0 else ROW_FILL_ODD
        put_text(ws, r, 1, label, size=FONT_SIZE, fill=fill, align="left", use_row_fill=False)
        put(ws, r, 2, ke_val, fill=fill, numfmt=fmt)
        put(ws, r, 3, w_val, fill=fill, numfmt=fmt)
        put_text(ws, r, 4, assess, size=FONT_SIZE, fill=fill, align="center", use_row_fill=False)

    r = row0 + len(rows) + 1
    put(ws, r, 1, "HHI < 0.15: Low concentration (diversified)", size=9)
    r += 1
    put(ws, r, 1, "HHI 0.15-0.25: Moderate concentration", size=9)
    r += 1
    put(ws, r, 1, "HHI > 0.25: High concentration", size=9)

    set_widths(ws, {"A": 42, "B": 18, "C": 18, "D": 24})
    ws.freeze_panes = f"A{row0}"


def make_concentration_chart(concentration, out_path):
    """Bar chart comparing Kenya vs World concentration metrics."""
    metrics = ["HHI", "Top-3 Share", "Top-5 Share"]
    ke_vals = [concentration["kenya"]["hhi"], concentration["kenya"]["top3"],
               concentration["kenya"]["top5"]]
    w_vals = [concentration["world"]["hhi"], concentration["world"]["top3"],
              concentration["world"]["top5"]]

    fig, ax = plt.subplots(figsize=(8, 5))
    x = np.arange(len(metrics))
    width = 0.35
    bars1 = ax.bar(x - width/2, ke_vals, width, label="Kenya", color="#2E75B6")
    bars2 = ax.bar(x + width/2, w_vals, width, label="World", color="#ED7D31")

    ax.set_xticks(x)
    ax.set_xticklabels(metrics, fontsize=10)
    ax.set_ylabel("Value", fontsize=10)
    ax.set_title("Kenya vs World: Export Concentration Metrics", fontsize=12, weight="bold")
    ax.legend(fontsize=10)
    ax.yaxis.grid(True, linestyle="--", alpha=0.35)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)

    for bar in bars1:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.005,
                f"{bar.get_height():.3f}", ha="center", va="bottom", fontsize=8)
    for bar in bars2:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.005,
                f"{bar.get_height():.3f}", ha="center", va="bottom", fontsize=8)

    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Diversification Potential analysis
# ---------------------------------------------------------------------------
def compute_diversification_potential(kenya_items, kenya_total, world_items, world_total,
                                       kenya_imp_items, kenya_imp_total, years):
    """Identify categories where Kenya could expand exports.

    Opportunity score = global_growth * (1 - kenya_global_share) * import_penetration
    """
    # Find latest year with actual Kenya category data
    latest_idx = None
    for k in range(len(years) - 1, -1, -1):
        has_data = any(
            k < len(it["vals"]) and it["vals"][k] is not None
            for it in kenya_items
        )
        if has_data:
            latest_idx = k
            break
    if latest_idx is None:
        latest_idx = len(years) - 1

    prev_idx = latest_idx - 1 if latest_idx > 0 else None

    world_by_code = {it["code"]: it for it in world_items}
    kenya_imp_by_code = {it["code"]: it for it in kenya_imp_items}

    ke_total_latest = kenya_total["vals"][latest_idx] if kenya_total and kenya_total.get("vals") and latest_idx < len(kenya_total["vals"]) else None
    ke_total_prev = kenya_total["vals"][prev_idx] if kenya_total and kenya_total.get("vals") and prev_idx is not None and prev_idx < len(kenya_total["vals"]) else None
    w_total_latest = world_total["vals"][latest_idx] if world_total and world_total.get("vals") and latest_idx < len(world_total["vals"]) else None
    w_total_prev = world_total["vals"][prev_idx] if world_total and world_total.get("vals") and prev_idx is not None and prev_idx < len(world_total["vals"]) else None

    results = []
    for kit in kenya_items:
        code = kit["code"]
        if code.upper() == "S" or kit.get("label", "").lower() == "all services":
            continue

        wit = world_by_code.get(code)
        kit_imp = kenya_imp_by_code.get(code)

        ke_val = kit["vals"][latest_idx] if kit.get("vals") and len(kit["vals"]) > abs(latest_idx) else None
        w_val = wit["vals"][latest_idx] if wit and wit.get("vals") and len(wit["vals"]) > abs(latest_idx) else None

        if ke_val is None or w_val is None or w_total_latest is None or w_total_latest == 0:
            continue

        ke_global_share = ke_val / w_total_latest

        ke_val_prev = kit["vals"][prev_idx] if kit.get("vals") and prev_idx is not None and prev_idx < len(kit["vals"]) else None
        w_val_prev = wit["vals"][prev_idx] if wit and wit.get("vals") and prev_idx is not None and prev_idx < len(wit["vals"]) else None

        if ke_val_prev is not None and ke_val_prev > 0:
            ke_growth = (ke_val - ke_val_prev) / ke_val_prev
        else:
            ke_growth = None

        if w_val_prev is not None and w_val_prev > 0:
            w_growth = (w_val - w_val_prev) / w_val_prev
        else:
            w_growth = None

        imp_penetration = 0
        if kit_imp and kenya_imp_total:
            imp_val = kit_imp["vals"][latest_idx] if kit_imp.get("vals") and len(kit_imp["vals"]) > abs(latest_idx) else None
            imp_total = kenya_imp_total["vals"][latest_idx] if kenya_imp_total.get("vals") and len(kenya_imp_total["vals"]) > abs(latest_idx) else None
            if imp_val is not None and imp_total is not None and imp_total > 0:
                imp_penetration = imp_val / imp_total

        growth_factor = max(0, w_growth) if w_growth is not None else 0
        gap_factor = 1.0 - ke_global_share
        opportunity_score = growth_factor * gap_factor * (1 + imp_penetration)

        results.append({
            "code": code,
            "label": kit.get("label", ""),
            "ke_val": ke_val,
            "ke_global_share": ke_global_share,
            "ke_growth": ke_growth,
            "w_growth": w_growth,
            "imp_penetration": imp_penetration,
            "opportunity_score": opportunity_score,
        })

    results.sort(key=lambda x: x["opportunity_score"], reverse=True)
    for i, it in enumerate(results, 1):
        it["rank"] = i
    return results


def write_diversification_table(ws, items, row1_label=None, row1_title=None):
    """Write diversification potential table."""
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=11, align="center")
        merge(ws, 1, 2, 1, 7)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Rank", "Code", "Service Category", "Kenya Export\n(USD Mn)",
            "Kenya Global\nShare %", "Global\nGrowth %", "Import\nShare %",
            "Opportunity\nScore"]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", 
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    # Opportunity score fills
    HIGH_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
    MED_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
    LOW_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")

    row0 = h + 1
    for i, it in enumerate(items[:15]):
        r = row0 + i
        score = it.get("opportunity_score", 0)
        if score >= 0.05:
            fill = HIGH_FILL
        elif score >= 0.02:
            fill = MED_FILL
        else:
            fill = LOW_FILL

        put_text(ws, r, 1, it.get("rank", ""), size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, str(it.get("code", "")), size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 3, it.get("label", ""), size=FONT_SIZE, fill=fill, wrap=True, align="left", use_row_fill=False)
        put_val(ws, r, 4, it.get("ke_val"), fill=fill, use_row_fill=False)
        put_share(ws, r, 5, it.get("ke_global_share"), fill=fill, use_row_fill=False)
        put_share(ws, r, 6, it.get("w_growth"), fill=fill, use_row_fill=False)
        put_share(ws, r, 7, it.get("imp_penetration"), fill=fill, use_row_fill=False)
        put(ws, r, 8, round(score, 4), fill=fill, numfmt="0.0000")

    r = row0 + min(len(items), 15) + 1
    put(ws, r, 1, "Higher opportunity score = greater diversification potential", size=9)

    set_widths(ws, {"A": 7, "B": 8, "C": 42, "D": 16, "E": 16, "F": 14, "G": 14, "H": 14})
    ws.freeze_panes = f"A{row0}"


def make_diversification_chart(items, out_path):
    """Horizontal bar chart of top diversification opportunities."""
    top = [it for it in items[:10] if it.get("opportunity_score", 0) > 0]
    if not top:
        return
    top.sort(key=lambda x: x["opportunity_score"])

    labels = [it["label"][:35] for it in top]
    scores = [it["opportunity_score"] for it in top]

    fig, ax = plt.subplots(figsize=(10, max(4, len(top) * 0.5)))
    colors = ["#C6EFCE" if s >= 0.05 else "#E2EFDA" if s >= 0.02 else "#F2F2F2" for s in scores]
    bars = ax.barh(labels, scores, color=colors, edgecolor="#808080", linewidth=0.5)
    ax.set_xlabel("Opportunity Score", fontsize=10)
    ax.set_title("Kenya's Top Diversification Opportunities in Services", fontsize=12, weight="bold")
    for bar, val in zip(bars, scores):
        ax.text(bar.get_width() + 0.001, bar.get_y() + bar.get_height() / 2,
                f"{val:.4f}", va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Value Composition Trajectory analysis
# ---------------------------------------------------------------------------
HIGH_VALUE_CODES = {"7", "8", "9", "10"}   # Financial, IP, ICT, Other business
TRADITIONAL_CODES = {"3", "4", "5"}         # Transport, Travel, Construction


def compute_value_trajectory(kenya_items, kenya_total, years):
    """Group categories into high-value, traditional, other and compute shares over time."""
    results = {"years": years, "high_value": [], "traditional": [], "other": []}
    n = len(years)

    hv_totals = [0.0] * n
    trad_totals = [0.0] * n
    other_totals = [0.0] * n

    for kit in kenya_items:
        code = kit.get("code", "")
        for k in range(n):
            v = kit["vals"][k] if kit.get("vals") and k < len(kit["vals"]) and kit["vals"][k] is not None else 0
            if code in HIGH_VALUE_CODES:
                hv_totals[k] += v
            elif code in TRADITIONAL_CODES:
                trad_totals[k] += v
            else:
                other_totals[k] += v

    total_vals = []
    for k in range(n):
        t = kenya_total["vals"][k] if kenya_total and kenya_total.get("vals") and k < len(kenya_total["vals"]) and kenya_total["vals"][k] is not None else 0
        total_vals.append(t)

    for k in range(n):
        t = total_vals[k] if total_vals[k] > 0 else 1
        results["high_value"].append(hv_totals[k] / t)
        results["traditional"].append(trad_totals[k] / t)
        results["other"].append(other_totals[k] / t)
        results.setdefault("abs_high_value", []).append(hv_totals[k])
        results.setdefault("abs_traditional", []).append(trad_totals[k])
        results.setdefault("abs_other", []).append(other_totals[k])
        results.setdefault("abs_total", []).append(total_vals[k])

    return results


def write_value_trajectory_table(ws, trajectory, row1_label=None, row1_title=None):
    """Write value composition trajectory table."""
    years = trajectory["years"]
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 3 + len(years))
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Category Group", "Service Codes"] + [str(y) for y in years]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", 
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    # Category group fills (professional muted tones)
    HV_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
    TRAD_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
    OTHER_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")

    row0 = h + 1
    groups = [
        ("High-Value Services", "7, 8, 9, 10", trajectory["high_value"], HV_FILL,
         "Financial, IP, ICT, Other business"),
        ("Traditional Services", "3, 4, 5", trajectory["traditional"], TRAD_FILL,
         "Transport, Travel, Construction"),
        ("Other Services", "1, 2, 6, 11, 12, SN", trajectory["other"], OTHER_FILL,
         "Manufacturing, Maintenance, Insurance, Personal, Government"),
    ]

    for i, (label, codes, shares, fill, desc) in enumerate(groups):
        r = row0 + i
        put_text(ws, r, 1, label, size=FONT_SIZE, fill=fill, align="left", bold=True, use_row_fill=False)
        put_text(ws, r, 2, codes, size=9, fill=fill, align="center", use_row_fill=False)
        for k, sh in enumerate(shares):
            put_share(ws, r, 3 + k, sh, fill=fill, use_row_fill=False)

    r = row0 + len(groups)
    put_text(ws, r, 1, "Total Exports (USD Mn)", size=FONT_SIZE, bold=True, align="left", use_row_fill=False)
    ws.cell(r, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
    put_text(ws, r, 2, "", size=9)
    for k, t in enumerate(trajectory.get("abs_total", [])):
        put_val(ws, r, 3 + k, t, fill=BAND_FILL, use_row_fill=False)

    r += 1
    put_text(ws, r, 1, "High-Value Exports (USD Mn)", size=FONT_SIZE, align="left", use_row_fill=False)
    put_text(ws, r, 2, "", size=9)
    for k, v in enumerate(trajectory.get("abs_high_value", [])):
        put_val(ws, r, 3 + k, v, fill=HV_FILL, use_row_fill=False)

    r += 1
    put_text(ws, r, 1, "Traditional Exports (USD Mn)", size=FONT_SIZE, align="left", use_row_fill=False)
    put_text(ws, r, 2, "", size=9)
    for k, v in enumerate(trajectory.get("abs_traditional", [])):
        put_val(ws, r, 3 + k, v, fill=TRAD_FILL, use_row_fill=False)

    widths = {"A": 26, "B": 18}
    for k in range(len(years)):
        widths[get_column_letter(3 + k)] = 14
    set_widths(ws, widths)
    ws.freeze_panes = f"A{row0}"


def make_value_trajectory_chart(trajectory, out_path):
    """Stacked area chart of value composition over time."""
    years = trajectory["years"]
    hv = [s * 100 for s in trajectory["high_value"]]
    trad = [s * 100 for s in trajectory["traditional"]]
    other = [s * 100 for s in trajectory["other"]]

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.stackplot(years, hv, trad, other,
                 labels=["High-Value Services", "Traditional Services", "Other Services"],
                 colors=["#2E75B6", "#70AD47", "#A5A5A5"], alpha=0.85)
    ax.set_xlabel("Year", fontsize=10)
    ax.set_ylabel("Share of Total Exports (%)", fontsize=10)
    ax.set_title("Kenya's Service Export Composition Over Time", fontsize=12, weight="bold")
    ax.legend(loc="upper right", fontsize=9)
    ax.set_ylim(0, 100)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.yaxis.grid(True, linestyle="--", alpha=0.35)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Kenya vs Peer comparison
# ---------------------------------------------------------------------------
PEER_COUNTRIES = {
    "African Peers": ["south africa", "egypt", "mauritius", "rwanda"],
    "Aspirational Peers": ["singapore", "malaysia"],
}


def write_peer_comparison_table(ws, items_exp, years_exp, row1_label=None, row1_title=None):
    """Write Kenya vs peer countries total service exports comparison."""
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 6)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Group", "Country", f"Exports\n({years_exp[-1]}) USD Bn",
            "Global\nShare %", "Growth\n(YoY) %", f"{years_exp[-1] - years_exp[0]}Y CAGR\n%", "Rank\n(Global)"]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", 
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    items_by_name = {it["label"].strip().lower(): it for it in items_exp}

    row0 = h + 1
    r = row0

    # Peer comparison fills
    GROUP_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
    PEER_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
    KENYA_PEER_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for group_name, peers in PEER_COUNTRIES.items():
        group_start = r
        for peer in peers:
            it = items_by_name.get(peer)
            if it is None:
                continue
            fill = PEER_FILL
            put_text(ws, r, 1, group_name if r == group_start else "", size=FONT_SIZE, fill=fill,
                     align="left", bold=(r == group_start), use_row_fill=False)
            put_text(ws, r, 2, it.get("label", ""), size=FONT_SIZE, fill=fill, align="left", use_row_fill=False)
            val = it["vals"][-1] if it.get("vals") and it["vals"][-1] is not None else None
            put_val(ws, r, 3, val, fill=fill, use_row_fill=False)
            put_share(ws, r, 4, it.get("share"), fill=fill, use_row_fill=False)
            growth = it.get("growth")
            if growth is not None:
                put_share(ws, r, 5, growth, fill=fill, use_row_fill=False)
            else:
                put_share(ws, r, 5, None, fill=fill, use_row_fill=False)
            put_share(ws, r, 7, it.get("cagr"), fill=fill, use_row_fill=False)
            put_text(ws, r, 8, it.get("rank", ""), size=FONT_SIZE, fill=fill, use_row_fill=False)
            r += 1

    ke_it = items_by_name.get("kenya")
    if ke_it:
        fill = KENYA_PEER_FILL
        put_text(ws, r, 1, "Focus Country", size=FONT_SIZE, fill=fill, align="left", bold=True, use_row_fill=False)
        ws.cell(r, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
        put_text(ws, r, 2, ke_it.get("label", ""), size=FONT_SIZE, fill=fill, align="left", bold=True, use_row_fill=False)
        ws.cell(r, 2).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
        val = ke_it["vals"][-1] if ke_it.get("vals") and ke_it["vals"][-1] is not None else None
        put_val(ws, r, 3, val, fill=fill, bold=True, use_row_fill=False)
        put_share(ws, r, 4, ke_it.get("share"), fill=fill, bold=True, use_row_fill=False)
        growth = ke_it.get("growth")
        if growth is not None:
            put_share(ws, r, 5, growth, fill=fill, bold=True, use_row_fill=False)
        else:
            put_share(ws, r, 5, None, fill=fill, bold=True, use_row_fill=False)
        put_share(ws, r, 7, ke_it.get("cagr"), fill=fill, bold=True, use_row_fill=False)
        put_text(ws, r, 8, ke_it.get("rank", ""), size=FONT_SIZE, fill=fill, bold=True, use_row_fill=False)
        r += 1

    set_widths(ws, {"A": 20, "B": 28, "C": 20, "D": 14, "E": 14, "F": 14, "G": 12})
    ws.freeze_panes = f"A{row0}"


def make_rca_radar_chart(rca_items, years, out_path):
    """Create a horizontal bar chart of RCA values for service categories."""
    items = [it for it in rca_items if it.get("rca") is not None]
    items.sort(key=lambda x: x["rca"])

    labels = [it["label"][:35] for it in items]
    rca_vals = [it["rca"] for it in items]

    fig, ax = plt.subplots(figsize=(10, max(4, len(items) * 0.5)))
    colors = ["#C6EFCE" if v >= 2.5 else "#E2EFDA" if v >= 1.0 else "#FCE4D6"
              for v in rca_vals]
    bars = ax.barh(labels, rca_vals, color=colors, edgecolor="#808080", linewidth=0.5)
    ax.axvline(x=1.0, color="#808080", linestyle="--", linewidth=0.8, label="RCA = 1.0")
    ax.axvline(x=2.5, color="#2E75B6", linestyle="--", linewidth=0.8, label="RCA = 2.5")
    ax.set_xlabel("Revealed Comparative Advantage (RCA)", fontsize=10)
    ax.set_title("Kenya's Service Export RCA by Category", fontsize=12, weight="bold")
    ax.legend(fontsize=8)
    for bar, val in zip(bars, rca_vals):
        ax.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}", va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Chart generators
# ---------------------------------------------------------------------------
def make_pie_chart(exports_by_category, years, out_path):
    """Create a pie chart of global service exports by category.

    exports_by_category: list of {'label': str, 'vals': [float]}
    years: list of year ints
    """
    latest_idx = -1
    data = []
    for item in exports_by_category:
        label = item.get("label", "")
        if label.lower() == "all services":
            continue
        val = item["vals"][latest_idx] if item.get("vals") and len(item["vals"]) > abs(latest_idx) else None
        if val is not None and val > 0:
            data.append((label, val))

    data.sort(key=lambda x: x[1], reverse=True)

    # Group small slices (< 3% of total) into "Other"
    total = sum(v for _, v in data)
    main_slices = []
    other_total = 0
    for label, val in data:
        if val / total >= 0.03:
            main_slices.append((label, val))
        else:
            other_total += val
    if other_total > 0:
        main_slices.append(("Other services", other_total))

    labels = [short_product_name(s[0], maxlen=40) for s in main_slices]
    sizes = [s[1] for s in main_slices]

    colors = ["#2E75B6", "#ED7D31", "#A5A5A5", "#FFC000", "#4472C4",
              "#70AD47", "#264478", "#9B57A0", "#636363", "#EB7E30"]

    fig, ax = new_fig(8, 6, 150)
    labels, sizes, wedges = draw_share_pie(
        ax, labels, sizes, colors,
        style="3d_exploded", other_label="Other services",
        max_slices=8, min_pct=3.0)
    total = sum(sizes) or 1.0
    ax.legend(wedges, [f"{l} - {v / total * 100:.1f}%" for l, v in zip(labels, sizes)],
              loc="center left", bbox_to_anchor=(1.0, 0.5), fontsize=8)
    ax.set_title("Global Service Exports by Category", fontsize=12, weight="bold")
    finish(fig, out_path)


def make_stacked_bar_chart(category_items, years, out_path):
    """Create a stacked bar chart of service exports by category across years.

    category_items: list of {'label': str, 'vals': [float]}
    years: list of year ints
    """
    # Filter out "All services" total
    items = [it for it in category_items if it.get("label", "").lower() != "all services"]
    items.sort(key=lambda it: it["vals"][-1] if it.get("vals") and it["vals"][-1] is not None else 0,
               reverse=True)

    # Take top 8 categories, group rest as "Other"
    top = items[:8]
    other_items = items[8:]

    fig, ax = plt.subplots(figsize=(10, 6))

    x = np.arange(len(years))
    width = 0.6
    colors = ["#2E75B6", "#ED7D31", "#A5A5A5", "#FFC000", "#4472C4",
              "#70AD47", "#264478", "#9B57A0", "#636363"]
    bottom = np.zeros(len(years))

    for i, item in enumerate(top):
        vals = []
        for k in range(len(years)):
            v = item["vals"][k] if k < len(item.get("vals", [])) and item["vals"][k] is not None else 0
            vals.append(v)
        vals = np.array(vals) / 1e3  # Convert to billion
        ax.bar(x, vals, width, bottom=bottom, label=item["label"],
               color=colors[i % len(colors)])
        bottom += vals

    if other_items:
        other_vals = np.zeros(len(years))
        for item in other_items:
            for k in range(len(years)):
                v = item["vals"][k] if k < len(item.get("vals", [])) and item["vals"][k] is not None else 0
                other_vals[k] += v / 1e3
        ax.bar(x, other_vals, width, bottom=bottom, label="Other services",
               color=colors[-1])

    ax.set_xticks(x)
    ax.set_xticklabels(years, fontsize=10)
    ax.set_ylabel("Value (USD Billion)", fontsize=10)
    ax.set_title("Global Service Exports by Category", fontsize=12, weight="bold")
    ax.legend(loc="upper left", bbox_to_anchor=(1.0, 1.0), fontsize=8)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def make_regional_bar_chart(by_region, years, out_path, title="Service Exports by Region"):
    """Create a grouped bar chart showing top countries per region.

    by_region: dict mapping region -> list of items
    years: list of year ints
    """
    # Get top 3 per region, order regions by total value
    region_totals = {}
    for region, items in by_region.items():
        total = sum(
            (it["vals"][-1] if it.get("vals") and it["vals"][-1] is not None else 0)
            for it in items
        )
        region_totals[region] = total

    ordered_regions = sorted(region_totals.keys(), key=lambda r: region_totals[r], reverse=True)

    fig, ax = plt.subplots(figsize=(12, 6))
    x = np.arange(len(ordered_regions))
    width = 0.2
    colors = ["#2E75B6", "#ED7D31", "#70AD47"]

    for rank in range(3):
        vals = []
        for region in ordered_regions:
            items = by_region.get(region, [])
            if rank < len(items):
                v = items[rank]["vals"][-1] if items[rank]["vals"] and items[rank]["vals"][-1] is not None else 0
                vals.append(v / 1e6)  # Convert to billion
            else:
                vals.append(0)
        offset = (rank - 1) * width
        label = f"Top {rank + 1}"
        ax.bar(x + offset, vals, width, label=label, color=colors[rank])

    ax.set_xticks(x)
    ax.set_xticklabels(ordered_regions, fontsize=9, rotation=15, ha="right")
    ax.set_ylabel("Value (USD Billion)", fontsize=10)
    ax.set_title(title, fontsize=12, weight="bold")
    ax.legend(fontsize=9)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate_service_tables(excel_dir, out_dir, top_n):
    """Build the service Table 1-4 + Figure 1 workbooks.

    top_n : number of top exporter/importer rows kept for Tables 1 and 2.
    """
    os.makedirs(out_dir, exist_ok=True)
    files = find_service_files(excel_dir)

    # ---- UNCTADstat service-by-category data (annual).
    # Preferred source: the R-export CSV (``unctad_services_<years>.csv``) that
    # is uploaded alongside the ITC files.  It supplies Kenya's per-category
    # detail (Tables 3/4), economy-level total services and the true published
    # World total (Tables 1/2, Table 13), and per-category world sums (world
    # pie / RCA / concentration).  Fallback: the all-economy gz from
    # fetch_unctad_tradeserv.py (``UNCTAD_*_tradeserv_annual_all.csv.gz``).
    un_annual = None
    if files.get("unctad_r") or files.get("unctad_annual"):
        try:
            un_annual = None
            last_err = None
            # 1) R-export CSV (the user's canonical UNCTAD file).
            for cand in sorted(files.get("unctad_r", [])):
                try:
                    un_annual = parse_unctad_services_r(cand)
                    if un_annual.get("country_s"):
                        break
                except Exception as exc:  # noqa: BLE001 - try next candidate
                    last_err = exc
                    un_annual = None
            # 2) Fallback: all-economy gz (prefer *_all, else Kenya-only shape).
            if not un_annual:
                cands = sorted(files.get("unctad_annual", []),
                               key=lambda p: (os.path.basename(p).lower().find("_all") == -1,
                                              os.path.basename(p)))
                for cand in cands:
                    try:
                        un_annual = parse_unctad_annual(cand)
                        if un_annual.get("country_s"):
                            break
                    except Exception as exc:  # noqa: BLE001 - try next candidate
                        last_err = exc
                        un_annual = None
            if isinstance(last_err, Exception) and not un_annual:
                raise last_err
        except Exception as exc:  # noqa: BLE001 - optional source
            print(f"Warning: ignoring unusable UNCTAD annual services file: {exc}")
            un_annual = None

    # R-file primary, all-economy gz fills Kenya category gaps (e.g. SA/SB that
    # the R download omits) so no genuinely-available cell is left blank.
    if un_annual:
        _gage_kenya_from_annual_gz(files, un_annual)

    # Kenya's latest *reported* total-services (S) year in the annual dataset.
    # This caps the global country-ranking axis (Tables 1/2) and the world pie so
    # they stay aligned with Kenya's coverage.  Kenya's per-category tables (3/4)
    # are separately capped at 2024 (2025 has no category breakdown for Kenya).
    max_kenya_year = None
    if un_annual and "kenya" in un_annual and "S" in un_annual["kenya"].get("Exports", {}):
        k_years = sorted(un_annual["kenya"]["Exports"]["S"])
        max_kenya_year = k_years[-1] if k_years else None

    # ---- Table 1: Global service exporters by country --------------------
    total_exp, items_exp, years_exp = _load_first(files, "exporters", parse_exporters)
    # Convert to billions
    for it in items_exp:
        it["vals"] = [v / 1e6 if v is not None else None for v in it["vals"]]
    if total_exp:
        total_exp["vals"] = [v / 1e6 if v is not None else None for v in total_exp["vals"]]

    if un_annual and un_annual.get("country_s"):
        items_exp, total_exp, years_exp, exp_note = extend_country_rankings_from_unctad(
            items_exp, total_exp, years_exp,
            un_annual["country_s"].get("Exports", {}), "Exports",
            max_year=max_kenya_year,
            world_s=un_annual.get("world_s", {}).get("Exports"))
        if exp_note:
            print(exp_note)

    n_years = len(years_exp)
    shown_exp = rank_and_top(items_exp, top_n)
    ao_exp = all_other_vals(total_exp, shown_exp, n_years)
    t_latest_exp = total_exp["vals"][-1] if total_exp else None
    for it in shown_exp:
        it["share"] = share(it["vals"][-1], t_latest_exp)
    if total_exp:
        total_exp["share"] = share(t_latest_exp, t_latest_exp)
    ao_share_exp = share(ao_exp[-1], t_latest_exp) if ao_exp else None
    d1 = {"years": years_exp, "total": total_exp, "shown": shown_exp,
          "all_other": {"vals": ao_exp, "share": ao_share_exp}}

    # ---- Table 2: Global service importers by country --------------------
    total_imp, items_imp, years_imp = _load_first(files, "importers", parse_importers)
    for it in items_imp:
        it["vals"] = [v / 1e6 if v is not None else None for v in it["vals"]]
    if total_imp:
        total_imp["vals"] = [v / 1e6 if v is not None else None for v in total_imp["vals"]]

    if un_annual and un_annual.get("country_s"):
        items_imp, total_imp, years_imp, imp_note = extend_country_rankings_from_unctad(
            items_imp, total_imp, years_imp,
            un_annual["country_s"].get("Imports", {}), "Imports",
            max_year=max_kenya_year,
            world_s=un_annual.get("world_s", {}).get("Imports"))
        if imp_note:
            print(imp_note)

    n_years_i = len(years_imp)
    shown_imp = rank_and_top(items_imp, top_n)
    ao_imp = all_other_vals(total_imp, shown_imp, n_years_i)
    t_latest_imp = total_imp["vals"][-1] if total_imp else None
    for it in shown_imp:
        it["share"] = share(it["vals"][-1], t_latest_imp)
    if total_imp:
        total_imp["share"] = share(t_latest_imp, t_latest_imp)
    ao_share_imp = share(ao_imp[-1], t_latest_imp) if ao_imp else None
    d2 = {"years": years_imp, "total": total_imp, "shown": shown_imp,
          "all_other": {"vals": ao_imp, "share": ao_share_imp}}

    # ---- Global service exports by category (for pie chart) ---------------
    total_gexp, items_gexp, years_gexp = _load_first(
        files, "exported_services", parse_exported_services)
    for it in items_gexp:
        it["vals"] = [v / 1e6 if v is not None else None for v in it["vals"]]
    if total_gexp:
        total_gexp["vals"] = [v / 1e6 if v is not None else None for v in total_gexp["vals"]]

    if un_annual and un_annual.get("world"):
        items_gexp, total_gexp, years_gexp, gexp_note = extend_world_categories_from_unctad(
            items_gexp, total_gexp, years_gexp,
            un_annual["world"].get("Exports", {}), "exports",
            max_year=max_kenya_year)
        if gexp_note:
            print(gexp_note)

    # ---- Table 3: Kenya service exports by category ----------------------
    total_kexp, items_kexp, years_kexp = _load_first(
        files, "kenya_exports", parse_kenya_services)
    for it in items_kexp:
        it["vals"] = [v / 1e3 if v is not None else None for v in it["vals"]]
    if total_kexp:
        total_kexp["vals"] = [v / 1e3 if v is not None else None for v in total_kexp["vals"]]
    if un_annual and "kenya" in un_annual:
        # The ITC Kenya file only breaks categories out through 2023; 2024 is
        # aggregate (S) only.  Fill the 2024 category detail and total from the
        # real UNCTAD annual dataset, then cap the table at the latest genuinely
        # complete year (2024).  2025 never enters: it is an estimated total only.
        items_kexp, total_kexp, years_kexp, kexpl_note = extend_kenya_categories_from_unctad(
            items_kexp, total_kexp, years_kexp, un_annual["kenya"].get("Exports", {}),
            "exports", max_year=2024)
        if kexpl_note:
            print(kexpl_note)
    if un_annual:
        years_kexp = _truncate_to_category_year(items_kexp, total_kexp, years_kexp)
    calc_growth_rates(items_kexp, years_kexp)

    n_years_ke = len(years_kexp)
    shown_kexp = rank_and_top(items_kexp, len(items_kexp))  # show all service categories
    ao_kexp = all_other_vals(total_kexp, shown_kexp, n_years_ke)
    t_latest_kexp = total_kexp["vals"][-1] if total_kexp else None
    for it in shown_kexp:
        it["share"] = share(it["vals"][-1], t_latest_kexp)
    if total_kexp:
        total_kexp["share"] = share(t_latest_kexp, t_latest_kexp)
    ao_share_kexp = share(ao_kexp[-1], t_latest_kexp) if ao_kexp else None
    d3 = {"years": years_kexp, "total": total_kexp, "shown": shown_kexp,
          "all_other": {"vals": ao_kexp, "share": ao_share_kexp}}

    # ---- Table 4: Kenya service imports by category ----------------------
    total_kimp, items_kimp, years_kimp = _load_first(
        files, "kenya_imports", parse_kenya_services)
    for it in items_kimp:
        it["vals"] = [v / 1e3 if v is not None else None for v in it["vals"]]
    if total_kimp:
        total_kimp["vals"] = [v / 1e3 if v is not None else None for v in total_kimp["vals"]]
    if un_annual and "kenya" in un_annual:
        # Same as exports: fill the 2024 import category detail + total from the
        # real UNCTAD annual dataset, capped at the latest complete year (2024).
        items_kimp, total_kimp, years_kimp, kimpl_note = extend_kenya_categories_from_unctad(
            items_kimp, total_kimp, years_kimp, un_annual["kenya"].get("Imports", {}),
            "imports", max_year=2024)
        if kimpl_note:
            print(kimpl_note)
    if un_annual:
        years_kimp = _truncate_to_category_year(items_kimp, total_kimp, years_kimp)
    calc_growth_rates(items_kimp, years_kimp)

    n_years_ki = len(years_kimp)
    shown_kimp = rank_and_top(items_kimp, len(items_kimp))
    ao_kimp = all_other_vals(total_kimp, shown_kimp, n_years_ki)
    t_latest_kimp = total_kimp["vals"][-1] if total_kimp else None
    for it in shown_kimp:
        it["share"] = share(it["vals"][-1], t_latest_kimp)
    if total_kimp:
        total_kimp["share"] = share(t_latest_kimp, t_latest_kimp)
    ao_share_kimp = share(ao_kimp[-1], t_latest_kimp) if ao_kimp else None
    d4 = {"years": years_kimp, "total": total_kimp, "shown": shown_kimp,
          "all_other": {"vals": ao_kimp, "share": ao_share_kimp}}

    # ---- Figure 1: Kenya services balance of trade ----------------------
    # Use the Kenya export/import years (consistent set) for the balance
    bal_years = years_kexp
    n_bal = len(bal_years)
    # Pad exports/imports to same length
    exports_full = [v / 1e3 if v is not None else None for v in total_kexp["vals"]] if total_kexp else []
    imports_full = [v / 1e3 if v is not None else None for v in total_kimp["vals"]] if total_kimp else []
    exports_full += [None] * (n_bal - len(exports_full))
    imports_full += [None] * (n_bal - len(imports_full))

    # Cross-check the latest-year balance against the ITC commercialized
    # services file (list of services commercialized by Kenya).
    comm_note = None
    comm_path = files.get("kenya_commercialized")
    if comm_path:
        try:
            comm_loaded = _load_first(files, "kenya_commercialized",
                                      parse_kenya_commercialized)
            comm_total, _, comm_years = comm_loaded
        except Exception as exc:  # noqa: BLE001 - optional source
            print(f"Warning: ignoring unusable commercialized services file: {exc}")
            comm_total, comm_years = None, []
        if comm_total and comm_total.get("export_val") is not None and comm_total.get("import_val") is not None:
            comm_export = comm_total["export_val"] / 1e6
            comm_import = comm_total["import_val"] / 1e6
            comm_balance = comm_export - comm_import
            bal_export = exports_full[-1] if exports_full else None
            bal_import = imports_full[-1] if imports_full else None
            if bal_export is not None and bal_import is not None:
                diff = abs(bal_export - comm_export) + abs(bal_import - comm_import)
                match = "consistent" if diff < 0.05 else "differs"
                comm_note = (
                    f"Balance cross-checked against the ITC list of services commercialized by Kenya: "
                    f"{comm_years[-1]} exports USD {comm_export:,.2f} bn, imports USD {comm_import:,.2f} bn, "
                    f"balance USD {comm_balance:,.2f} bn (source {match})."
                )

    # ---- RCA: Kenya services revealed comparative advantage ---------------
    rca_items, rca_year = compute_services_rca(
        items_kexp, total_kexp, items_gexp, total_gexp, years_gexp)

    # ---- Development status classification ------------------------------
    classification = load_development_classification()
    calc_growth_rates(items_exp, years_exp)
    calc_growth_rates(items_imp, years_imp)

    # Add region data to items
    for it in items_exp:
        it["region"] = get_region(it.get("label", ""))
    for it in items_imp:
        it["region"] = get_region(it.get("label", ""))

    # Top exporters by development status
    exp_dev, exp_devel, exp_ldc = rank_by_dev_status(
        [it for it in items_exp if it.get("label", "").lower() != "world"],
        years_exp, classification, top_n=10)
    # Top importers by development status
    imp_dev, imp_devel, imp_ldc = rank_by_dev_status(
        [it for it in items_imp if it.get("label", "").lower() != "world"],
        years_imp, classification, top_n=10)

    # Top exporters by region (top 5 per region)
    exp_by_region = rank_by_region(
        [it for it in items_exp if it.get("label", "").lower() != "world"],
        years_exp, top_n=5)
    # Top importers by region (top 5 per region)
    imp_by_region = rank_by_region(
        [it for it in items_imp if it.get("label", "").lower() != "world"],
        years_imp, top_n=5)

    # ---- Write Excel files -----------------------------------------------
    out = {}

    # Table 1
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 1"
    cache = write_market_table(ws, d1, "Exporters",
                       f"List of Top Service Exporters\nValue in USD Billion",
                       unit_row=True, kenya_highlight=False,
                       row1_label="Table 1:",
                       row1_title="Top Global Service Exporters")
    set_widths(ws, {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12, "J": 12})
    out["t1"] = os.path.join(out_dir, "Table 1 Top Global Service Exporters.xlsx")
    _finalize(ws, out["t1"], cache)

    # Table 2
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 2"
    cache = write_market_table(ws, d2, "Importers",
                       f"List of Top Service Importers\nValue in USD Billion",
                       unit_row=True, kenya_highlight=False,
                       row1_label="Table 2:",
                       row1_title="Top Global Service Importers")
    set_widths(ws, {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12, "J": 12})
    out["t2"] = os.path.join(out_dir, "Table 2 Top Global Service Importers.xlsx")
    _finalize(ws, out["t2"], cache)

    # Table 3
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 3"
    cache = write_product_table(ws, d3,
                        "Kenya's Service Exports by Category\nValue in USD Million",
                        unit_row="Value in USD Million",
                        row1_label="Table 3",
                        row1_title="Kenya's Service Exports by Category")
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12, "J": 12, "K": 12})
    out["t3"] = os.path.join(out_dir, "Table 3 Kenya Service Exports.xlsx")
    _finalize(ws, out["t3"], cache)

    # Table 4
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 4"
    cache = write_product_table(ws, d4,
                        "Kenya's Service Imports by Category\nValue in USD Million",
                        unit_row="Value in USD Million",
                        row1_label="Table 4",
                        row1_title="Kenya's Service Imports by Category")
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12, "J": 12, "K": 12})
    out["t4"] = os.path.join(out_dir, "Table 4 Kenya Service Imports.xlsx")
    _finalize(ws, out["t4"], cache)

    # Table 5: Top Service Exporters by Development Status
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 5"
    cache = []
    write_dev_status_table(ws, exp_dev, exp_devel, exp_ldc, years_exp,
                           "Exporters", "Value in USD Billion",
                           row1_label="Table 5:",
                           row1_title="Top Global Service Exporters by Development Status")
    out["t5"] = os.path.join(out_dir, "Table 5 Service Exporters by Dev Status.xlsx")
    _finalize(ws, out["t5"], cache)

    # Table 6: Top Service Importers by Development Status
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 6"
    cache = []
    write_dev_status_table(ws, imp_dev, imp_devel, imp_ldc, years_imp,
                           "Importers", "Value in USD Billion",
                           row1_label="Table 6:",
                           row1_title="Top Global Service Importers by Development Status")
    out["t6"] = os.path.join(out_dir, "Table 6 Service Importers by Dev Status.xlsx")
    _finalize(ws, out["t6"], cache)

    # Table 7: Top Service Exporters by Region
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 7"
    cache = []
    write_region_table(ws, exp_by_region, years_exp,
                       "Exporters", "Value in USD Billion",
                       row1_label="Table 7:",
                       row1_title="Top Service Exporters by Region")
    out["t7"] = os.path.join(out_dir, "Table 7 Service Exporters by Region.xlsx")
    _finalize(ws, out["t7"], cache)

    # Table 8: Top Service Importers by Region
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 8"
    cache = []
    write_region_table(ws, imp_by_region, years_imp,
                       "Importers", "Value in USD Billion",
                       row1_label="Table 8:",
                       row1_title="Top Service Importers by Region")
    out["t8"] = os.path.join(out_dir, "Table 8 Service Importers by Region.xlsx")
    _finalize(ws, out["t8"], cache)

    # Table 9: Kenya's Revealed Comparative Advantage in Services
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 9"
    cache = []
    write_rca_table(ws, rca_items, [rca_year],
                    row1_label="Table 9:",
                    row1_title="Kenya's Revealed Comparative Advantage (RCA) in Services")
    out["t9_rca"] = os.path.join(out_dir, "Table 9 Kenya Services RCA.xlsx")
    _finalize(ws, out["t9_rca"], cache)

    # Figure 6: RCA bar chart
    rca_chart_path = os.path.join(out_dir, "Figure 6 Kenya Services RCA.png")
    try:
        make_rca_radar_chart(rca_items, [rca_year], rca_chart_path)
        out["rca_chart"] = rca_chart_path
    except Exception as e:
        print(f"Warning: Could not create RCA chart: {e}")

    # ---- Table 10: Export Concentration Index (HHI) -----------------------
    concentration = compute_concentration_index(
        items_kexp, total_kexp, items_gexp, total_gexp, years_kexp)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 10"
    cache = []
    write_concentration_table(ws, concentration,
                              row1_label="Table 10:",
                              row1_title="Export Concentration Index (HHI)")
    out["t10_conc"] = os.path.join(out_dir, "Table 10 Export Concentration Index.xlsx")
    _finalize(ws, out["t10_conc"], cache)

    # Figure 7: Concentration comparison chart
    conc_chart_path = os.path.join(out_dir, "Figure 7 Concentration Comparison.png")
    try:
        make_concentration_chart(concentration, conc_chart_path)
        out["conc_chart"] = conc_chart_path
    except Exception as e:
        print(f"Warning: Could not create concentration chart: {e}")

    # ---- Table 11: Diversification Potential -----------------------------
    div_items = compute_diversification_potential(
        items_kexp, total_kexp, items_gexp, total_gexp,
        items_kimp, total_kimp, years_kexp)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 11"
    cache = []
    write_diversification_table(ws, div_items,
                                row1_label="Table 11:",
                                row1_title="Diversification Potential in Services")
    out["t11_div"] = os.path.join(out_dir, "Table 11 Diversification Potential.xlsx")
    _finalize(ws, out["t11_div"], cache)

    # Figure 8: Diversification chart
    div_chart_path = os.path.join(out_dir, "Figure 8 Diversification Opportunities.png")
    try:
        make_diversification_chart(div_items, div_chart_path)
        out["div_chart"] = div_chart_path
    except Exception as e:
        print(f"Warning: Could not create diversification chart: {e}")

    # ---- Table 12: Value Composition Trajectory --------------------------
    trajectory = compute_value_trajectory(items_kexp, total_kexp, years_kexp)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 12"
    cache = []
    write_value_trajectory_table(ws, trajectory,
                                 row1_label="Table 12:",
                                 row1_title="Service Export Composition Trajectory")
    out["t12_traj"] = os.path.join(out_dir, "Table 12 Value Composition Trajectory.xlsx")
    _finalize(ws, out["t12_traj"], cache)

    # Figure 9: Value composition chart
    traj_chart_path = os.path.join(out_dir, "Figure 9 Value Composition Trajectory.png")
    try:
        make_value_trajectory_chart(trajectory, traj_chart_path)
        out["traj_chart"] = traj_chart_path
    except Exception as e:
        print(f"Warning: Could not create trajectory chart: {e}")

    # ---- Table 13: Kenya vs Peer Total Exports ---------------------------
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 13"
    cache = []
    write_peer_comparison_table(ws, items_exp, years_exp,
                                row1_label="Table 13:",
                                row1_title="Kenya vs Peer Countries: Service Exports")
    out["t13_peer"] = os.path.join(out_dir, "Table 13 Kenya vs Peers Service Exports.xlsx")
    _finalize(ws, out["t13_peer"], cache)

    # Figure 1: Kenya Services Balance
    bal_path = os.path.join(out_dir, "Figure 1 Services Balance.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Figure 1"
    cache = []
    write_balance(ws, bal_years, exports_full, imports_full, cache=cache)
    set_widths(ws, BALANCE_WIDTHS)
    if comm_note:
        note_row = 8
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2 + n_bal)
        cell = ws.cell(note_row, 1, comm_note)
        cell.font = Font(name=FONT, size=9, italic=True, color=FONT_GRAY if "FONT_GRAY" in globals() else "595959")
        ws.row_dimensions[note_row].height = 28
    out["balance"] = bal_path
    _finalize(ws, out["balance"], cache)

    # Figure 2: Pie chart of global service exports by category
    pie_path = os.path.join(out_dir, "Figure 2 Service Exports Structure.png")
    try:
        make_pie_chart(items_gexp, years_gexp, pie_path)
        out["pie"] = pie_path
    except Exception as e:
        print(f"Warning: Could not create pie chart: {e}")

    # Figure 3: Stacked bar chart of service exports by category
    bar_path = os.path.join(out_dir, "Figure 3 Service Exports by Category.png")
    try:
        make_stacked_bar_chart(items_gexp, years_gexp, bar_path)
        out["stacked_bar"] = bar_path
    except Exception as e:
        print(f"Warning: Could not create stacked bar chart: {e}")

    # Figure 4: Regional bar chart of exporters
    reg_exp_path = os.path.join(out_dir, "Figure 4 Service Exports by Region.png")
    try:
        make_regional_bar_chart(exp_by_region, years_exp, reg_exp_path,
                                title="Top Service Exporters by Region")
        out["reg_exp_chart"] = reg_exp_path
    except Exception as e:
        print(f"Warning: Could not create regional exporter chart: {e}")

    # Figure 5: Regional bar chart of importers
    reg_imp_path = os.path.join(out_dir, "Figure 5 Service Imports by Region.png")
    try:
        make_regional_bar_chart(imp_by_region, years_imp, reg_imp_path,
                                title="Top Service Importers by Region")
        out["reg_imp_chart"] = reg_imp_path
    except Exception as e:
        print(f"Warning: Could not create regional importer chart: {e}")

    # ---- Combined workbook: all sheets in one file -----------------------
    wb_all = openpyxl.Workbook()
    wb_all.remove(wb_all.active)
    widths_mkt = {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12, "J": 12}
    widths_prd = {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12}
    cache_all = {}

    for s in ("Table 1", "Table 2", "Table 3", "Table 4", "Table 5", "Table 6",
              "Table 7", "Table 8", "Table 9", "Table 10", "Table 11", "Table 12",
              "Table 13", "Figure 1"):
        wb_all.create_sheet(s)

    ws = wb_all["Table 1"]; c = write_market_table(ws, d1, "Exporters",
                       "List of Top Service Exporters\nValue in USD Billion",
                       unit_row=True, kenya_highlight=False,
                       row1_label="Table 1:",
                       row1_title="Top Global Service Exporters")
    set_widths(ws, widths_mkt); cache_all["Table 1"] = dict(c)

    ws = wb_all["Table 2"]; c = write_market_table(ws, d2, "Importers",
                       "List of Top Service Importers\nValue in USD Billion",
                       unit_row=True, kenya_highlight=False,
                       row1_label="Table 2:",
                       row1_title="Top Global Service Importers")
    set_widths(ws, widths_mkt); cache_all["Table 2"] = dict(c)

    ws = wb_all["Table 3"]; c = write_product_table(ws, d3,
                        "Kenya's Service Exports by Category\nValue in USD Million",
                        unit_row="Value in USD Million",
                        row1_label="Table 3",
                        row1_title="Kenya's Service Exports by Category")
    set_widths(ws, widths_prd); cache_all["Table 3"] = dict(c)

    ws = wb_all["Table 4"]; c = write_product_table(ws, d4,
                        "Kenya's Service Imports by Category\nValue in USD Million",
                        unit_row="Value in USD Million",
                        row1_label="Table 4",
                        row1_title="Kenya's Service Imports by Category")
    set_widths(ws, widths_prd); cache_all["Table 4"] = dict(c)

    ws = wb_all["Table 5"]; c = []
    write_dev_status_table(ws, exp_dev, exp_devel, exp_ldc, years_exp,
                           "Exporters", "Value in USD Billion",
                           row1_label="Table 5:",
                           row1_title="Top Global Service Exporters by Development Status")
    cache_all["Table 5"] = dict(c)

    ws = wb_all["Table 6"]; c = []
    write_dev_status_table(ws, imp_dev, imp_devel, imp_ldc, years_imp,
                           "Importers", "Value in USD Billion",
                           row1_label="Table 6:",
                           row1_title="Top Global Service Importers by Development Status")
    cache_all["Table 6"] = dict(c)

    ws = wb_all["Table 7"]; c = []
    write_region_table(ws, exp_by_region, years_exp,
                       "Exporters", "Value in USD Billion",
                       row1_label="Table 7:",
                       row1_title="Top Service Exporters by Region")
    cache_all["Table 7"] = dict(c)

    ws = wb_all["Table 8"]; c = []
    write_region_table(ws, imp_by_region, years_imp,
                       "Importers", "Value in USD Billion",
                       row1_label="Table 8:",
                       row1_title="Top Service Importers by Region")
    cache_all["Table 8"] = dict(c)

    ws = wb_all["Table 9"]; c = []
    write_rca_table(ws, rca_items, [rca_year],
                    row1_label="Table 9:",
                    row1_title="Kenya's Revealed Comparative Advantage (RCA) in Services")
    cache_all["Table 9"] = dict(c)

    ws = wb_all["Table 10"]; c = []
    write_concentration_table(ws, concentration,
                              row1_label="Table 10:",
                              row1_title="Export Concentration Index (HHI)")
    cache_all["Table 10"] = dict(c)

    ws = wb_all["Table 11"]; c = []
    write_diversification_table(ws, div_items,
                                row1_label="Table 11:",
                                row1_title="Diversification Potential in Services")
    cache_all["Table 11"] = dict(c)

    ws = wb_all["Table 12"]; c = []
    write_value_trajectory_table(ws, trajectory,
                                 row1_label="Table 12:",
                                 row1_title="Service Export Composition Trajectory")
    cache_all["Table 12"] = dict(c)

    ws = wb_all["Table 13"]; c = []
    write_peer_comparison_table(ws, items_exp, years_exp,
                                row1_label="Table 13:",
                                row1_title="Kenya vs Peer Countries: Service Exports")
    cache_all["Table 13"] = dict(c)

    ws = wb_all["Figure 1"]; c = []
    write_balance(ws, bal_years, exports_full, imports_full, cache=c)
    set_widths(ws, BALANCE_WIDTHS); cache_all["Figure 1"] = dict(c)

    out["all"] = os.path.join(out_dir, "All Service Tables.xlsx")
    try:
        wb_all.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb_all.save(out["all"])
    try:
        _inject_cached_values(out["all"], cache_all)
    except Exception:
        pass

    # Ship the UNCTAD R-export CSV alongside the ITC tables so the packaged
    # "services" deliverable carries both the ITC analysis and the raw UNCTAD
    # source file.  (The analyst keeps the companion *.xlsx locally.)
    _copy_unctad_source_files(files, out_dir)

    # Extract country name from Kenya exports title
    rep = "Kenya"
    partner = "the World"

    return out, (rep, partner)


def main():
    ap = argparse.ArgumentParser(
        description="Build service Table 1-4 + Figure 1 from raw ITC source files.")
    ap.add_argument("--excel-dir", default="services",
                    help="Folder with the seven raw ITC service files (default: services)")
    ap.add_argument("--out-dir", default=os.path.join("output", "service_tables"),
                    help="Where to write the generated tables (default: output/service_tables)")
    ap.add_argument("--top", type=int, default=20,
                    help="Number of top exporter/importer rows to keep (default: 20)")
    args = ap.parse_args()

    if args.top < 1:
        sys.exit("[ERROR] --top must be >= 1")

    try:
        out, (rep, partner) = generate_service_tables(args.excel_dir, args.out_dir, args.top)
    except SystemExit:
        raise
    except Exception as e:
        sys.exit(f"[ERROR] Failed to generate service tables: {e}")

    print("[1/2] Read raw service source files from :", os.path.abspath(args.excel_dir))
    print("[2/2] Wrote service tables for            :", rep, "<->", partner)
    for key in ("t1", "t2", "t3", "t4", "balance", "all"):
        print(f"      - {os.path.basename(out[key])}")
    print()
    print("Next step: run")
    print(f'  python generate_services_report.py --excel-dir "{os.path.abspath(args.out_dir)}"')


if __name__ == "__main__":
    main()
