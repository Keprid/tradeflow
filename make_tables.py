#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_tables.py
==============

Turn the raw ITC (International Trade Centre) Excel downloads for Kenya's
trade with any partner country into the six "Table 1 .. Table 6" workbooks
that the report generator (`generate_report.py`) expects.

Raw source files expected in ``--excel-dir`` (located by keywords in the
filename, case-insensitive; all Excel formats accepted: .xls, .xlsx, .xlsm,
.xlsb, .csv):

    <country>s-imports-from-world-by-exporter_all.xlsx   -> Table 1  (import source markets)
    <country>s-imports-from-world-by-product_all.xlsx    -> Table 2  (import products)
    <country>s-exports-to-world-by-importer_all.xlsx     -> Table 3  (export destinations)
    <country>s-exports-to-world-by-product_all.xlsx      -> Table 4  (export products)
    kenyas-exports-to-<country>-by-product_all.xlsx      -> Table 5  (Kenya exports to the country)
    kenyas-imports-from-<country>-by-product_all.xlsx    -> Table 6  (Kenya imports from the country)

Data manipulation (mirrors how the reference tables were prepared):

* Source values are in USD Thousand.
* Tables 1-4  : divide by 1,000,000  ->  "Value in USD Billion".
* Tables 5-6  : divide by 1,000      ->  "Value in USD Million".
* The first column ranks rows by the latest year, highest first.
* Market tables (Tables 1 and 3) always keep the top 25 exporters/partners;
  product tables (Tables 2, 4, 5, 6) keep the top ``--top`` rows
  (default 20).
* For the partner-country tables (1 and 3): if Kenya is not inside the top N
  it is still included with its actual rank and the whole row is highlighted.
* The values of every row not shown are summed into an
  "All other countries/products" row; the last row is the World / All products
  total.
* The final column is the share of the latest-year total
  (row value / total), shown as a percentage.

Runs on Windows, Linux and macOS (Python 3.9+, openpyxl only).

Usage
-----
    python make_tables.py --excel-dir sourcefiles --out-dir output/tables

Then feed the generated folder to ``generate_report.py``. The script also
writes a ``Figure 1 Trade Balance.xlsx`` derived from the Table 5 / Table 6
totals (Kenya's exports minus imports with the partner country).
"""

import argparse
import os
import re
import shutil
import sys
import tempfile
import zipfile

import openpyxl
from lxml import etree
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from country_names import display_name, short_product_name
from xlsx_compat import convert_to_xlsx, is_spreadsheet

# ---------------------------------------------------------------------------
# Constants / styles (matching make_services_tables.py)
# ---------------------------------------------------------------------------
# Professional color palette (muted, modern tones)
FONT = "Century Gothic"
FONT_SIZE = 10
HEADER_FONT_SIZE = 10

TOP_MARKETS = 25

# Header styles - deep navy with white text
HDR_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
HDR_FONT_COLOR = "FFFFFF"

# Alternating row fills (very subtle gray tones)
ROW_FILL_EVEN = PatternFill(fill_type="solid", fgColor="F2F2F2")
ROW_FILL_ODD = None  # No fill for odd rows (white)

# Kenya highlight (soft gold)
KENYA_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")

# Band/summary row (light blue-gray)
BAND_FILL = PatternFill(fill_type="solid", fgColor="D6DCE4")

# Border styles - thin bottom only for headers, no side borders
THIN = Side(style="thin", color="B4C6E7")  # Light blue-gray
MEDIUM = Side(style="medium", color="1F3864")  # Navy accent
HEADER_BORDER = Border(bottom=MEDIUM)
DATA_BORDER = Border(bottom=Side(style="thin", color="D6DCE4"))
TOTAL_BORDER = Border(top=Side(style="thin", color="808080"),
                      bottom=Side(style="double", color="808080"))

# Legacy border (kept for compatibility)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Number formats
FMT_VALUE = "#,##0.0"
FMT_VALUE_INT = "#,##0"
FMT_SHARE = "0.0%"
FMT_GROWTH = "0.0%"

# Column widths for Figure 1 (label + up to ten year columns).
BALANCE_WIDTHS = {"A": 16, "B": 30, "C": 12, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12, "J": 12, "K": 12}


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def to_float(v):
    """Convert a raw cell value to float; None for blanks / '...' / garbage."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", "")
        if v in ("", "...", "..", "-"):
            return None
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def truncate_label(label, limit=80):
    """Shorten a long product label for the product tables.

    Labels are long HS descriptions built from ``;``-separated clauses. The
    result keeps whole clauses as long as they fit within ``limit`` chars and
    never appends a trailing ellipsis. Any trailing dots/punctuation (which
    the raw ITC download sometimes leaves on a cut description) are removed so
    every label reads as a clean, complete sentence.
    """
    if not label:
        return ""
    label = str(label).strip().rstrip(".,;: ")
    if not label:
        return ""
    parts = re.split(r";\s*", label)
    out = parts[0]
    for p in parts[1:]:
        candidate = out + "; " + p
        if len(candidate) <= limit:
            out = candidate
        else:
            break
    if len(out) > limit:
        head = out[:limit]
        cut = head.rsplit(" ", 1)[0]
        out = (cut or head).rstrip(".,;: ")
    return out


def _hs_granularity(path, limit=400):
    """Probe a Trade Map workbook and return the granularity of its product
    codes (2, 4, 6 or None when not a product table / undetectable).

    The product-code column is located by header text (`code`, `product code`,
    `hs code` -- but *not* reporter/partner/country codes), then the leading
    digit-strings of its data cells are sampled and the modal length returned.
    Trade Map ships product tables at HS2 by default, so callers use this to
    prefer the HS4 copy when several versions of the same table are present.
    """
    try:
        wb, _used = convert_to_xlsx(path)
        ws = wb[wb.sheetnames[0]]
        grid = [[_cell_text(c) for c in row]
                for row in ws.iter_rows(values_only=True, max_row=120)]
        wb.close()
    except Exception:
        return None

    if not grid:
        return None
    header = grid[0]
    code_col = None
    for i, h in enumerate(header):
        hl = h.lower().replace(" ", "_")
        # product code column: mentions a code ('cd'/'code'/'product code') but
        # is not a partner/reporter/market one
        if ("cd" in hl or "code" in hl) and not any(k in hl for k in
                                                    ("reporter", "partner", "country",
                                                     "importer", "exporter", "market",
                                                     "destination", "origin", "supplier")):
            code_col = i
            break
    # fallback: classic layouts often name it "Product code" or just "Code"
    if code_col is None:
        for i, h in enumerate(header):
            if h.lower().replace(" ", "_") in ("product_code", "code", "hs_code",
                                               "harmonized_system_code", "productcd"):
                code_col = i
                break
    if code_col is None:
        return None

    lengths = []
    for row in grid[1:]:
        if code_col >= len(row):
            continue
        raw = row[code_col].lstrip("'").replace("\u00a0", " ").strip()
        if raw.isdigit():
            lengths.append(len(raw))
    if not lengths:
        return None
    # modal non-zero length
    from collections import Counter
    counts = Counter(l for l in lengths if l > 0)
    if not counts:
        return None
    return counts.most_common(1)[0][0]


_PRODUCT_ROLES = ("table2", "table4", "table5", "table6")


def _prefer_hs4(candidates, excel_dir):
    """Given a list of candidate filenames for a product-table role, prefer
    the HS4 copy when several are present (Trade Map offers HS2 by default but
    users often also download HS4).  Falls back to first alphabetic match."""
    if len(candidates) < 2:
        return candidates[0]
    scanned = []
    for fname in candidates:
        lvl = _hs_granularity(os.path.join(excel_dir, fname))
        scanned.append((lvl, fname))
    best = None
    for lvl, fname in scanned:
        if lvl == 4:
            return fname                       # exact HS4 match wins
        if lvl is not None:
            if best is None or lvl > best[0]:
                best = (lvl, fname)
    return best[1] if best else candidates[0]


def find_source_files(excel_dir):
    """Locate the six raw source files by filename keywords.

    Both Trade Map download styles are recognised:

    * **beta** slugs -- ``kenyas-exports-to-malaysia-by-product_all.xlsx``,
      ``malaysias-imports-from-world-by-product_all.xlsx`` ...
    * **classic ("previous") names** -- ``Trade_Map_-_List_of_supplying_
      markets_for_a_product_imported_by_Malaysia.xls``,
      ``Trade_Map_-_List_of_products_imported_by_Malaysia.xls``,
      ``Trade_Map_-_List_of_importing_markets_for_a_product_exported_by_...``,
      ``Trade_Map_-_List_of_products_exported_by_Malaysia.xls`` and the two
      ``Trade_Map_-_Bilateral_trade_between_Kenya_and_<Partner>`` workbooks.
      The bilateral pair share one filename (browser " (1)" copies), so they
      are told apart by peeking at the trade-flow title inside each file.
    """
    found = {}
    _cands = {r: [] for r in _PRODUCT_ROLES}
    bilaterals = []
    for fname in sorted(os.listdir(excel_dir)):
        if not is_spreadsheet(fname):
            continue
        low = fname.lower().replace(" ", "_")
        if "bilateral_trade_between_kenya_and" in low:
            bilaterals.append(fname)
        elif "list_of_supplying_markets_for_a_product_imported" in low \
                or ("imports-from-world" in low and "by-exporter" in low):
            found.setdefault("table1", fname)
        elif "list_of_products_imported" in low \
                or ("imports-from-world" in low and "by-product" in low):
            _cands["table2"].append(fname)
        elif "list_of_importing_markets_for_a_product_exported" in low \
                or ("exports-to-world" in low and "by-importer" in low):
            found.setdefault("table3", fname)
        elif "list_of_products_exported" in low \
                or ("exports-to-world" in low and "by-product" in low):
            _cands["table4"].append(fname)
        elif "imports-from" in low and "by-product" in low:
            _cands["table6"].append(fname)
        elif "exports-to" in low and "by-product" in low:
            _cands["table5"].append(fname)

    # Classic bilateral downloads: same filename for both directions, so
    # classify each by the flow named inside the workbook.
    for fname in bilaterals:
        try:
            kind = _bilateral_direction(os.path.join(excel_dir, fname))
        except Exception:
            continue
        if kind == "exports":
            _cands["table5"].append(fname)
        elif kind == "imports":
            _cands["table6"].append(fname)

    # Product tables: when several versions are present (e.g. HS2 *and* HS4 of
    # the same Tabel Map download), record the HS4 copy.
    for role in _PRODUCT_ROLES:
        if _cands[role]:
            found[role] = _prefer_hs4(_cands[role], excel_dir)

    missing = [k for k in ("table1", "table2", "table3", "table4", "table5", "table6")
               if k not in found]
    if missing:
        expected = {
            "table1": "imports-from-world-by-exporter (or 'List of supplying "
                      "markets for a product imported by <Partner>')",
            "table2": "imports-from-world-by-product (or 'List of products "
                      "imported by <Partner>')",
            "table3": "exports-to-world-by-importer (or 'List of importing "
                      "markets for a product exported by <Partner>')",
            "table4": "exports-to-world-by-product (or 'List of products "
                      "exported by <Partner>')",
            "table5": "kenyas-exports-to-*-by-product (or 'Bilateral trade "
                      "between Kenya and <Partner>' holding Kenya's exports)",
            "table6": "kenyas-imports-from-*-by-product (or 'Bilateral trade "
                      "between Kenya and <Partner>' holding Kenya's imports)",
        }
        missing_detail = []
        for k in missing:
            missing_detail.append(f"  - {k}: expected filename containing '{expected[k]}'")
        sys.exit(
            "[ERROR] Missing required source files in '%s'.\n"
            "Found %d file(s): %s\n"
            "Missing %d required file(s):\n%s"
            % (excel_dir, len(found),
               ", ".join(f"{k}={v}" for k, v in sorted(found.items())),
               len(missing), "\n".join(missing_detail)))
    return {k: os.path.join(excel_dir, v) for k, v in found.items()}


def _bilateral_direction(path):
    """Classify a classic 'Bilateral trade between A and B' workbook.

    Returns ``'exports'`` when it holds Kenya's exports to the partner,
    ``'imports'`` when it holds Kenya's imports from the partner, else
    ``None``. The first non-World flow title in the header row decides.
    """
    tmpdir = tempfile.mkdtemp(prefix="tradeflow_peek_")
    try:
        wb, _ = convert_to_xlsx(path, out_dir=tmpdir)
        try:
            ws = wb[wb.sheetnames[0]]
            titles = next(ws.iter_rows(max_row=1, values_only=True))
        finally:
            wb.close()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    for t in titles:
        tl = str(t or "").strip().lower()
        if not tl or "world" in tl:
            continue
        if re.search(r"exports?\s+to", tl):
            return "exports"
        if re.search(r"imports?\s+from", tl):
            return "imports"
    return None


# ---------------------------------------------------------------------------
# Source parsing
# ---------------------------------------------------------------------------
_YEAR_RE = re.compile(r"(?:19|20)\d{2}")
_VALUE_IN_RE = re.compile(r"value\s+in\s+((?:19|20)\d{2})", re.IGNORECASE)


def _cell_text(v):
    return str(v).strip() if v is not None else ""


def _clean_code(v):
    """Strip the leading apostrophe / stray whitespace Trade Map puts on
    product codes ('1511 -> 1511)."""
    return _cell_text(v).lstrip("'").replace("\u00a0", " ").strip()


def _select_year_cols(ycols, years):
    """Apply the caller's year restriction (or auto-pick when None)."""
    if years is None:
        return _pick_year_cols(ycols)
    allowed = set(years)
    return [(i, y) for i, y in ycols if y in allowed]


def _norm_ycols(chosen):
    """Year columns of a normalised row: values start right after the six
    beta metadata columns."""
    return [(6 + k, y) for k, (_, y) in enumerate(chosen)]


def _country_from_download_name(path):
    """Best-effort country label from a Trade Map download filename.

    '..._by_Malaysia.xls' -> 'Malaysia'; '..._between_Kenya_and_Malaysia'
    -> 'Malaysia'. Browser '(1)' copies and the beta '_all' suffix are
    ignored; multi-word slugs become Title Case before the SHORT_NAMES
    lookup in display_name.
    """
    stem = os.path.splitext(os.path.basename(str(path)))[0].lower()
    stem = stem.replace(" ", "_")
    stem = re.sub(r"_*\(\d+\)$", "", stem)
    stem = re.sub(r"_all$", "", stem)
    m = re.search(r"(?:^|_)by_([a-z0-9_]+)$", stem) \
        or re.search(r"_and_([a-z0-9_]+)$", stem)
    if not m:
        return ""
    return display_name(m.group(1).replace("_", " ").strip().title())


def parse_source(path, years=None):
    """Read an ITC workbook (either Trade Map layout).

    Returns ``(rows, year_columns, years, labels)`` where rows/ycols are
    normalised to the beta shape -- six leading metadata columns
    (reporterCd, reporterLabel, partnerCd, partnerLabel, code, label)
    followed by one column per selected year -- so ``extract_markets`` /
    ``extract_products`` / ``extract_kenya`` stay layout-agnostic.

    ``year_columns`` is a list of (column_index, year); when ``years`` is
    given only those years are kept (all tables in one report must share
    the same year set), otherwise they are chosen via _pick_years.
    """
    wb, _used = convert_to_xlsx(path)
    ws = wb[wb.sheetnames[0]]
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    kind = _detect_classic_layout(grid)
    if kind == "markets":
        return _parse_classic_markets(grid, path, years)
    if kind == "products":
        return _parse_classic_products(grid, path, years)
    if kind == "bilateral":
        return _parse_classic_bilateral(grid, path, years)

    # ---- beta layout (flat header: reporterCd ... productLabel, years) ----
    header = [_cell_text(c) for c in grid[0]]
    ycols = []
    for i, h in enumerate(header):
        m = re.match(r"^(\d{4})", h)
        if m:
            ycols.append((i, int(m.group(1))))
    ycols.sort(key=lambda t: t[1])
    chosen = _select_year_cols(ycols, years)
    years_out = [y for _, y in chosen]

    rows = [r for r in grid[1:]
            if r and any(c is not None and str(c).strip() != "" for c in r)]
    labels = {
        "reporter": str(grid[1][1]).strip() if len(grid) > 1 and grid[1][1] else "",
        "partner": str(grid[1][3]).strip() if len(grid) > 1 and grid[1][3] else "",
    }
    return rows, chosen, years_out, labels


def _detect_classic_layout(grid):
    """Classify a classic ('previous version') Trade Map sheet.

    'markets'   -- single header 'Exporters/Importers | value in YYYY ...'
    'products'  -- single header 'Code | Product label | value in YYYY ...'
    'bilateral' -- two-row header ('Product code' over 'Value in YYYY'),
                   several side-by-side data blocks
    None        -- not recognisably classic (assume the beta layout)
    """
    if not grid or not grid[0]:
        return None
    first = _cell_text(grid[0][0]).lower().lstrip("'")
    if first in ("exporters", "importers"):
        return "markets"
    if first in ("code", "product code"):
        second = [_cell_text(c).lower() for c in grid[1]] if len(grid) > 1 else []
        if any(_VALUE_IN_RE.match(t) for t in second):
            return "bilateral"
        return "products"
    return None


def _classic_year_cols(header_row):
    """(col, year) pairs from header cells like 'Imported value in 2021'."""
    ycols = []
    for i, c in enumerate(header_row):
        m = _YEAR_RE.search(_cell_text(c))
        if m:
            ycols.append((i, int(m.group(0))))
    ycols.sort(key=lambda t: t[1])
    return ycols


def _parse_classic_markets(grid, path, years):
    """Classic Table 1/3 source: one column of market names + year columns;
    the 'World' row carries the total."""
    chosen = _select_year_cols(_classic_year_cols(grid[0]), years)
    partner = _country_from_download_name(path)

    rows = []
    for r in grid[1:]:
        if not any(_cell_text(c) for c in r):
            continue
        label = _cell_text(r[0])
        vals = [r[i] if i < len(r) else None for i, _ in chosen]
        cd = "000" if label.lower() == "world" else "---"
        rows.append(["", "", cd, label, "TOTAL", label] + vals)

    labels = {"reporter": partner or "Partner", "partner": "World"}
    return rows, _norm_ycols(chosen), [y for _, y in chosen], labels


def _parse_classic_products(grid, path, years):
    """Classic Table 2/4 source: Code | Product label | value columns;
    the 'TOTAL' code row carries the world total."""
    chosen = _select_year_cols(_classic_year_cols(grid[0]), years)
    partner = _country_from_download_name(path)

    rows = []
    for r in grid[1:]:
        if not any(_cell_text(c) for c in r):
            continue
        code = _clean_code(r[0])
        label = _cell_text(r[1])
        vals = [r[i] if i < len(r) else None for i, _ in chosen]
        rows.append(["", "", "000", "", code, label] + vals)

    labels = {"reporter": partner or "Partner", "partner": "World"}
    return rows, _norm_ycols(chosen), [y for _, y in chosen], labels


def _parse_classic_bilateral(grid, path, years):
    """Classic bilateral source ('Trade Map - Bilateral trade between A and
    B'): a two-row header whose sub-row reads 'Value in YYYY' once per data
    block. Block 0 holds Kenya<->partner trade by product (Tables 5/6);
    later blocks repeat the partners' world-trade series and are ignored.

    The 'Value in YYYY' sub-row sits two columns left of its data (the
    'Product code'/'Product label' headers only span the first header row),
    so year columns are re-anchored onto the first data row before use.
    """
    hdr_idx = None
    for ri, row in enumerate(grid[:6]):
        if any(_VALUE_IN_RE.search(_cell_text(c)) for c in row):
            hdr_idx = ri
            break
    if hdr_idx is None:
        raise ValueError(f"No 'Value in <year>' header row found in {path}")

    sub_row = grid[hdr_idx]
    sub_years = {}
    for i, c in enumerate(sub_row):
        m = _VALUE_IN_RE.search(_cell_text(c))
        if m:
            sub_years[i] = int(m.group(1))

    # first non-empty data row anchors the real value columns
    data_row = None
    for r in grid[hdr_idx + 1:]:
        if any(_cell_text(c) for c in r):
            data_row = r
            break
    if data_row is None:
        raise ValueError(f"No data rows under the header of {path}")
    data_cols = [i for i, c in enumerate(data_row)
                 if i >= 2 and _cell_text(c) != ""]
    sub_blocks = _column_runs(sorted(sub_years))
    data_blocks = _column_runs(data_cols)
    if len(sub_blocks) != len(data_blocks) \
            or any(len(s) != len(d)
                   for s, d in zip(sub_blocks, data_blocks)):
        raise ValueError(f"Mismatched header/data blocks in {path}")

    # block 0 is the Kenya<->partner flow; shift its year columns by the
    # header-to-data offset observed on the anchor row
    off = data_blocks[0][0] - sub_blocks[0][0]
    block0 = [(i + off, sub_years[i]) for i in sub_blocks[0]]
    chosen = _select_year_cols(block0, years)

    krep, kpartner = _bilateral_parties(grid[hdr_idx - 1], path)
    rows = []
    for r in grid[hdr_idx + 1:]:
        if not any(_cell_text(c) for c in r):
            continue
        code = _clean_code(r[0])
        label = _cell_text(r[1])
        vals = [r[i] if i < len(r) else None for i, _ in chosen]
        rows.append(["", "", "000", "", code, label] + vals)

    labels = {"reporter": krep or "Kenya", "partner": kpartner}
    return rows, _norm_ycols(chosen), [y for _, y in chosen], labels


def _column_runs(cols):
    """Split a sorted column-index list into runs of consecutive indexes."""
    runs, cur = [], []
    for i in cols:
        if cur and i != cur[-1] + 1:
            runs.append(cur)
            cur = []
        cur.append(i)
    if cur:
        runs.append(cur)
    return runs


def _bilateral_parties(title_row, path):
    """(reporter, partner) from a bilateral workbook.

    Prefers the flow title cell ("Kenya's exports to Malaysia" /
    "Kenya's imports from Malaysia"); falls back to the filename
    ('..._between_Kenya_and_<Partner>').
    """
    for c in title_row:
        t = _cell_text(c)
        if not t or "world" in t.lower():
            continue
        m = re.match(r"^(.+?)['\u2019]s\s+exports?\s+to\s+(.+?)\s*$", t,
                     re.IGNORECASE) \
            or re.match(r"^(.+?)['\u2019]s\s+imports?\s+from\s+(.+?)\s*$", t,
                        re.IGNORECASE)
        if m:
            return display_name(m.group(1)), display_name(m.group(2))
    return "Kenya", _country_from_download_name(path) or "Partner"


def _pick_years(years):
    """Choose the report years: the longest run of 4..10 consecutive years
    (ties broken by the most recent run). Falls back to the last five years
    if no valid run exists.
    """
    if not years:
        return []
    runs = []
    cur = [years[0]]
    for y in years[1:]:
        if y == cur[-1] + 1:
            cur.append(y)
        else:
            runs.append(cur)
            cur = [y]
    runs.append(cur)
    valid = [r for r in runs if 4 <= len(r) <= 10]
    if not valid:
        return years[-min(5, len(years)):]
    return max(valid, key=lambda r: (len(r), r[-1]))


def _pick_year_cols(ycols):
    """Choose the year columns the report will cover (see _pick_years)."""
    selected = set(_pick_years([y for _, y in ycols]))
    return [(i, y) for i, y in ycols if y in selected]


def _balance_years(years, n=10):
    """Trailing run of consecutive years (at most `n`) ending at the latest.

    Used for the trade-balance series: unlike Tables 1-6 (last five years),
    the balance reaches back up to ten years whenever the Kenya export and
    import sources carry that history.
    """
    if not years:
        return []
    years = sorted(set(years))
    run = [years[-1]]
    for y in reversed(years[:-1]):
        if y == run[0] - 1:
            run.insert(0, y)
        else:
            break
    return run[-n:]


def _vals(row, ycols):
    return [to_float(row[i]) if i < len(row) else None for i, _ in ycols]


def extract_markets(rows, ycols):
    """Partner-country tables (Table 1 / Table 3): one row per partner."""
    total = None
    items = []
    for r in rows:
        if len(r) < 6:
            continue
        if str(r[4] or "").strip().upper() != "TOTAL":
            continue
        vals = _vals(r, ycols)
        if str(r[2] or "").strip() == "000":
            total = {"label": str(r[3]).strip(), "vals": vals}
        else:
            items.append({"label": str(r[3]).strip(), "vals": vals})
    return total, items


def extract_products(rows, ycols):
    """Product tables (Table 2 / Table 4): World rows only."""
    total = None
    items = []
    for r in rows:
        if str(r[2] or "").strip() != "000":
            continue
        code = str(r[4] or "").strip()
        vals = _vals(r, ycols)
        if code.upper() == "TOTAL":
            total = {"code": code, "label": str(r[5]).strip(), "vals": vals}
        else:
            items.append({"code": code, "label": str(r[5]).strip(), "vals": vals})
    return total, items


def extract_kenya(rows, ycols):
    """Bilateral product tables (Table 5 / Table 6): every row is Kenya's
    exports/imports with the partner."""
    total = None
    items = []
    for r in rows:
        code = str(r[4] or "").strip()
        vals = _vals(r, ycols)
        if code.upper() == "TOTAL":
            total = {"code": code, "label": str(r[5]).strip(), "vals": vals}
        else:
            items.append({"code": code, "label": str(r[5]).strip(), "vals": vals})
    return total, items


# ---------------------------------------------------------------------------
# Ranking / aggregation
# ---------------------------------------------------------------------------
def rank_and_top(items, top_n):
    """Sort by latest-year value (desc), assign ranks to every item and
    return the top `top_n` items."""
    items.sort(
        key=lambda it: (it["vals"][-1] if it["vals"][-1] is not None else float("-inf")),
        reverse=True,
    )
    for i, it in enumerate(items, 1):
        it["rank"] = i
    return items[:top_n]


def find_kenya(items):
    for it in items:
        if str(it.get("label", "")).strip().lower() == "kenya":
            return it
    return None


def sum_shown(shown, n_years):
    out = []
    for k in range(n_years):
        s = 0.0
        for it in shown:
            v = it["vals"][k]
            s += v if v is not None else 0.0
        out.append(s)
    return out


def all_other_vals(total, shown, n_years):
    """Total minus the sum of the displayed rows, per year."""
    if total is None:
        return [None] * n_years
    sums = sum_shown(shown, n_years)
    out = []
    for k in range(n_years):
        t = total["vals"][k]
        out.append(t - sums[k] if t is not None else None)
    return out


def share(v, total_v):
    if v is None or total_v is None or total_v == 0:
        return None
    return v / total_v


def prepare(extracted, years, top_n, divisor, is_markets):
    """Build the display structure for one table.

    extracted : dict with 'total' and 'items' (raw USD Thousand values)
    divisor   : 1e6 for billion tables, 1e3 for million tables
    """
    total = extracted["total"]
    items = extracted["items"]

    for it in items:
        it["vals"] = [v / divisor if v is not None else None for v in it["vals"]]
    if total:
        total["vals"] = [v / divisor if v is not None else None for v in total["vals"]]

    shown = list(rank_and_top(items, top_n))
    if is_markets:
        kenya = find_kenya(items)
        if kenya is not None and kenya not in shown:
            shown.append(kenya)

    n = len(years)
    ao_vals = all_other_vals(total, shown, n)
    t_latest = total["vals"][-1] if total else None
    for it in shown:
        it["share"] = share(it["vals"][-1], t_latest)
    total["share"] = share(t_latest, t_latest) if total else None
    all_other = {"vals": ao_vals,
                 "share": share(ao_vals[-1], t_latest) if total else None}

    return {"years": years, "total": total, "shown": shown,
            "all_other": all_other}


# ---------------------------------------------------------------------------
# Worksheet styling helpers
# ---------------------------------------------------------------------------
def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_cell(cell, bold=False, size=FONT_SIZE, fill=None, numfmt=None,
               wrap=False, align="center", border_style="data"):
    """Apply professional styling to a cell.
    
    border_style: 'header', 'data', 'total', or 'none'
    """
    cell.font = Font(name=FONT, size=size, bold=bold, color="000000" if not bold else "1F3864")
    
    # Apply appropriate border
    if border_style == "header":
        cell.border = HEADER_BORDER
    elif border_style == "total":
        cell.border = TOTAL_BORDER
    elif border_style == "none":
        cell.border = Border()
    else:
        cell.border = DATA_BORDER
    
    if fill is not None:
        cell.fill = fill
    if numfmt is not None:
        cell.number_format = numfmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap,
                               indent=1 if align == "left" else 0)


def style_header(cell, text, size=HEADER_FONT_SIZE, fill=HDR_FILL, 
                 font_color=HDR_FONT_COLOR, align="center", wrap=True):
    """Style a header cell with professional dark header."""
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=font_color)
    cell.fill = fill
    cell.border = HEADER_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def get_row_fill(row_index):
    """Get alternating row fill for data rows (0-indexed)."""
    return ROW_FILL_EVEN if row_index % 2 == 0 else ROW_FILL_ODD


def put(ws, row, col, value, **kw):
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, **kw)
    return cell


def put_text(ws, row, col, text, bold=False, size=FONT_SIZE, fill=None, wrap=False,
             align="left", use_row_fill=True):
    """Write text with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, text, bold=bold, size=size, fill=fill,
               wrap=wrap, align=align)


def put_val(ws, row, col, value, bold=False, fill=None, fmt=FMT_VALUE, use_row_fill=True):
    """Write numeric value with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, value, bold=bold, fill=fill, numfmt=fmt)


def put_share(ws, row, col, value, bold=False, fill=None, use_row_fill=True):
    """Write share percentage with optional alternating row fill."""
    if fill is None and use_row_fill:
        fill = get_row_fill(row)
    return put(ws, row, col, value, bold=bold, fill=fill, numfmt=FMT_SHARE)


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


def _put_formula(ws, cache, row, col, formula, cached, bold=False, fill=None,
                 numfmt=None):
    """Write an Excel formula and register the numeric result for caching."""
    put(ws, row, col, formula, bold=bold, fill=fill, numfmt=numfmt)
    cache.append((_cell_ref(row, col), cached))


def _sheet_order(path):
    """Sheet names in document order (openpyxl names the xml files
    sheet1.xml, sheet2.xml ... matching that order)."""
    with zipfile.ZipFile(path) as zin:
        xml = zin.read("xl/workbook.xml").decode("utf-8", "ignore")
    return re.findall(r'<sheet[^>]*name="([^"]+)"', xml)


def _inject_cached_values(path, cache):
    """Add the cached <v> results next to the <f> formulas in every sheet.

    ``cache`` is either a list of (cell_ref, value) pairs for a single-sheet
    workbook or a dict {sheet_name: {cell_ref: value}}. openpyxl writes a
    formula without a cached value, so a plain data_only=True read would
    return None. This patches the saved workbook (which is exactly how Excel
    itself stores a recalculated formula) so the report pipeline can read the
    numbers while Excel still shows live formulas that recalculate when a
    value is edited.
    """
    if isinstance(cache, dict):
        per_sheet = {k: {r: v for r, v in m.items() if v is not None}
                     for k, m in cache.items()}
    else:
        per_sheet = {"Sheet1": {r: v for r, v in cache if v is not None}}
    if not any(per_sheet.values()):
        return
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        blobs = {n: zin.read(n) for n in names}
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    order = _sheet_order(path)
    patched = 0
    for i, sheet in enumerate(order):
        refs = per_sheet.get(sheet)
        if not refs:
            continue
        fname = f"xl/worksheets/sheet{i + 1}.xml"
        if fname not in blobs:
            continue
        root = etree.fromstring(blobs[fname])
        for cell in root.iter("{%s}c" % ns):
            ref = cell.get("r")
            if ref not in refs:
                continue
            if cell.find("{%s}f" % ns) is None:
                continue
            v = cell.find("{%s}v" % ns)
            if v is None:
                v = etree.SubElement(cell, "{%s}v" % ns)
            v.text = repr(refs[ref])
        blobs[fname] = etree.tostring(root, xml_declaration=True,
                                      encoding="UTF-8")
        patched += 1
    if not patched:
        raise RuntimeError("no worksheet patched with cached values")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, blobs[n])


def _missing_cached_refs(path, refs):
    """Verify patched-in formula caches without paying a full workbook load.

    Streams the file in read-only mode and skips raw record sheets ('Data…')
    -- they never hold formulas but dominate the quarterly workbooks' size.
    Returns [(sheet, ref)] for every cached ref still reading as None.
    """
    wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        bad = []
        for sh, m in refs.items():
            if sh not in wb2.sheetnames:
                continue
            want = {coordinate_to_tuple(r): r for r, v in m.items()
                    if v is not None}
            if not want or sh.lower().startswith("data"):
                continue
            ws2 = wb2[sh]
            seen = set()
            for row in ws2.iter_rows():
                for c in row:
                    key = (c.row, c.column)
                    if key in want and c.value is not None:
                        seen.add(key)
                if len(seen) == len(want):
                    break
            bad.extend((sh, want[k]) for k in want.keys() - seen)
        return bad
    finally:
        wb2.close()


def _save_workbook(wb, path, cache):
    """Save the workbook, patch in cached formula values, and fall back to
    plain values if anything goes wrong so the pipeline never breaks."""
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(path)
    try:
        _inject_cached_values(path, cache)
        refs = cache if isinstance(cache, dict) else \
            {"Sheet1": {r: v for r, v in cache if v is not None}}
        bad = _missing_cached_refs(path, refs)
        if bad:
            raise RuntimeError(f"cached values missing for {bad}")
    except Exception:
        wb3 = openpyxl.load_workbook(path)
        refs = cache if isinstance(cache, dict) else \
            {"Sheet1": {r: v for r, v in cache if v is not None}}
        for sh, m in refs.items():
            if sh not in wb3.sheetnames:
                continue
            ws3 = wb3[sh]
            for ref, val in m.items():
                ws3[ref] = val
        wb3.save(path)


def _finalize(ws, path, cache):
    """Convenience wrapper for _save_workbook with a single worksheet."""
    _save_workbook(ws.parent, path, {ws.title: {r: v for r, v in cache
                                                if v is not None}})



# ---------------------------------------------------------------------------
# Table writers
# ---------------------------------------------------------------------------
def write_market_table(ws, data, rep, is_exports, unit_row, kenya_highlight,
                       row1_label=None, row1_title=None, cache=None):
    """Table 1 (import source markets) and Table 3 (export destinations).

    The share column, the "All other countries" row and the "World" row are
    written as live Excel formulas (with cached results), so the workbook
    recalculates when a value is edited.
    """
    cache = [] if cache is None else cache
    years = data["years"]
    n = len(years)
    latest = years[-1]
    total = data["total"]
    shown = data["shown"]

    label_col = "Importers" if is_exports else "Exporters"
    verb = ("Importing Countries from " if is_exports else "List of Exporters to ")
    verb += rep
    if not unit_row:
        verb += "\nValue in USD Billion"

    # optional title row (Table 1 only)
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 2 + n)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    share_col = 3 + n
    
    # Header row with dark navy fill
    put_text(ws, h, 1, f"Rank in {latest}", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 1).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 1, h + 1, 1)
    
    put_text(ws, h, 2, label_col, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 2).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 3, h, 2 + n)
    put_text(ws, h, 3, verb, bold=True, size=HEADER_FONT_SIZE if unit_row else HEADER_FONT_SIZE,
             fill=None if unit_row else HDR_FILL, wrap=True, align="center", use_row_fill=False)
    if not unit_row:
        ws.cell(h, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, share_col, h + 1, share_col)
    put_text(ws, h, share_col, f"Share in {latest} %", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, share_col).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    # years row
    for k, y in enumerate(years):
        put_text(ws, h + 1, 3 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 1, 3 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    if unit_row:
        merge(ws, h + 2, 3, h + 2, 2 + n)
        put_text(ws, h + 2, 3, "Value in USD Billion", bold=True, size=HEADER_FONT_SIZE,
                 fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 2, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
        put_text(ws, h + 2, 2, "", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        row0 = h + 3
    else:
        row0 = h + 2

    # data rows
    r_first = row0
    r_last = row0 + len(shown) - 1
    r_ao = row0 + len(shown)
    r_world = r_ao + 1
    for i, it in enumerate(shown):
        r = row0 + i
        is_kenya = kenya_highlight and str(it["label"]).strip().lower() == "kenya"
        fill = KENYA_FILL if is_kenya else get_row_fill(i)
        put_text(ws, r, 1, it["rank"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, display_name(it["label"]), size=FONT_SIZE, fill=fill, wrap=True, use_row_fill=False)
        for k in range(n):
            put_val(ws, r, 3 + k, it["vals"][k], fill=fill, use_row_fill=False)
        if it["share"] is not None and total and total["vals"][-1] is not None:
            _put_formula(ws, cache, r, share_col,
                         f"={_cell_ref(r, 3 + n - 1)}/{_cell_ref(r_world, 3 + n - 1)}",
                         it["share"], fill=fill, numfmt=FMT_SHARE)
        else:
            put_share(ws, r, share_col, it["share"], fill=fill, use_row_fill=False)

    # all other row
    r = r_ao
    ao_fill = BAND_FILL
    put_text(ws, r, 2, "All other countries", size=FONT_SIZE, fill=ao_fill, wrap=True, use_row_fill=False)
    for k in range(n):
        c = 3 + k
        val = data["all_other"]["vals"][k]
        if (len(shown) and val is not None
                and total and total["vals"][k] is not None):
            _put_formula(ws, cache, r, c,
                         f"={_cell_ref(r_world, c)}"
                         f"-SUM({_cell_ref(r_first, c)}:{_cell_ref(r_last, c)})",
                         val, fill=ao_fill, numfmt=FMT_VALUE)
        else:
            put_val(ws, r, c, val, fill=ao_fill, use_row_fill=False)
    ao_share = data["all_other"]["share"]
    if ao_share is not None and total and total["vals"][-1] is not None:
        _put_formula(ws, cache, r, share_col,
                     f"={_cell_ref(r_ao, 3 + n - 1)}/{_cell_ref(r_world, 3 + n - 1)}",
                     ao_share, fill=ao_fill, numfmt=FMT_SHARE)
    else:
        put_share(ws, r, share_col, ao_share, fill=ao_fill, use_row_fill=False)

    # world row
    r = r_world
    total_fill = BAND_FILL
    put_text(ws, r, 2, "World", bold=True, size=FONT_SIZE, fill=total_fill, use_row_fill=False)
    ws.cell(r, 2).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
    for k in range(n):
        c = 3 + k
        val = total["vals"][k] if total else None
        put_val(ws, r, c, val, bold=True, fill=total_fill, use_row_fill=False)
    put_share(ws, r, share_col, total["share"] if total else None, bold=True, fill=total_fill, use_row_fill=False)
    ws.freeze_panes = f"A{row0}"


def write_product_table(ws, data, hdr, unit_row,
                        row1_label=None, row1_title=None,
                        prefix_apostrophe=True, cache=None):
    """Tables 2, 4, 5, 6 (product tables).

    The share column, the "All other products" row and the "All products"
    total row are written as live Excel formulas (with cached results).
    """
    cache = [] if cache is None else cache
    years = data["years"]
    n = len(years)
    latest = years[-1]
    total = data["total"]
    shown = data["shown"]

    # optional title row (Tables 2, 4, 6)
    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 2 + n)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    # Header row with dark navy fill
    put_text(ws, h, 1, f"Rank in {latest}", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 1).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 2, h + 1, 2)
    put_text(ws, h, 2, "Code", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 2).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 3, h + 1, 3)
    put_text(ws, h, 3, "Product label", bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 3).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 4, h, 3 + n)
    put_text(ws, h, 4, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, wrap=True, align="center", use_row_fill=False)
    ws.cell(h, 4).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    merge(ws, h, 4 + n, h + 1, 4 + n)
    put_text(ws, h, 4 + n, f"Share in {latest} %", bold=True, size=HEADER_FONT_SIZE, 
             fill=HDR_FILL, align="center", use_row_fill=False)
    ws.cell(h, 4 + n).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    # years row
    for k, y in enumerate(years):
        put_text(ws, h + 1, 4 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 1, 4 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    if unit_row:
        merge(ws, h + 2, 4, h + 2, 3 + n)
        put_text(ws, h + 2, 4, "Value in USD Billion", bold=True, size=HEADER_FONT_SIZE,
                 fill=HDR_FILL, align="center", use_row_fill=False)
        ws.cell(h + 2, 4).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
        row0 = h + 3
    else:
        row0 = h + 2

    # data rows
    r_first = row0
    r_last = row0 + len(shown) - 1
    r_ao = row0 + len(shown)
    r_total = r_ao + 1
    for i, it in enumerate(shown):
        r = row0 + i
        fill = get_row_fill(i)
        put_text(ws, r, 1, it["rank"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        code = it["code"]
        if prefix_apostrophe and code:
            code = "'" + code
        put_text(ws, r, 2, code, size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 3, short_product_name(it["label"], it.get("code"), maxlen=48),
                 size=FONT_SIZE, fill=fill, wrap=True, use_row_fill=False)
        for k in range(n):
            put_val(ws, r, 4 + k, it["vals"][k], fill=fill, use_row_fill=False)
        if it["share"] is not None and total and total["vals"][-1] is not None:
            _put_formula(ws, cache, r, 4 + n,
                         f"={_cell_ref(r, 4 + n - 1)}/{_cell_ref(r_total, 4 + n - 1)}",
                         it["share"], fill=fill, numfmt=FMT_SHARE)
        else:
            put_share(ws, r, 4 + n, it["share"], fill=fill, use_row_fill=False)

    # all other row
    r = r_ao
    ao_fill = BAND_FILL
    put_text(ws, r, 3, "All other products", size=FONT_SIZE, fill=ao_fill, wrap=True, use_row_fill=False)
    for k in range(n):
        c = 4 + k
        val = data["all_other"]["vals"][k]
        if (len(shown) and val is not None
                and total and total["vals"][k] is not None):
            _put_formula(ws, cache, r, c,
                         f"={_cell_ref(r_total, c)}"
                         f"-SUM({_cell_ref(r_first, c)}:{_cell_ref(r_last, c)})",
                         val, fill=ao_fill, numfmt=FMT_VALUE)
        else:
            put_val(ws, r, c, val, fill=ao_fill, use_row_fill=False)
    ao_share = data["all_other"]["share"]
    if ao_share is not None and total and total["vals"][-1] is not None:
        _put_formula(ws, cache, r, 4 + n,
                     f"={_cell_ref(r_ao, 4 + n - 1)}/{_cell_ref(r_total, 4 + n - 1)}",
                     ao_share, fill=ao_fill, numfmt=FMT_SHARE)
    else:
        put_share(ws, r, 4 + n, ao_share, fill=ao_fill, use_row_fill=False)

    # total row
    r = r_total
    total_fill = BAND_FILL
    if total and total.get("code"):
        put_text(ws, r, 2, ("'" if prefix_apostrophe else "") + str(total["code"]),
                 bold=True, size=FONT_SIZE, fill=total_fill, wrap=True, use_row_fill=False)
    put_text(ws, r, 3, "All products", bold=True, size=FONT_SIZE, fill=total_fill, wrap=True, use_row_fill=False)
    ws.cell(r, 3).font = Font(name=FONT, size=FONT_SIZE, bold=True, color="1F3864")
    for k in range(n):
        c = 4 + k
        val = total["vals"][k] if total else None
        put_val(ws, r, c, val, bold=True, fill=total_fill, use_row_fill=False)
    put_share(ws, r, 4 + n, total["share"] if total else None, bold=True, fill=total_fill, use_row_fill=False)
    ws.freeze_panes = f"A{row0}"


def write_balance(ws, krep, kpartner, years, exports, imports, cache=None, skip_chart=False):
    """Figure 1: bilateral trade balance derived from the Table 5/6 totals.

    exports = Kenya's exports to the partner (Table 5 total, USD Million)
    imports = Kenya's imports from the partner (Table 6 total, USD Million)
    Balance of Trade = exports - imports (a live Excel formula).
    """
    cache = [] if cache is None else cache
    hdr = HDR_FILL
    put_text(ws, 1, 1, "Figure 1", bold=True, size=12, align="center", use_row_fill=False)
    put_text(ws, 1, 2, f"BOT Kenya- {kpartner}", bold=True, size=12, align="center", use_row_fill=False)
    put_text(ws, 2, 2, f"Kenya's exports to {kpartner}", size=FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    ws.cell(2, 2).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    put_text(ws, 3, 2, "", bold=True, size=HEADER_FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    for k, y in enumerate(years):
        put_text(ws, 3, 2 + k, y, bold=True, size=HEADER_FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
        ws.cell(3, 2 + k).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    def row(label, values, fill_color=None):
        fill = fill_color if fill_color else hdr
        put_text(ws, r, 1, label, bold=True, size=FONT_SIZE, fill=fill, align="center", use_row_fill=False)
        ws.cell(r, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR if fill == hdr else "1F3864")
        for k, v in enumerate(values):
            data_fill = ROW_FILL_EVEN if k % 2 == 0 else ROW_FILL_ODD
            put_val(ws, r, 2 + k, v, fill=data_fill, use_row_fill=False)

    r = 4
    row("Exports", exports, hdr)
    r += 1
    row("Imports", imports, hdr)
    r += 1
    put_text(ws, r, 1, "Balance of Trade", bold=True, size=FONT_SIZE, fill=hdr, align="center", use_row_fill=False)
    ws.cell(r, 1).font = Font(name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    for k, (e, i) in enumerate(zip(exports, imports)):
        c = 2 + k
        if e is not None and i is not None:
            _put_formula(ws, cache, r, c, f"={_cell_ref(4, c)}-{_cell_ref(5, c)}",
                         e - i, fill=BAND_FILL)
        else:
            put_val(ws, r, c, (e - i) if (e is not None and i is not None) else None, fill=BAND_FILL, use_row_fill=False)
    if not skip_chart:
        _add_balance_chart(ws, krep, kpartner, years)


def _add_balance_chart(ws, krep, kpartner, years):
    """Embed a clustered bar chart of exports/imports/balance below the table."""
    try:
        from openpyxl.chart import BarChart, Reference
        n = len(years)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"{krep}-{kpartner} Balance of Trade"
        chart.y_axis.title = "USD Million"
        chart.x_axis.title = "Year"
        chart.y_axis.crosses = "autoZero"
        chart.height = 9
        chart.width = 20
        cats = Reference(ws, min_col=2, max_col=1 + n, min_row=3)
        chart.add_data(Reference(ws, min_col=1, max_col=1 + n,
                                 min_row=4, max_row=6),
                       titles_from_data=True, from_rows=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{8 if len(years) <= 5 else 9}")
    except Exception:
        pass


def read_existing_balance(path):
    """Read (years, exports, imports) from an existing balance workbook.

    Handles both the layout written by write_balance and the reference
    "Figure 1 Trade Balance.xlsx" (which may carry two series blocks and extra
    columns). Returns None if the file is unreadable or has no valid series.
    """
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        grid = [[c.value for c in row] for row in ws.iter_rows()]
    except Exception:
        return None

    for ri, row in enumerate(grid):
        if not row or not isinstance(row[0], str):
            continue
        lbl = row[0].strip()
        if lbl == "Exports" and ri + 2 < len(grid):
            l1 = grid[ri + 1][0].strip() if isinstance(grid[ri + 1][0], str) else ""
            l2 = grid[ri + 2][0].strip() if isinstance(grid[ri + 2][0], str) else ""
            if l1 == "Imports" and l2 == "Balance of Trade":
                years = [int(v) for v in grid[ri - 1][1:]
                         if isinstance(v, (int, float)) and not isinstance(v, bool)
                         and 1990 <= v <= 2100]
                if not years:
                    continue
                exports = [grid[ri][1 + i] if isinstance(grid[ri][1 + i], (int, float)) else None
                           for i in range(len(years))]
                imports = [grid[ri + 1][1 + i] if isinstance(grid[ri + 1][1 + i], (int, float)) else None
                           for i in range(len(years))]
                return {"years": years, "exports": exports, "imports": imports}
    return None


def merge_balance_series(existing, years, exports, imports):
    """Merge new data over an existing series, keeping older history."""
    if not existing:
        return years, exports, imports
    ymap = dict(zip(existing["years"], existing["exports"]))
    imap = dict(zip(existing["years"], existing["imports"]))
    for y, e, i in zip(years, exports, imports):
        ymap[y] = e
        imap[y] = i
    ys = sorted(ymap)
    return ys, [ymap[y] for y in ys], [imap[y] for y in ys]


# ---------------------------------------------------------------------------
# Bilateral export composition analysis
# ---------------------------------------------------------------------------
def compute_bilateral_alignment(kenya_items, kenya_total, partner_items,
                                partner_total, years):
    """Compare Kenya's export products to the partner vs partner's import demand.

    Matches by HS code. For each Kenya product, computes:
    - The product's share within Kenya's exports to the partner
    - Partner's import share for the same product
    - Alignment score: how well Kenya's exports match partner's import demand

    Inputs must already be scaled to USD Million (prepare() rescales the
    extracted structures in place, so the "raw" references below are in
    millions by the time this is called).
    """
    partner_by_code = {}
    for it in partner_items:
        partner_by_code[it["code"]] = it

    latest = -1
    results = []
    for kit in kenya_items:
        code = kit["code"]
        pit = partner_by_code.get(code)
        if pit is None:
            continue

        kv = kit["vals"][latest] if kit["vals"] and len(kit["vals"]) > abs(latest) else None
        kt = kenya_total["vals"][latest] if kenya_total and kenya_total["vals"] else None
        pv = pit["vals"][latest] if pit["vals"] and len(pit["vals"]) > abs(latest) else None
        pt = partner_total["vals"][latest] if partner_total and partner_total["vals"] else None

        if kv is None or kt is None or kt == 0:
            continue
        k_share = kv / kt

        p_share = None
        if pv is not None and pt is not None and pt != 0:
            p_share = pv / pt

        # Alignment: how relevant is this product to the partner's imports
        alignment = p_share if p_share is not None else 0

        # Growth
        growth = None
        if (len(kit["vals"]) >= 2 and kit["vals"][-1] is not None
                and kit["vals"][-2] is not None and kit["vals"][-2] != 0):
            growth = (kit["vals"][-1] - kit["vals"][-2]) / abs(kit["vals"][-2])

        # Values arrive in USD Million (prepare() rescales the extracted
        # items in place before this runs), matching Tables 5 and 6.
        results.append({
            "code": code,
            "label": kit["label"],
            "kenya_val": kv,
            "kenya_share": k_share,
            "partner_val": pv,
            "partner_share": p_share,
            "alignment": alignment,
            "growth": growth,
        })

    results.sort(key=lambda x: x["kenya_val"] if x["kenya_val"] is not None else -1,
                 reverse=True)
    results = results[:30]
    for i, it in enumerate(results, 1):
        it["rank"] = i
    return results


def write_bilateral_table(ws, items, years, krep, kpartner,
                          row1_label=None, row1_title=None):
    """Write Kenya's bilateral export composition + market alignment table.

    Columns: Rank | Code | Product | Kenya Value | Kenya Share | Partner Import Share | Growth %
    """
    latest = years[-1] if years else ""

    if row1_label or row1_title:
        if row1_label:
            put_text(ws, 1, 1, row1_label, bold=True, size=12, align="center", use_row_fill=False)
        merge(ws, 1, 2, 1, 7)
        put_text(ws, 1, 2, row1_title, bold=True, size=12, wrap=True, align="center", use_row_fill=False)
        title_rows = 1
    else:
        title_rows = 0

    h = 1 + title_rows
    cols = ["Rank", "HS Code", "Product",
            f"Kenya Value USD M ({latest})",
            f"Share in Kenya's Exports to {kpartner} (%)",
            f"{kpartner} Import Share (%)",
            "Annual Growth %"]
    for c, hdr in enumerate(cols, 1):
        put_text(ws, h, c, hdr, bold=True, size=HEADER_FONT_SIZE, fill=HDR_FILL, align="center",
                 wrap=True, use_row_fill=False)
        ws.cell(h, c).font = Font(name=FONT, size=HEADER_FONT_SIZE, bold=True, color=HDR_FONT_COLOR)

    row0 = h + 1
    for i, it in enumerate(items):
        r = row0 + it["rank"] - 1
        fill = get_row_fill(i)
        put_text(ws, r, 1, it["rank"], size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 2, str(it["code"]), size=FONT_SIZE, fill=fill, use_row_fill=False)
        put_text(ws, r, 3, short_product_name(it["label"], it.get("code"), maxlen=55),
                 size=FONT_SIZE, fill=fill, wrap=True, align="left", use_row_fill=False)
        put_val(ws, r, 4, it["kenya_val"], fill=fill, use_row_fill=False)
        put_share(ws, r, 5, it["kenya_share"], fill=fill, use_row_fill=False)
        put_share(ws, r, 6, it["partner_share"], fill=fill, use_row_fill=False)
        if it["growth"] is not None:
            put_share(ws, r, 7, it["growth"], fill=fill, use_row_fill=False)
        else:
            put_share(ws, r, 7, None, fill=fill, use_row_fill=False)

    set_widths(ws, {"A": 7, "B": 10, "C": 48, "D": 18, "E": 14, "F": 20, "G": 14})
    ws.freeze_panes = f"A{row0}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate_tables(excel_dir, out_dir, top_n):
    """Build the Table 1-6 workbooks.

    top_n : number of top product rows kept for Tables 2, 4, 5 and 6.
            Market tables (Tables 1 and 3) always keep TOP_MARKETS rows.
    """
    os.makedirs(out_dir, exist_ok=True)
    files = find_source_files(excel_dir)

    # ---- common year set: all six tables must share the same years ---------
    keys = ("table1", "table2", "table3", "table4", "table5", "table6")
    years_sets = []
    for k in keys:
        _, _, ys, _ = parse_source(files[k])
        years_sets.append(set(ys))
    full_years = _pick_years(sorted(set.intersection(*years_sets)))
    # Tables 1-6 show only the last five years; Figure 1 (the trade balance)
    # uses the last up-to-TEN consecutive years so the analysis covers e.g.
    # 2016-2025 whenever the source data carries that history.
    table_years = full_years[-5:] if len(full_years) > 5 else full_years
    # The balance window is derived from the Kenya exports (Table 5) and
    # imports (Table 6) sources alone -- it must not shrink just because the
    # narrower Tables 1-4 cover fewer years. Older history already captured
    # in an existing balance workbook is merged back in further below.
    _, _, ys5_all, _ = parse_source(files["table5"])
    _, _, ys6_all, _ = parse_source(files["table6"])
    balance_window = _balance_years(set(ys5_all) & set(ys6_all), n=10) or full_years

    # ---- Table 1: import source markets --------------------------------
    rows, ycols, years, labels = parse_source(files["table1"], years=table_years)
    total, items = extract_markets(rows, ycols)
    rep = labels["reporter"]
    d1 = prepare({"total": total, "items": items}, years, TOP_MARKETS, 1e6, is_markets=True)

    # ---- Table 3: export destinations -----------------------------------
    rows, ycols, years, labels = parse_source(files["table3"], years=table_years)
    total, items = extract_markets(rows, ycols)
    d3 = prepare({"total": total, "items": items}, years, TOP_MARKETS, 1e6, is_markets=True)

    # ---- Table 2: import products ---------------------------------------
    rows, ycols, years, labels = parse_source(files["table2"], years=table_years)
    total, items = extract_products(rows, ycols)
    d2 = prepare({"total": total, "items": items}, years, top_n, 1e6, is_markets=False)
    p_total_raw = total   # NOTE: USD Million -- prepare() rescaled in place
    p_items_raw = items   # (same object), not raw USD Thousand

    # ---- Table 4: export products ---------------------------------------
    rows, ycols, years, labels = parse_source(files["table4"], years=table_years)
    total, items = extract_products(rows, ycols)
    d4 = prepare({"total": total, "items": items}, years, top_n, 1e6, is_markets=False)

    # ---- Table 5: Kenya exports to partner ------------------------------
    rows, ycols, years, labels = parse_source(files["table5"], years=table_years)
    total, items = extract_kenya(rows, ycols)
    krep = labels["reporter"] or "Kenya"
    kpartner = labels["partner"]
    d5 = prepare({"total": total, "items": items}, years, top_n, 1e3, is_markets=False)
    k_total_raw = total   # NOTE: USD Million -- prepare() rescaled in place
    k_items_raw = items   # (same object), not raw USD Thousand

    # ---- Table 6: Kenya imports from partner ----------------------------
    rows, ycols, years, labels = parse_source(files["table6"], years=table_years)
    total, items = extract_kenya(rows, ycols)
    d6 = prepare({"total": total, "items": items}, years, top_n, 1e3, is_markets=False)

    # ---- Table 7: Kenya's export composition + market alignment -----------
    bilateral_items = compute_bilateral_alignment(
        k_items_raw, k_total_raw, p_items_raw, p_total_raw, years)

    rep_possessive = rep if rep.endswith("s") else rep + "'s"

    out = {}

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 1"
    cache = []
    write_market_table(ws, d1, rep, is_exports=False, unit_row=False,
                       kenya_highlight=True, cache=cache,
                       row1_label="Table 1:",
                       row1_title=f"{rep} Import Source Markets")
    set_widths(ws, {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12})
    out["t1"] = os.path.join(out_dir, f"Table 1 {rep} Import Source Markets.xlsx")
    _finalize(ws, out["t1"], cache)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 2"
    cache = []
    write_product_table(ws, d2, "Imported Products from the World",
                        unit_row=True, cache=cache,
                        row1_title=f"List of products imported by {rep}")
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12})
    out["t2"] = os.path.join(out_dir, f"Table 2 {rep} Lead Imports.xlsx")
    _finalize(ws, out["t2"], cache)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 3"
    cache = []
    write_market_table(ws, d3, rep, is_exports=True, unit_row=True,
                       kenya_highlight=True, cache=cache)
    set_widths(ws, {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12})
    out["t3"] = os.path.join(out_dir, f"Table 3 {rep} Lead Export Destinations.xlsx")
    _finalize(ws, out["t3"], cache)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 4"
    cache = []
    write_product_table(
        ws, d4, "Exported Products to the World\nValue in USD Billions",
        unit_row=False, cache=cache,
        row1_title=f"{rep_possessive} Lead Export Products to the World")
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12})
    out["t4"] = os.path.join(out_dir, f"Table 4 {rep} Lead Export Products.xlsx")
    _finalize(ws, out["t4"], cache)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 5"
    cache = []
    write_product_table(
        ws, d5, f"{krep}'s exports to {kpartner}\nValue in USD Million",
        unit_row=False, cache=cache)
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12})
    out["t5"] = os.path.join(out_dir, f"Table 5 {krep} Exports to {kpartner}.xlsx")
    _finalize(ws, out["t5"], cache)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 6"
    cache = []
    write_product_table(
        ws, d6, f"{krep}'s imports from {kpartner}\nValue in USD Million",
        unit_row=False, cache=cache,
        row1_label="Table 6",
        row1_title=f"{krep}'s Top imports from {kpartner}")
    set_widths(ws, {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 12})
    out["t6"] = os.path.join(out_dir, f"Table 6 {krep} Top Imports from {kpartner}.xlsx")
    _finalize(ws, out["t6"], cache)

    # Table 7: Kenya's export composition + market alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table 7"
    cache = []
    write_bilateral_table(ws, bilateral_items, years, krep, kpartner,
                          row1_label="Table 7:",
                          row1_title=f"{krep}'s Export Composition and Market Alignment with {kpartner}")
    set_widths(ws, {"A": 7, "B": 10, "C": 48, "D": 18, "E": 14, "F": 20, "G": 14})
    out["t7"] = os.path.join(out_dir, f"Table 7 {krep} Export Composition and Market Alignment.xlsx")
    _finalize(ws, out["t7"], cache)

    # ---- Figure 1: trade balance derived from Table 5 / Table 6 ----------
    # The balance spans the last up-to-10 consecutive years even though
    # Tables 1-6 only show the last five.
    bal_path = os.path.join(out_dir, "Figure 1 Trade Balance.xlsx")
    rows5f, ycols5f, years5f, _ = parse_source(files["table5"], years=balance_window)
    rows6f, ycols6f, years6f, _ = parse_source(files["table6"], years=balance_window)
    tot5f, _ = extract_kenya(rows5f, ycols5f)
    tot6f, _ = extract_kenya(rows6f, ycols6f)
    exports_full = [v / 1e3 if v is not None else None for v in tot5f["vals"]]
    imports_full = [v / 1e3 if v is not None else None for v in tot6f["vals"]]
    years_m, exports_m, imports_m = merge_balance_series(
        read_existing_balance(bal_path),
        years5f, exports_full, imports_full)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Figure 1"
    cache = []
    write_balance(ws, krep, kpartner, years_m, exports_m, imports_m, cache=cache)
    set_widths(ws, BALANCE_WIDTHS)
    out["balance"] = bal_path
    _finalize(ws, out["balance"], cache)

    # ---- Combined workbook: all seven sheets in one file -------------------
    wb_all = openpyxl.Workbook()
    wb_all.remove(wb_all.active)
    for s in ("Table 1", "Table 2", "Table 3", "Table 4", "Table 5",
              "Table 6", "Table 7", "Figure 1"):
        wb_all.create_sheet(s)
    cache_all = {}
    widths_mkt = {"A": 7, "B": 32, "C": 12, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12}
    widths_prd = {"A": 7, "B": 9, "C": 45, "D": 12, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12}
    ws = wb_all["Table 1"]; c = []
    write_market_table(ws, d1, rep, is_exports=False, unit_row=False,
                       kenya_highlight=True, cache=c,
                       row1_label="Table 1:",
                       row1_title=f"{rep} Import Source Markets")
    set_widths(ws, widths_mkt); cache_all["Table 1"] = dict(c)
    ws = wb_all["Table 2"]; c = []
    write_product_table(ws, d2, "Imported Products from the World",
                        unit_row=True, cache=c,
                        row1_title=f"List of products imported by {rep}")
    set_widths(ws, widths_prd); cache_all["Table 2"] = dict(c)
    ws = wb_all["Table 3"]; c = []
    write_market_table(ws, d3, rep, is_exports=True, unit_row=True,
                       kenya_highlight=True, cache=c)
    set_widths(ws, widths_mkt); cache_all["Table 3"] = dict(c)
    ws = wb_all["Table 4"]; c = []
    write_product_table(
        ws, d4, "Exported Products to the World\nValue in USD Billions",
        unit_row=False, cache=c,
        row1_title=f"{rep_possessive} Lead Export Products to the World")
    set_widths(ws, widths_prd); cache_all["Table 4"] = dict(c)
    ws = wb_all["Table 5"]; c = []
    write_product_table(
        ws, d5, f"{krep}'s exports to {kpartner}\nValue in USD Million",
        unit_row=False, cache=c)
    set_widths(ws, widths_prd); cache_all["Table 5"] = dict(c)
    ws = wb_all["Table 6"]; c = []
    write_product_table(
        ws, d6, f"{krep}'s imports from {kpartner}\nValue in USD Million",
        unit_row=False, cache=c,
        row1_label="Table 6",
        row1_title=f"{krep}'s Top imports from {kpartner}")
    set_widths(ws, widths_prd); cache_all["Table 6"] = dict(c)
    ws = wb_all["Table 7"]; c = []
    write_bilateral_table(ws, bilateral_items, years, krep, kpartner,
                          row1_label="Table 7:",
                          row1_title=f"{krep}'s Export Composition and Market Alignment with {kpartner}")
    set_widths(ws, {"A": 7, "B": 10, "C": 48, "D": 18, "E": 14, "F": 20, "G": 14})
    cache_all["Table 7"] = dict(c)
    ws = wb_all["Figure 1"]; c = []
    write_balance(ws, krep, kpartner, years_m, exports_m, imports_m, cache=c, skip_chart=True)
    set_widths(ws, BALANCE_WIDTHS)
    cache_all["Figure 1"] = dict(c)
    out["all"] = os.path.join(out_dir, "All Tables.xlsx")
    _save_workbook(wb_all, out["all"], cache_all)

    return out, (rep, kpartner)


def main():
    ap = argparse.ArgumentParser(
        description="Build Table 1-6 workbooks from raw ITC source files.")
    ap.add_argument("--excel-dir", default="sourcefiles",
                    help="Folder with the six raw ITC source files (default: sourcefiles)")
    ap.add_argument("--out-dir", default=os.path.join("output", "tables"),
                    help="Where to write the generated Table 1-6 files "
                         "(default: output/tables)")
    ap.add_argument("--top", type=int, default=20,
                    help="Number of top product rows to keep (default: 20). "
                         "Tables 1 and 3 always keep the top 25 exporters")
    args = ap.parse_args()

    if args.top < 1:
        sys.exit("[ERROR] --top must be >= 1")

    try:
        out, (rep, partner) = generate_tables(args.excel_dir, args.out_dir, args.top)
    except SystemExit:
        raise
    except Exception as e:
        sys.exit(f"[ERROR] Failed to generate tables: {e}")

    print("[1/2] Read raw source files from :", os.path.abspath(args.excel_dir))
    print("[2/2] Wrote tables for           :", rep, "<-> Kenya")
    for key in ("t1", "t2", "t3", "t4", "t5", "t6", "t7", "balance", "all"):
        print(f"      - {os.path.basename(out[key])}")
    print()
    print("The Figure 1 file is derived from the Table 5 and Table 6 totals")
    print("(Kenya's exports to / imports from the partner country). The")
    print("'All Tables.xlsx' workbook bundles every sheet into one file.")
    print()
    print("Next step: run")
    print(f'  python generate_report.py --excel-dir "{os.path.abspath(args.out_dir)}"')


if __name__ == "__main__":
    main()
