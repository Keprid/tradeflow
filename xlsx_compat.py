"""
xlsx_compat.py
==============

Universal spreadsheet reader / converter.

ITC Trade Map downloads may arrive in any of these formats:
  .xlsx  / .xlsm  – modern Open XML (openpyxl native)
  .xlsb           – binary Open XML (pyxlsb)
  .xls            – legacy binary (xlrd) **or** HTML-table-in-disguise
  .csv / .tsv     – plain text

All public helpers convert an arbitrary file into an openpyxl ``Workbook``
that the caller can read with ``data_only=True``.  The conversion is
in-place-safe: the original file is kept, and a sibling ``*.xlsx`` is
written to a temp-like path only if needed.
"""

import csv
import html.parser
import io
import os
import re
import tempfile
from pathlib import Path

import openpyxl

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name, fallback):
    """A valid openpyxl sheet title (non-empty, <= 31 chars, no bad chars)."""
    n = _INVALID_SHEET_CHARS.sub("_", (name or "").strip())
    return n[:31] or fallback


def _read_text_any(path):
    """Read a text file, trying common encodings in turn."""
    p = Path(path)
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return p.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return p.read_text(encoding="latin-1", errors="replace")


def _sniff_spreadsheet_kind(path):
    """Classify a non-.xlsx file: 'html' | 'xls' | 'csv'.

    ITC / Trade Map downloads are often HTML tables (or comma/tab-delimited
    text) saved with a ``.xls`` extension.  ``xlrd`` only understands the
    real binary format, so sniff the first bytes to pick the right parser.
    """
    with open(path, "rb") as f:
        head = f.read(8192)
    low = head.lower()
    if b"<html" in low or b"<table" in low or b"<!doctype" in low \
            or b"<head" in low or b"<meta" in low:
        return "html"
    if b"\x00" in head:
        return "xls"
    try:
        text = head.decode("utf-8", "ignore")
    except Exception:
        return "xls"
    if any(ch in text for ch in ("\t", ",", ";", "|")):
        return "csv"
    return "xls"


# ---------------------------------------------------------------------------
# HTML table parser (for ITC .xls files that are really HTML)
# ---------------------------------------------------------------------------
class HTMLTableParser(html.parser.HTMLParser):
    """Extract the largest <table> in an HTML file as a list of row lists."""

    def __init__(self):
        super().__init__()
        self.tables = []
        self._cur = None
        self._row = None
        self._cell = None
        self._in_cell = False
        self._data = []

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._cur = []
        elif tag == "tr" and self._cur is not None:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._in_cell = True
            self._data = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._in_cell:
            self._row.append("".join(self._data).strip())
            self._in_cell = False
            self._data = []
        elif tag == "tr" and self._row is not None:
            if self._cur is not None:
                self._cur.append(self._row)
            self._row = None
        elif tag == "table" and self._cur is not None:
            self.tables.append(self._cur)
            self._cur = None

    def handle_data(self, data):
        if self._in_cell:
            self._data.append(data)


def _write_html_as_xlsx(src, out):
    """Parse an HTML file (often a .xls download) into the first .xlsx sheet."""
    text = _read_text_any(src)
    parser = HTMLTableParser()
    parser.feed(text)
    if not parser.tables:
        raise ValueError(f"No HTML table found in {src}")
    table = max(parser.tables, key=len)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, row in enumerate(table, start=1):
        for c, v in enumerate(row, start=1):
            if v != "":
                ws.cell(row=r, column=c, value=v)
    wb.save(str(out))


def _write_xls_as_xlsx(src, out):
    """Convert a real binary .xls (xlrd) into .xlsx."""
    import xlrd
    book = xlrd.open_workbook(str(src), formatting_info=False)
    wb = openpyxl.Workbook()
    for idx in range(book.nsheets):
        sheet = book.sheet_by_index(idx)
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = _safe_sheet_name(sheet.name, f"Sheet{idx + 1}")
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        v = xlrd.xldate_as_datetime(v, book.datemode)
                    except (ValueError, OverflowError):
                        pass
                elif cell.ctype == xlrd.XL_CELL_NUMBER \
                        and isinstance(v, float) and v.is_integer():
                    v = int(v)
                if v is not None:
                    ws.cell(row=r + 1, column=c + 1, value=v)
    wb.save(str(out))


def _write_xlsb_as_xlsx(src, out):
    """Convert a .xlsb file into .xlsx."""
    from pyxlsb import open_workbook
    wb = openpyxl.Workbook()
    first = True
    with open_workbook(str(src)) as book:
        for name in book.sheets:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = _safe_sheet_name(name, f"Sheet{len(wb.sheetnames) + 1}")
            with book.get_sheet(name) as sheet:
                r = 1
                for row in sheet.rows():
                    for c, cell in enumerate(row):
                        v = cell.v
                        if isinstance(v, float) and v.is_integer():
                            v = int(v)
                        if v is not None:
                            ws.cell(row=r, column=c + 1, value=v)
                    r += 1
    wb.save(str(out))


def _write_csv_as_xlsx(src, out):
    """Convert a CSV/TSV file into .xlsx."""
    text = _read_text_any(src)
    lines = text.splitlines()
    first = lines[0] if lines else ""
    counts = {d: first.count(d) for d in ("\t", ";", ",", "|")}
    delim = max(counts, key=counts.get) if any(counts.values()) else ","
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, row in enumerate(csv.reader(io.StringIO(text), delimiter=delim), start=1):
        for c, v in enumerate(row, start=1):
            if v != "":
                ws.cell(row=r, column=c, value=v)
    wb.save(str(out))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".xlsb", ".csv", ".tsv")


def is_spreadsheet(filename):
    """Return True if *filename* has a recognised spreadsheet extension."""
    return os.path.splitext(filename)[1].lower() in SUPPORTED_EXTENSIONS


def convert_to_xlsx(src, out_dir=None):
    """Convert any supported spreadsheet into .xlsx.

    Returns ``(openpyxl.Workbook, path_used)`` where *path_used* is the
    path that was actually opened (either the original or a converted
    copy).  Caller is responsible for closing the workbook.

    Files already in .xlsx/.xlsm are opened directly.  Everything else
    is converted to a temp ``.xlsx`` in *out_dir* (defaults to the
    source file's parent directory).  The temp file is NOT deleted —
    the caller should clean up if needed.
    """
    src = Path(src)
    ext = src.suffix.lower()

    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(str(src), data_only=True)
        return wb, str(src)

    if out_dir is None:
        out_dir = src.parent
    out_dir = Path(out_dir)

    out = out_dir / (src.stem + ".xlsx")
    i = 1
    while out.exists():
        out = out_dir / f"{src.stem}__{i}.xlsx"
        i += 1

    if ext == ".xls":
        kind = _sniff_spreadsheet_kind(src)
        if kind == "html":
            _write_html_as_xlsx(src, out)
        elif kind == "csv":
            _write_csv_as_xlsx(src, out)
        else:
            _write_xls_as_xlsx(src, out)
    elif ext == ".xlsb":
        _write_xlsb_as_xlsx(src, out)
    elif ext in (".csv", ".tsv"):
        _write_csv_as_xlsx(src, out)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {ext}")

    wb = openpyxl.load_workbook(str(out), data_only=True)
    return wb, str(out)
