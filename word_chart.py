"""Native, editable Word charts (Figures 1/2/3) for the report .docx.

Word's "editable chart" (the kind you click and get the Chart Design /
Chart Tools ribbon, plus "Edit Data" opening Excel) is a rich object made of
several OPC parts that python-docx does not expose:

* ``word/charts/chartN.xml``   -- the ``c:chartSpace`` XML describing the chart.
* ``word/charts/_rels/chartN.xml.rels``
* ``word/embeddings/...xlsx``  -- the embedded Excel workbook the chart is bound
  to.  Word opens exactly this file on "Edit Data", so it is what makes the
  chart editable.
* A DrawingML block inside the document body pointing at the chart part through
  ``c:chart/@r:id``.

Instead of hand-writing the fragile ``chartSpace`` XML we reuse openpyxl, which
produces the *same* schema-valid ``chartSpace`` it writes into an xlsx, then we
install the resulting parts into the package via python-docx's low-level
``Part`` / ``relate_to`` hooks.  The embedded workbook shown in Excel's
"Edit Data" is built with openpyxl too, so series/categories are live and
editable.
"""

import io

from lxml import etree

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from docx.opc.packuri import PackURI
from docx.opc.part import Part

# Namespaces
_NS_C = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_WP = ("http://schemas.openxmlformats.org/drawingml/2006/"
          "wordprocessingDrawing")
_NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

_C = "{%s}" % _NS_C
_A = "{%s}" % _NS_A
_R = "{%s}" % _NS_R
_WP = "{%s}" % _NS_WP
_W = "{%s}" % _NS_W

_REL_CHART = ("http://schemas.openxmlformats.org/officeDocument/2006/"
              "relationships/chart")
_REL_PKG = ("http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/package")

_CT_CHART = ("application/vnd.openxmlformats-officedocument.drawingml."
             "chart+xml")
_CT_XLSX = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

_CHART_URI = "http://schemas.openxmlformats.org/drawingml/2006/chart"

THEME = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
         "70AD47", "FF0000", "7030A0", "00B0F0", "F7A7A7"]


def _col_letter(idx):
    """1-based column index -> Excel column letter."""
    s = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


# ---------------------------------------------------------------------------
# openpyxl chart + workbook construction
# ---------------------------------------------------------------------------

def _write_chart_data(wb, kind, categories, series, colors):
    """Populate the first sheet (the embedded 'Edit Data' workbook)."""
    ws = wb.active
    ws.title = "Sheet1"

    if kind == "pie":
        ws.cell(1, 1, "Category")
        ws.cell(1, 2, "Share")
        for i, (lab, sh) in enumerate(zip(categories, series), start=2):
            ws.cell(i, 1, lab)
            c = ws.cell(i, 2)
            c.value = sh
            c.number_format = "0.0%"
    else:  # bar
        ws.cell(1, 1, "Year")
        for sidx, (name, values) in enumerate(series, start=1):
            ws.cell(1, 1 + sidx, name)
            for ri, v in enumerate(values, start=2):
                ws.cell(ri, 1, categories[ri - 2])
                ws.cell(ri, 1 + sidx, v)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(horizontal="right")


def _build_openpyxl_chart(kind, title, categories, series, colors):
    """Construct and return (chart, workbook) ready for Word packaging."""
    wb = openpyxl.Workbook()
    _write_chart_data(wb, kind, categories, series, colors)
    ws = wb.active

    if kind == "pie":
        chart = PieChart()
        chart.title = title
        labels = Reference(ws, min_col=1, min_row=2,
                           max_row=1 + len(categories))
        data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(categories))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.legend = None
        # per-point fill colours (openpyxl keeps default Office palette when
        # points aren't specified; setting them keeps parity with the pie).
        chart.series[0].data_points = [
            DataPoint(idx=i) for i in range(len(categories))]
        for i in range(len(categories)):
            col = colors[i % len(colors)]
            chart.series[0].data_points[i].graphicalProperties.solidFill = col
    else:  # bar
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = title
        data = Reference(ws, min_col=1, min_row=1,
                         max_col=1 + len(series),
                         max_row=1 + len(categories))
        cats = Reference(ws, min_col=1, min_row=2,
                         max_row=1 + len(categories))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    return chart, wb


def _embedded_workbook_bytes(wb):
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# cache population (Word renders best when each ref carries a value cache,
# matching what Excel itself writes)
# ---------------------------------------------------------------------------

def _num_cache(values, fmt="General"):
    cache = etree.Element(_C + "numCache")
    etree.SubElement(cache, _C + "formatCode").text = fmt
    etree.SubElement(cache, _C + "ptCount", val=str(len(values)))
    for i, v in enumerate(values):
        pt = etree.SubElement(cache, _C + "pt", idx=str(i))
        etree.SubElement(pt, _C + "v").text = _num_str(v)
    return cache


def _str_cache(values):
    cache = etree.Element(_C + "strCache")
    etree.SubElement(cache, _C + "ptCount", val=str(len(values)))
    for i, v in enumerate(values):
        pt = etree.SubElement(cache, _C + "pt", idx=str(i))
        etree.SubElement(pt, _C + "v").text = str(v)
    return cache


def _num_str(v):
    if isinstance(v, float):
        return ("%g" % v)
    return str(v)


def _populate_caches(root, kind, categories, series):
    """Add strCache/numCache to every tx / cat / val ref in the chartSpace.

    ``categories`` is a list of category labels; ``series`` is the series data
    as passed to ``inject_chart`` ([[(name, [..]), ...]] for bar, [..] for pie).
    """
    chart = root.find(_C + "chart")
    if chart is None:
        return
    plot_area = chart.find(_C + "plotArea")
    if plot_area is None:
        return
    # series live inside the *Chart element (barChart / pieChart / ...)
    chart_kinds = ("barChart", "pieChart", "lineChart", "areaChart",
                   "doughnutChart", "radarChart", "scatterChart",
                   "bubbleChart", "ofPieChart", "stockChart",
                   "surfaceChart", "bubbleChart")
    for cname in chart_kinds:
        parent = plot_area.find(_C + cname)
        if parent is None:
            continue
        series_els = parent.findall(_C + "ser")
        for si, ser in enumerate(series_els):
            # tx (series name, single cell)
            tx = ser.find(_C + "tx")
            if tx is not None:
                ref = tx.find(_C + "strRef")
                if ref is not None:
                    ref.append(_str_cache([_series_name(kind, series, si)]))
            # cat (categories) -- fresh cache element per series
            cat = ser.find(_C + "cat")
            if cat is not None:
                numref = cat.find(_C + "numRef")
                if numref is not None:
                    strref = etree.Element(_C + "strRef")
                    f_el = numref.find(_C + "f")
                    if f_el is not None:
                        healthy = etree.Element(_C + "f")
                        healthy.text = f_el.text
                        strref.append(healthy)
                    strref.append(_str_cache(categories))
                    cat.remove(numref)
                    cat.append(strref)
                else:
                    strref = cat.find(_C + "strRef")
                    if strref is not None and strref.find(_C + "strCache") is None:
                        strref.append(_str_cache(categories))
            # val (series values)
            val = ser.find(_C + "val")
            if val is not None:
                numref = val.find(_C + "numRef")
                if numref is not None and numref.find(_C + "numCache") is None:
                    numref.append(_num_cache(_series_values(kind, series, si)))


def _series_name(kind, series, si):
    if kind == "pie":
        return "Share"
    if si < len(series):
        return series[si][0]
    return ""


def _series_values(kind, series, si):
    if kind == "pie":
        return list(series)
    if si < len(series):
        return list(series[si][1])
    return []


# ---------------------------------------------------------------------------
# DrawingML + package plumbing
# ---------------------------------------------------------------------------

def _drawing_inline(chart_r_id, doc_pr_id, name, cx, cy):
    inline = etree.Element(_WP + "inline", distT="0", distB="0",
                           distL="0", distR="0")
    etree.SubElement(inline, _WP + "extent", cx=str(cx), cy=str(cy))
    etree.SubElement(inline, _WP + "effectExtent", l="0", t="0", r="0", b="0")
    doc_pr = etree.SubElement(inline, _WP + "docPr", id=str(doc_pr_id),
                              name=name)
    cpr = etree.SubElement(doc_pr, _WP + "cNvGraphicFramePr")
    etree.SubElement(cpr, _A + "graphicFrameLocks", noGrp="1")
    graphic = etree.SubElement(inline, _A + "graphic")
    gdata = etree.SubElement(graphic, _A + "graphicData", uri=_CHART_URI)
    chart_el = etree.SubElement(gdata, _C + "chart")
    chart_el.set(_R + "id", chart_r_id)
    etree.SubElement(chart_el, _C + "autoTitleDeleted", val="0")
    etree.SubElement(chart_el, _C + "plotVisOnly", val="1")
    etree.SubElement(chart_el, _C + "rAdjH", val="0")
    etree.SubElement(chart_el, _C + "rAdjW", val="0")
    return inline


def inject_chart(paragraph, kind, title, categories, series, colors=None,
                 width_in=6.3, height_in=3.6, name="Chart"):
    """Install a native editable chart into an existing *paragraph*.

    kind: "bar"  -> clustered column (series = [(name,[...]), ...])
          "pie"  -> pie            (series = [share, ...]; categories = labels)

    Returns the document-part rId of the chart part (useful for tests).
    """
    colors = colors or THEME
    chart, wb = _build_openpyxl_chart(kind, title, categories, series, colors)
    # openpyxl builds a namespace-free tree (xmlns applied only on write), so
    # serialise then re-parse to attach real namespaces before we enrich it
    # with value caches that Word needs for reliable rendering.
    raw = etree.tostring(chart._write(), encoding="UTF-8")
    space_root = etree.fromstring(raw)
    _populate_caches(space_root, kind, categories, series)
    chart_space_xml = etree.tostring(
        space_root, xml_declaration=True, encoding="UTF-8",
        standalone=True)
    wb_bytes = _embedded_workbook_bytes(wb)

    doc_part = paragraph.part
    package = doc_part.package

    n = _next_chart_index(package)
    wb_part = Part(
        PackURI("/word/embeddings/Microsoft_Excel_Worksheet%d.xlsx" % n),
        _CT_XLSX, blob=wb_bytes, package=package)
    chart_part = Part(
        PackURI("/word/charts/chart%d.xml" % n),
        _CT_CHART, blob=chart_space_xml, package=package)

    chart_r_id = doc_part.relate_to(chart_part, _REL_CHART)
    chart_part.relate_to(wb_part, _REL_PKG)

    cx = int(width_in * 914400)
    cy = int(height_in * 914400)
    drawing = _drawing_inline(
        chart_r_id, _next_doc_object_id(doc_part), name, cx, cy)

    p = paragraph._p
    run = p.makeelement(_W + "r", {})
    p.append(run)
    run.append(drawing)
    return chart_r_id


def _next_chart_index(package):
    idx = 1
    while True:
        pn = PackURI("/word/charts/chart%d.xml" % idx)
        if pn not in {p.partname for p in package.iter_parts()}:
            return idx
        idx += 1


def _next_doc_object_id(doc_part):
    """Unique drawing (``wp:docPr/@id``) number across the whole package.

    ``DocumentPart.next_id`` only scans the document body, so any drawing in a
    header/footer (e.g. the letterhead) can already own an id the body will
    happily duplicate.  Word rejects documents whose drawing ids collide
    ("found unreadable content"), so walk every XML part to find ids already in
    use and return the next value missing from that set.
    """
    used = set()
    package = doc_part.package
    for part in package.iter_parts():
        element = getattr(part, "_element", None)
        if element is None:
            continue
        for value in element.xpath("//@id"):
            if value.isdigit():
                used.add(int(value))
    candidate = doc_part.next_id
    while candidate in used:
        candidate += 1
    return candidate
