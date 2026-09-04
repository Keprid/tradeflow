#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_quarterly_tables.py
========================

Build the quarterly "Kenya Export Performance" pivot workbooks
(``Exports.xlsx`` and ``Imports.xlsx``) straight from the raw KRA
customs extracts:

  - ``... Exports by HS Destination ...`` .xls
    (Year, Month, Destin, Destination Name, HS_C, SITC_C, SHORT_DESC,
     Net Wt, Quantity, UNIT, CIF Value -- values in KES)
  - ``... Imports by HS Origin ...`` .xls (same layout, Origin instead of
    Destination)

Each report table compares the current quarter with the same quarter of the
previous year, so the extracts should cover both years (e.g. Jan-Jun 2026
plus Apr-Jun 2025). When the previous year is missing it can be synthesised
for testing with ``--synthesize-prev-year`` (random variation around the
current year, fixed seed).

Output workbooks (all values in Ksh. Billion):

  Exports.xlsx   Data | By Partner <Y> | By Partner <Y-1> |
                 By Product <Y> | By Product <Y-1> |
                 Table 1 Top Destination Markets |
                 Table 2 Top Export Products | Balance of Trade
  Imports.xlsx   Data | By Partner <Y> | By Partner <Y-1> |
                 By Product <Y> | By Product <Y-1> |
                 Annex 1 Imports by Supplying Countries |
                 Annex 2 Imports by Products

Report-table layout (comparison):

    Rank | Item | <Y-1> months + Total | <Y> months + Total |
    Change in <Y-1>-<Y> (Total change | %)

Data rules replicated from the manual process:

  * only the shared review months (the quarter, e.g. April-June) are kept;
  * the statistical artefact destination ``AIRCRAFT & SHIPSTORES`` is
    excluded from exports;
  * records without a product description become ``All others
    (non-defined)`` (exports) / ``All others`` (imports);
  * every value is converted from KES to Ksh. Billion (/1e9).

Derived cells (change, %, All-other rows, totals) are live Excel formulas
carrying cached results, so the workbooks recalculate in Excel while the
report generator can still read plain values.

Usage
-----
    python3 make_quarterly_tables.py \
        --excel-dir "EXPORT PERFORMANCE FOR Q2" --out-dir output/quarterly

Dependencies: pandas, numpy, openpyxl, xlrd (for legacy .xls).
"""

import argparse
import gc
import os
import sys

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from make_tables import _put_formula, _cell_ref, _save_workbook  # noqa: E402

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MONTH_NAMES = {1: "January", 2: "February", 3: "March", 4: "April",
               5: "May", 6: "June", 7: "July", 8: "August", 9: "September",
               10: "October", 11: "November", 12: "December"}

EXCLUDED_DESTINATIONS = {"AIRCRAFT & SHIPSTORES"}

UNDEFINED_EXPORT_LABEL = "All others (non-defined)"
UNDEFINED_IMPORT_LABEL = "All others"

FONT = "Century Gothic"
FONT_SIZE = 11
HDR_FILL = PatternFill("solid", fgColor="156082")
HDR_FONT_COLOR = "FFFFFF"
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
ROW_FILL_EVEN = PatternFill("solid", fgColor="F2F7FB")
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="B0B0B0")
DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_VALUE = "#,##0.0"

# green up-arrow for increases, red down-arrow for decreases (Excel renders
# the arrow glyphs; sign is implied by the arrow so no +/- is shown)
FMT_CHANGE = '[Color10]"▲ "#,##0.0;[Red]"▼ "#,##0.0;"— "#,##0.0'
FMT_PCT = '[Color10]"▲ "0.0%;[Red]"▼ "0.0%;"— "0.0%'

KSH_BILLION = "Value in Ksh. Billion"


# ---------------------------------------------------------------------------
# Locating and loading the raw KRA files
# ---------------------------------------------------------------------------
def find_kra_files(excel_dir):
    """Locate the raw KRA extracts by keyword in their filenames.

    Returns ``{"exports": [paths...], "imports": [paths...]}``. Several
    files may match each side (e.g. one extract per year, or a single file
    covering both years) - they are all loaded and merged.
    """
    found = {"exports": [], "imports": []}
    for fname in sorted(os.listdir(excel_dir)):
        low = fname.lower()
        if not low.endswith((".xls", ".xlsx", ".xlsm")):
            continue
        path = os.path.join(excel_dir, fname)
        if "export" in low and "destination" in low:
            found["exports"].append(path)
        elif "import" in low and "origin" in low:
            found["imports"].append(path)
    missing = [k for k in ("exports", "imports") if not found[k]]
    if missing:
        names = ", ".join(sorted(os.listdir(excel_dir)))
        hints = "\n".join(
            "  - %s: filename(s) containing '%s'"
            % (k, "Exports by HS Destination" if k == "exports"
               else "Imports by HS Origin")
            for k in missing)
        sys.exit("[ERROR] Missing KRA raw file(s) in '%s'.\n"
                 "Files present: %s\nExpected:\n%s" % (excel_dir, names, hints))
    return found


def load_side(paths, side):
    """Load and merge every extract of one side (exports or imports)."""
    frames = [load_kra(p, side) for p in paths]
    if len(frames) == 1:
        return frames[0]
    merged = pd.concat(frames, ignore_index=True)
    # the same year/month/partner/product row must not be counted twice
    key = ["Year", "Month", "Partner", "Product"]
    dupes = merged.duplicated(subset=key + ["Value"]).sum()
    exact = merged.duplicated(subset=key).sum()
    if exact and not dupes:
        sys.exit(
            "[ERROR] Conflicting extracts for %s: several files contain "
            "rows for the same Year/Month/Partner/Product with different "
            "values (%d rows). Check that each file covers a distinct "
            "period." % (side, exact))
    merged = merged.drop_duplicates(subset=key + ["Value"])
    return merged


def load_kra(path, side):
    """Read one raw KRA extract into a normalised DataFrame.

    Returns columns: Year (int), Month (int), Partner (str),
    Product (str), Value (float, Ksh. Billion).
    """
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    col_year = pick("Year")
    col_month = pick("Month")
    col_name = pick("Destination Name", "Origin Name", "Destination", "Origin")
    col_desc = pick("SHORT_DESC", "Short Description", "DESCRIPTION")
    col_val = pick("CIF Value", "FOB Value", "Value", "NET VALUE")
    missing = [n for n, c in (("Year", col_year), ("Month", col_month),
                              ("partner", col_name), ("product", col_desc),
                              ("value", col_val)) if c is None]
    if missing:
        sys.exit("[ERROR] %s: could not locate column(s) %s.\nFound: %s"
                 % (os.path.basename(path), ", ".join(missing),
                    ", ".join(df.columns)))

    out = pd.DataFrame()
    out["Year"] = pd.to_numeric(df[col_year], errors="coerce").astype("Int64")
    out["Month"] = pd.to_numeric(df[col_month], errors="coerce").astype("Int64")
    out["Partner"] = df[col_name].astype(str).str.strip().str.upper()
    fallback = (UNDEFINED_EXPORT_LABEL if side == "exports"
                else UNDEFINED_IMPORT_LABEL)
    desc = df[col_desc].where(df[col_desc].notna(), fallback)
    desc = desc.astype(str).str.strip().replace({"": fallback, "nan": fallback})
    out["Product"] = desc
    out["Value"] = pd.to_numeric(df[col_val], errors="coerce").fillna(0.0) / 1e9

    out = out.dropna(subset=["Year", "Month"])
    if side == "exports":
        out = out[~out["Partner"].isin(EXCLUDED_DESTINATIONS)]
    return out


def synthesize_previous_year(frame, months, year, seed=42):
    """Create synthetic rows for ``year - 1`` (testing only).

    Every record of the review quarter is duplicated into the previous year
    with a random factor in [0.70, 1.30] applied to its value (fixed seed ->
    reproducible). Returns a copy of the frame extended with the new rows.
    """
    prev = year - 1
    rng = np.random.RandomState(seed)
    q = frame[frame["Month"].isin(months) & (frame["Year"] == year)]
    if q.empty:
        return frame
    synth = q.copy()
    synth["Year"] = prev
    n = len(synth)
    # per-row factor, mildly autocorrelated within partner to look plausible:
    # draw one factor per partner-month and jitter each row around it
    factors = rng.uniform(0.70, 1.30, size=n)
    synth["Value"] = (synth["Value"].to_numpy() * factors)
    out = pd.concat([frame, synth], ignore_index=True)
    return out


# ---------------------------------------------------------------------------
# Aggregation helpers
# ---------------------------------------------------------------------------
def detect_quarter(frames):
    """Months and review year covered by both extracts.

    The review year is the latest year present in both files.
    """
    months_e = set(frames["exports"]["Month"].dropna().astype(int))
    months_i = set(frames["imports"]["Month"].dropna().astype(int))
    months = sorted(months_e & months_i)
    years_e = set(frames["exports"]["Year"].dropna().astype(int))
    years_i = set(frames["imports"]["Year"].dropna().astype(int))
    years = sorted(years_e & years_i) or sorted(years_e | years_i)
    if len(months) < 2:
        sys.exit("[ERROR] The extracts share fewer than two months; "
                 "cannot define a review period. "
                 "Exports months: %s, imports months: %s"
                 % (sorted(months_e), sorted(months_i)))
    return months, (years[-1] if years else None)


def quarter_label(months):
    """'April-June' style label for a list of month numbers."""
    parts = []
    run_start = prev = months[0]
    for m in months[1:]:
        if m == prev + 1:
            prev = m
            continue
        parts.append((run_start, prev))
        run_start = prev = m
    parts.append((run_start, prev))
    return "-".join(MONTH_NAMES[a] if a == b else
                    "%s-%s" % (MONTH_NAMES[a], MONTH_NAMES[b])
                    for a, b in parts)


def pivot_by(frame, key, months=None, year=None):
    """Pivot one flow into key x month totals plus a __total__ column."""
    f = frame
    if months is not None:
        f = f[f["Month"].isin(months)]
    if year is not None:
        f = f[f["Year"] == year]
    month_list = sorted(f["Month"].dropna().astype(int).unique())
    piv = f.pivot_table(index=key, columns="Month", values="Value",
                        aggfunc="sum")
    piv = piv.reindex(columns=month_list)
    piv["__total__"] = piv.sum(axis=1, skipna=True)
    return piv


def _aligned_vals(piv, name, months):
    """Month values for one item from a pivot (0.0 where missing)."""
    vals = []
    for m in months:
        try:
            v = piv.at[name, m]
        except KeyError:
            v = np.nan
        vals.append(float(v) if pd.notna(v) else 0.0)
    return vals


def rank_table_comparison(cur_piv, prev_piv, months, top_n, all_label):
    """Rows for one comparison report table (current vs previous year).

    Items are ranked by the current-year total. Each row carries both
    years' monthly values and totals plus the change and % change.
    """
    ranked_cur = cur_piv.sort_values("__total__", ascending=False)
    top_names = [str(n).strip() for n in ranked_cur.head(top_n).index]

    def make_row(rank, name, cur_vals, cur_tot, prev_vals, prev_tot):
        chg = cur_tot - prev_tot
        pct = (chg / prev_tot) if prev_tot else None
        return {"rank": rank, "name": name,
                "cur_months": cur_vals, "cur_total": cur_tot,
                "prev_months": prev_vals, "prev_total": prev_tot,
                "change": chg, "pct": pct}

    rows = []
    for rank, name in enumerate(top_names, start=1):
        cur_vals = _aligned_vals(cur_piv, name, months)
        prev_vals = _aligned_vals(prev_piv, name, months)
        rows.append(make_row(
            rank, name, cur_vals,
            float(cur_piv.at[name, "__total__"]),
            prev_vals,
            float(prev_piv.at[name, "__total__"])
            if name in prev_piv.index else sum(prev_vals)))

    def grand_row(name):
        cm = [float(cur_piv[m].sum(skipna=True)) for m in months]
        pm = [float(prev_piv[m].sum(skipna=True)) if m in prev_piv.columns
              else 0.0 for m in months]
        ct = float(cur_piv["__total__"].sum(skipna=True))
        pt = float(prev_piv["__total__"].sum(skipna=True))
        return make_row(None, name, cm, ct, pm, pt)

    all_other = None
    rest = ranked_cur.iloc[top_n:]
    if len(rest):
        ao = grand_row(all_label)
        shown_cur = [sum(r["cur_months"][k] for r in rows)
                     for k in range(len(months))]
        shown_prev = [sum(r["prev_months"][k] for r in rows)
                      for k in range(len(months))]
        st_c, st_p = sum(r["cur_total"] for r in rows), \
            sum(r["prev_total"] for r in rows)
        all_other = grand_row(all_label)
        all_other["cur_months"] = [a - s for a, s
                                   in zip(ao["cur_months"], shown_cur)]
        all_other["prev_months"] = [a - s for a, s
                                    in zip(ao["prev_months"], shown_prev)]
        all_other["cur_total"] -= st_c
        all_other["prev_total"] -= st_p
        all_other["change"] = all_other["cur_total"] - all_other["prev_total"]
        all_other["pct"] = (all_other["change"] / all_other["prev_total"]
                            if all_other["prev_total"] else None)
    grand = grand_row("Grand Total")
    return rows, all_other, grand


def rank_table(piv, top_n, all_label):
    """Rows for one single-year report table: top items + All-other + Grand.

    Used when no previous-year data exists (no comparison columns).
    """
    months = [c for c in piv.columns if c != "__total__"]
    ranked = piv.sort_values("__total__", ascending=False)
    top = ranked.head(top_n)

    rows = []
    grand_total = float(piv["__total__"].sum(skipna=True))
    month_totals = [float(piv[m].sum(skipna=True)) for m in months]
    for rank, (name, r) in enumerate(top.iterrows(), start=1):
        tot = float(r["__total__"])
        rows.append({"rank": rank, "name": str(name).strip(),
                     "months": [float(r[m]) if pd.notna(r[m]) else 0.0
                                for m in months],
                     "total": tot,
                     "share": (tot / grand_total) if grand_total else 0.0})
    all_other = None
    rest = ranked.iloc[top_n:]
    if len(rest):
        shown_months = [sum(r["months"][k] for r in rows)
                        for k in range(len(months))]
        shown_total = sum(r["total"] for r in rows)
        ao_tot = grand_total - shown_total
        all_other = {"rank": None, "name": all_label,
                     "months": [mt - sm for mt, sm
                                in zip(month_totals, shown_months)],
                     "total": ao_tot,
                     "share": (ao_tot / grand_total) if grand_total else 0.0}
    grand = {"rank": None, "name": "Grand Total",
             "months": month_totals, "total": grand_total, "share": 1.0}
    return rows, all_other, grand


# ---------------------------------------------------------------------------
# Workbook styling helpers
# ---------------------------------------------------------------------------
def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_cell(cell, bold=False, size=FONT_SIZE, fill=None, numfmt=None,
               wrap=False, align="center"):
    cell.font = Font(name=FONT, size=size, bold=bold, color="000000")
    cell.border = DATA_BORDER
    if fill is not None:
        cell.fill = fill
    if numfmt is not None:
        cell.number_format = numfmt
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap,
                               indent=1 if align == "left" else 0)


def put(ws, row, col, value, **kw):
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, **kw)
    return cell


def get_fill(row_index):
    return ROW_FILL_EVEN if row_index % 2 == 0 else ROW_FILL_ODD


def put_text(ws, row, col, text, bold=False, fill=None, wrap=False,
             align="left"):
    if fill is None:
        fill = get_fill(row)
    return put(ws, row, col, text, bold=bold, fill=fill, wrap=wrap,
               align=align)


def put_val(ws, row, col, value, bold=False, fill=None, fmt=FMT_VALUE):
    if fill is None:
        fill = get_fill(row)
    return put(ws, row, col, value, bold=bold, fill=fill, numfmt=fmt)


def header_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT, size=FONT_SIZE, bold=True,
                     color=HDR_FONT_COLOR)
    cell.fill = HDR_FILL
    cell.border = DATA_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    return cell


def _putf(ws, cache, row, col, formula, cached, **kw):
    """_put_formula variant registering the cached value per sheet title."""
    put(ws, row, col, formula, **kw)
    cache.setdefault(ws.title, {})[_cell_ref(row, col)] = cached


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
def write_data_sheet(wb, frame, title):
    """Flattened record sheet (Year | Month | Partner | Product | Value).

    Raw extracts run to tens of thousands of rows, so the records are
    written as plain cells -- no fonts/fills/borders/formats. Styling every
    cell costs minutes of CPU and hundreds of MB of RAM inside openpyxl's
    style registry (which made web jobs die mid-request), and buys nothing
    on a raw data dump. Only the header row is styled.
    """
    ws = wb.create_sheet(title[:31])
    heads = ["Year", "Month", "Partner", "Product", KSH_BILLION]
    for c, h in enumerate(heads, start=1):
        header_cell(ws, 1, c, h)
    cell = ws.cell
    r = 2
    for rec in frame.itertuples(index=False):
        cell(row=r, column=1, value=int(rec.Year))
        cell(row=r, column=2, value=int(rec.Month))
        cell(row=r, column=3, value=rec.Partner)
        cell(row=r, column=4, value=rec.Product)
        # rounded so the default General format still reads cleanly
        cell(row=r, column=5, value=round(float(rec.Value), 4))
        r += 1
    set_widths(ws, {"A": 7, "B": 8, "C": 26, "D": 70, "E": 18})
    ws.freeze_panes = "A2"
    return ws


def write_pivot_sheet(wb, piv, title, key_header, cache):
    """Pivot sheet: key x months + Total (formulas for totals)."""
    ws = wb.create_sheet(title[:31])
    months = [c for c in piv.columns if c != "__total__"]
    header_cell(ws, 1, 1, "Sum of Value")
    header_cell(ws, 2, 1, key_header)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    for k, m in enumerate(months):
        header_cell(ws, 2, 2 + k, MONTH_NAMES.get(int(m), str(m)))
    tcol = 2 + len(months)
    header_cell(ws, 2, tcol, "Total")
    r = 3
    for name, row in piv.sort_values("__total__", ascending=False).iterrows():
        put_text(ws, r, 1, str(name).strip())
        for k, m in enumerate(months):
            v = row[m]
            put_val(ws, r, 2 + k, float(v) if pd.notna(v) else None)
        first_cl = get_column_letter(2)
        last_cl = get_column_letter(1 + len(months))
        _putf(ws, cache, r, tcol,
              "=SUM(%s%d:%s%d)" % (first_cl, r, last_cl, r),
              float(row["__total__"]), fill=get_fill(r))
        r += 1
    # grand total row
    put_text(ws, r, 1, "Grand Total", bold=True, fill=TOTAL_FILL)
    for k, m in enumerate(months):
        col_letter = get_column_letter(2 + k)
        tot = float(piv[m].sum(skipna=True))
        _putf(ws, cache, r, 2 + k,
              "=SUM(%s3:%s%d)" % (col_letter, col_letter, r - 1),
              tot, bold=True, fill=TOTAL_FILL)
    col_letter = get_column_letter(tcol)
    _putf(ws, cache, r, tcol,
          "=SUM(%s3:%s%d)" % (col_letter, col_letter, r - 1),
          float(piv["__total__"].sum(skipna=True)),
          bold=True, fill=TOTAL_FILL)
    set_widths(ws, dict([("A", 30)] +
                        [(get_column_letter(2 + k), 13)
                         for k in range(len(months) + 1)]))
    ws.freeze_panes = "B3"
    return ws


def write_report_table_sheet(wb, rows, all_other, grand, title, month_names,
                             item_header, cache, year):
    """Comparison report table (current year vs previous year).

    Layout (columns):
        A Rank | B Item | C.. prev-year months | prev Total |
        cur-year months | cur Total | Total change | %

    Rows: 1 note, 2 group header (years merged / Change merged),
    3 month labels, data from row 4, then All-other and Grand Total.
    Change cells are live formulas (=cur-prev); % is change/prev-total;
    both carry green-up/red-down arrow number formats.
    """
    ws = wb.create_sheet(title[:31])
    n = len(month_names)
    Yp, Yc = year - 1, year

    # column indices -------------------------------------------------------
    c_rank, c_item = 1, 2
    p_first = 3                       # first prev-year month column
    p_total = p_first + n             # prev-year Total
    c_first = p_total + 1             # first cur-year month column
    c_total = c_first + n             # cur-year Total
    c_chg = c_total + 1               # Total change
    c_pct = c_chg + 1                 # %
    ncols = c_pct

    # note row -------------------------------------------------------------
    note = ws.cell(row=1, column=1, value="(%s)" % KSH_BILLION)
    note.font = Font(name=FONT, size=9, italic=True, color="555555")

    # group header (row 2) ---------------------------------------------------
    hdr_r, lbl_r = 2, 3
    ws.merge_cells(start_row=hdr_r, start_column=c_rank,
                   end_row=lbl_r, end_column=c_rank)
    header_cell(ws, hdr_r, c_rank, "Rank")
    ws.merge_cells(start_row=hdr_r, start_column=c_item,
                   end_row=lbl_r, end_column=c_item)
    header_cell(ws, hdr_r, c_item, item_header)
    ws.merge_cells(start_row=hdr_r, start_column=p_first,
                   end_row=hdr_r, end_column=p_total)
    header_cell(ws, hdr_r, p_first, str(Yp))
    ws.merge_cells(start_row=hdr_r, start_column=c_first,
                   end_row=hdr_r, end_column=c_total)
    header_cell(ws, hdr_r, c_first, str(Yc))
    ws.merge_cells(start_row=hdr_r, start_column=c_chg,
                   end_row=hdr_r, end_column=c_pct)
    header_cell(ws, hdr_r, c_chg, "Change in\n%d-%d" % (Yp, Yc))

    # label row (row 3) ------------------------------------------------------
    for base, total_col in ((p_first, p_total), (c_first, c_total)):
        for k, m in enumerate(month_names):
            header_cell(ws, lbl_r, base + k, m)
        header_cell(ws, lbl_r, total_col, "Total")
    header_cell(ws, lbl_r, c_chg, "Total change")
    header_cell(ws, lbl_r, c_pct, "%")

    # data rows ----------------------------------------------------------------
    first_data = 4
    top_last = first_data + len(rows) - 1
    ao_row = top_last + 1 if all_other is not None else None
    grand_row = (ao_row or top_last) + 1

    def cl(col):
        return get_column_letter(col)

    def emit(d, ri, is_ao=False, is_grand=False):
        fill = TOTAL_FILL if (is_ao or is_grand) else get_fill(ri)
        put(ws, ri, c_rank, d["rank"] if d["rank"] is not None else "",
            fill=fill)
        put_text(ws, ri, c_item, d["name"], fill=fill, wrap=True)
        for base, total_col, mvals, tot in (
                (p_first, p_total, d["prev_months"], d["prev_total"]),
                (c_first, c_total, d["cur_months"], d["cur_total"])):
            for k, v in enumerate(mvals):
                if is_ao:
                    _putf(ws, cache, ri, base + k,
                          "=%s%d-SUM(%s%d:%s%d)"
                          % (cl(base + k), grand_row,
                             cl(base + k), first_data,
                             cl(base + k), top_last),
                          v, fill=fill)
                elif is_grand:
                    _putf(ws, cache, ri, base + k,
                          "=SUM(%s%d:%s%d)"
                          % (cl(base + k), first_data,
                             cl(base + k), grand_row - 1),
                          v, bold=True, fill=fill)
                else:
                    put_val(ws, ri, base + k, v, fill=fill)
            if is_ao:
                _putf(ws, cache, ri, total_col,
                      "=%s%d-SUM(%s%d:%s%d)"
                      % (cl(total_col), grand_row,
                         cl(total_col), first_data,
                         cl(total_col), top_last),
                      tot, bold=True, fill=fill)
            elif is_grand:
                _putf(ws, cache, ri, total_col,
                      "=SUM(%s%d:%s%d)"
                      % (cl(total_col), first_data,
                         cl(total_col), grand_row - 1),
                      tot, bold=True, fill=fill)
            else:
                put_val(ws, ri, total_col, tot, fill=fill)
        # change + %
        _putf(ws, cache, ri, c_chg,
              "=%s%d-%s%d" % (cl(c_total), ri, cl(p_total), ri),
              d["change"], numfmt=FMT_CHANGE,
              bold=is_ao or is_grand, fill=fill)
        cached_pct = d["pct"] if d["pct"] is not None else ""
        _putf(ws, cache, ri, c_pct,
              '=IF(%s%d=0,"",%s%d/%s%d)'
              % (cl(p_total), ri, cl(c_chg), ri, cl(p_total), ri),
              cached_pct, numfmt=FMT_PCT,
              bold=is_ao or is_grand, fill=fill)

    data_rows = list(rows) + ([all_other] if all_other is not None else [])
    r = first_data
    for d in rows:
        emit(d, r)
        r += 1
    if all_other is not None:
        emit(all_other, r, is_ao=True)
        r += 1
    emit(grand, r, is_grand=True)

    set_widths(ws, dict([(cl(c_rank), 6), (cl(c_item), 44)] +
                        [(cl(p_first + k), 10) for k in range(n)] +
                        [(cl(p_total), 11)] +
                        [(cl(c_first + k), 10) for k in range(n)] +
                        [(cl(c_total), 11), (cl(c_chg), 13),
                         (cl(c_pct), 10)]))
    ws.freeze_panes = ws.cell(row=first_data, column=c_item + 1)
    return ws


def write_report_table_simple(wb, rows, all_other, grand, title, month_names,
                              item_header, cache):
    """Single-year report table: Rank | Item | months | Total | Share.

    Layout: row 1 header, row 2 unit sub-header, data from row 3,
    then the All-other row and the Grand Total row.
    """
    ws = wb.create_sheet(title[:31])
    n = len(month_names)

    header_cell(ws, 1, 1, "Rank")
    header_cell(ws, 1, 2, item_header)
    for k, m in enumerate(month_names):
        header_cell(ws, 1, 3 + k, m)
    header_cell(ws, 1, 3 + n, "Total")
    header_cell(ws, 1, 4 + n, "Share")

    # unit sub-header row
    unit_row = 2
    put(ws, unit_row, 1, "", fill=HDR_FILL)
    hdr_unit = ws.cell(row=unit_row, column=2, value=KSH_BILLION)
    style_cell(hdr_unit, bold=True, fill=HDR_FILL)
    hdr_unit.font = Font(name=FONT, size=FONT_SIZE, bold=True,
                         color=HDR_FONT_COLOR)
    ws.merge_cells(start_row=unit_row, start_column=2,
                   end_row=unit_row, end_column=2 + n)
    put(ws, unit_row, 3 + n, KSH_BILLION, bold=True, fill=HDR_FILL)
    ws.cell(row=unit_row, column=3 + n).font = Font(
        name=FONT, size=FONT_SIZE, bold=True, color=HDR_FONT_COLOR)
    put(ws, unit_row, 4 + n, "%", bold=True, fill=HDR_FILL)

    first_data = 3
    top_last = first_data + len(rows) - 1          # last ranked row
    ao_row = top_last + 1 if all_other else None   # All-other row
    grand_row = (ao_row or top_last) + 1           # Grand Total row
    tcl = get_column_letter(3 + n)

    def data_fill(row, bold=False):
        return TOTAL_FILL if bold else get_fill(row)

    r = first_data
    data_rows = list(rows) + ([all_other] if all_other else [])
    for d in data_rows:
        is_ao = all_other is not None and d is all_other
        fill = data_fill(r, is_ao)
        put(ws, r, 1, d["rank"] if d["rank"] is not None else "", fill=fill)
        put_text(ws, r, 2, d["name"], fill=fill, wrap=True)
        for k, v in enumerate(d["months"]):
            cl_ = get_column_letter(3 + k)
            if is_ao:
                _putf(ws, cache, r, 3 + k,
                      "=%s%d-SUM(%s%d:%s%d)"
                      % (cl_, grand_row, cl_, first_data, cl_, top_last),
                      v, fill=fill)
            else:
                put_val(ws, r, 3 + k, v, fill=fill)
        if is_ao:
            _putf(ws, cache, r, 3 + n,
                  "=%s%d-SUM(%s%d:%s%d)"
                  % (tcl, grand_row, tcl, first_data, tcl, top_last),
                  d["total"], bold=True, fill=fill)
        else:
            put_val(ws, r, 3 + n, d["total"], fill=fill)
        r += 1

    put(ws, grand_row, 1, "", fill=TOTAL_FILL)
    put(ws, grand_row, 2, grand["name"], bold=True, fill=TOTAL_FILL)
    last_data = grand_row - 1
    for k in range(n):
        cl_ = get_column_letter(3 + k)
        _putf(ws, cache, grand_row, 3 + k,
              "=SUM(%s%d:%s%d)" % (cl_, first_data, cl_, last_data),
              grand["months"][k], bold=True, fill=TOTAL_FILL)
    _putf(ws, cache, grand_row, 3 + n,
          "=SUM(%s%d:%s%d)" % (tcl, first_data, tcl, last_data),
          grand["total"], bold=True, fill=TOTAL_FILL)

    share_col = 4 + n
    gref = "$%s$%d" % (tcl, grand_row)
    r = first_data
    for d in data_rows:
        is_ao = all_other is not None and d is all_other
        fill = data_fill(r, is_ao)
        cl_ = get_column_letter(3 + n)
        _putf(ws, cache, r, share_col,
              "=IF(%s=0,0,%s/%s)" % (gref, cl_, gref),
              d["share"], bold=is_ao, numfmt="0.0%", fill=fill)
        r += 1
    _putf(ws, cache, grand_row, share_col,
          "=IF(%s=0,0,%s/%s)" % (gref, gref, gref),
          1.0, bold=True, numfmt="0.0%", fill=TOTAL_FILL)

    set_widths(ws, dict([("A", 6), ("B", 46)] +
                        [(get_column_letter(3 + k), 11)
                         for k in range(n)] +
                        [(get_column_letter(3 + n), 11),
                         (get_column_letter(4 + n), 9)]))
    ws.freeze_panes = "C3"
    return ws


def write_balance_sheet(wb, exp_piv_cur, imp_piv_cur, exp_piv_prev,
                        imp_piv_prev, months, cache, year):
    """Monthly Exports / Imports / Balance blocks (prev year if given)."""
    ws = wb.create_sheet("Balance of Trade")

    def block(start, y, ep, ip):
        header_cell(ws, start, 1, str(y))
        for k, m in enumerate(months):
            header_cell(ws, start, 2 + k, MONTH_NAMES.get(int(m), str(m)))
        labels = ["Exports", "Imports", "Bal. of Trade"]
        for i, lbl in enumerate(labels):
            put_text(ws, start + 1 + i, 1, lbl, bold=(i == 2))
        for k, m in enumerate(months):
            e = float(ep[m].sum(skipna=True)) if m in ep.columns else 0.0
            i_ = float(ip[m].sum(skipna=True)) if m in ip.columns else 0.0
            put_val(ws, start + 1, 2 + k, e)
            put_val(ws, start + 2, 2 + k, i_)
            c = get_column_letter(2 + k)
            _putf(ws, cache, start + 3, 2 + k, "=%s%d-%s%d"
                  % (c, start + 1, c, start + 2), e - i_, bold=True)
        put_text(ws, start + 4, 1, "Quarter Balance", bold=True)
        e_tot = sum(float(ep[m].sum(skipna=True))
                    for m in months if m in ep.columns)
        i_tot = sum(float(ip[m].sum(skipna=True))
                    for m in months if m in ip.columns)
        _putf(ws, cache, start + 4, 2,
              "=SUM(B%d:%s%d)" % (start + 3,
                                  get_column_letter(1 + len(months)),
                                  start + 3),
              e_tot - i_tot, bold=True)

    block(2, year, exp_piv_cur, imp_piv_cur)
    if exp_piv_prev is not None and imp_piv_prev is not None:
        block(8, year - 1, exp_piv_prev, imp_piv_prev)
    set_widths(ws, dict([("A", 16)] +
                        [(get_column_letter(2 + k), 14)
                         for k in range(len(months))]))
    return ws


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def _has_prev_quarter(frame, months, prev_year):
    """True when the frame holds rows for the review months of prev_year."""
    return bool(len(frame[(frame["Year"] == prev_year)
                          & (frame["Month"].isin(months))]))


def generate_quarterly_tables(excel_dir, out_dir,
                              top_markets=25, top_products=20,
                              synthesize_prev=False, seed=42):
    """Build Exports.xlsx + Imports.xlsx from the raw KRA extracts.

    Several files per side are allowed and merged, so the review year and
    the previous year may arrive as one file each (four uploads) or as a
    single extract covering both years.

    When ``synthesize_prev`` is true and the previous year is absent,
    synthetic prior-year rows are generated (random factors, fixed seed)
    so the comparison layout can be tested.

    When no previous-year data exists (and none is synthesised), the
    workbooks fall back to the single-year layout (Rank | Item | months |
    Total | Share) without comparison columns. The returned metadata dict
    reports this in ``comparison``.
    """
    files = find_kra_files(excel_dir)
    frames = {"exports": load_side(files["exports"], "exports"),
              "imports": load_side(files["imports"], "imports")}
    months, year = detect_quarter(frames)
    prev_year = year - 1
    synthesized = []
    if synthesize_prev:
        for side, fr in frames.items():
            if not _has_prev_quarter(fr, months, prev_year):
                frames[side] = synthesize_previous_year(fr, months, year,
                                                        seed=seed)
                synthesized.append(side)

    comparison = all(_has_prev_quarter(frames[s], months, prev_year)
                     for s in ("exports", "imports"))
    qlabel = quarter_label(months)
    month_names = [MONTH_NAMES.get(int(m), str(m)) for m in months]

    os.makedirs(out_dir, exist_ok=True)

    # slice out the review quarter from both sides, then drop the full
    # extracts -- everything below works on the quarter frames only
    exp_q = frames["exports"][frames["exports"]["Month"].isin(months)]
    imp_q = frames["imports"][frames["imports"]["Month"].isin(months)]
    frames.clear()
    gc.collect()

    # ---------------- exports workbook ----------------
    exp_cur = exp_q[exp_q["Year"] == year]
    exp_prev = (exp_q[exp_q["Year"] == prev_year] if comparison else None)

    wb_e = openpyxl.Workbook()
    wb_e.remove(wb_e.active)
    cache_e = {}
    write_data_sheet(wb_e, exp_q.sort_values(["Year", "Month", "Partner"]),
                     "Data")
    piv_ep_cur = pivot_by(exp_cur, "Partner")
    piv_ep_prev = (pivot_by(exp_prev, "Partner", months=months)
                   if comparison else None)
    piv_et_cur = pivot_by(exp_cur, "Product")
    piv_et_prev = (pivot_by(exp_prev, "Product", months=months)
                   if comparison else None)
    write_pivot_sheet(wb_e, piv_ep_cur, "By Partner %d" % year,
                      "Destination", cache_e)
    write_pivot_sheet(wb_e, piv_et_cur, "By Product %d" % year,
                      "Product label", cache_e)
    if comparison:
        write_pivot_sheet(wb_e, piv_ep_prev, "By Partner %d" % prev_year,
                          "Destination", cache_e)
        write_pivot_sheet(wb_e, piv_et_prev, "By Product %d" % prev_year,
                          "Product label", cache_e)

    t1 = rank_table_comparison(
        piv_ep_cur, piv_ep_prev, months, top_markets, "All Other Countries") \
        if comparison else rank_table(piv_ep_cur, top_markets,
                                      "All Other Countries")
    t2 = rank_table_comparison(
        piv_et_cur, piv_et_prev, months, top_products, "All other products") \
        if comparison else rank_table(piv_et_cur, top_products,
                                      "All other products")

    imp_cur = imp_q[imp_q["Year"] == year]
    imp_prev = (imp_q[imp_q["Year"] == prev_year] if comparison else None)
    piv_ip_cur = pivot_by(imp_cur, "Partner")
    piv_ip_prev = (pivot_by(imp_prev, "Partner", months=months)
                   if comparison else None)
    if comparison:
        write_balance_sheet(wb_e, piv_ep_cur, piv_ip_cur,
                            piv_ep_prev, piv_ip_prev, months, cache_e, year)
    else:
        write_balance_sheet(wb_e, piv_ep_cur, piv_ip_cur,
                            None, None, months, cache_e, year)

    def emit_table(wb_, cache_, data, title, header):
        if comparison:
            write_report_table_sheet(wb_, data[0], data[1], data[2],
                                     title, month_names, header, cache_,
                                     year)
        else:
            write_report_table_simple(wb_, data[0], data[1], data[2],
                                      title, month_names, header, cache_)

    emit_table(wb_e, cache_e, t1,
               "Table 1 Top Destination Markets", "PARTNER")
    emit_table(wb_e, cache_e, t2,
               "Table 2 Top Export Products", "Product label")
    wb_e.properties.title = "Kenya Export Performance %s %d" % (qlabel, year)
    _save_workbook(wb_e, os.path.join(out_dir, "Exports.xlsx"), cache_e)
    del wb_e, cache_e
    gc.collect()   # release the exports cells before the imports build

    # ---------------- imports workbook ----------------
    wb_i = openpyxl.Workbook()
    wb_i.remove(wb_i.active)
    cache_i = {}
    write_data_sheet(wb_i, imp_q.sort_values(["Year", "Month", "Partner"]),
                     "Data")
    piv_it_cur = pivot_by(imp_cur, "Product")
    piv_it_prev = (pivot_by(imp_prev, "Product", months=months)
                   if comparison else None)
    write_pivot_sheet(wb_i, piv_ip_cur, "By Partner %d" % year, "PARTNER",
                      cache_i)
    write_pivot_sheet(wb_i, piv_it_cur, "By Product %d" % year,
                      "Product Label", cache_i)
    if comparison:
        write_pivot_sheet(wb_i, piv_ip_prev, "By Partner %d" % prev_year,
                          "PARTNER", cache_i)
        write_pivot_sheet(wb_i, piv_it_prev, "By Product %d" % prev_year,
                          "Product Label", cache_i)
    a1 = rank_table_comparison(
        piv_ip_cur, piv_ip_prev, months, top_markets, "All other countries") \
        if comparison else rank_table(piv_ip_cur, top_markets,
                                      "All other countries")
    a2 = rank_table_comparison(
        piv_it_cur, piv_it_prev, months, top_products, "All other products") \
        if comparison else rank_table(piv_it_cur, top_products,
                                      "All other products")
    if comparison:
        write_report_table_sheet(wb_i, a1[0], a1[1], a1[2],
                                 "Annex 1 Imports by Partner", month_names,
                                 "PARTNER", cache_i, year)
        write_report_table_sheet(wb_i, a2[0], a2[1], a2[2],
                                 "Annex 2 Imports by Products", month_names,
                                 "Product Label", cache_i, year)
    else:
        write_report_table_simple(wb_i, a1[0], a1[1], a1[2],
                                  "Annex 1 Imports by Partner", month_names,
                                  "PARTNER", cache_i)
        write_report_table_simple(wb_i, a2[0], a2[1], a2[2],
                                  "Annex 2 Imports by Products", month_names,
                                  "Product Label", cache_i)
    wb_i.properties.title = "Kenya Import Performance %s %d" % (qlabel, year)
    _save_workbook(wb_i, os.path.join(out_dir, "Imports.xlsx"), cache_i)
    del wb_i, cache_i
    gc.collect()

    return {"months": months, "year": year, "prev_year": prev_year,
            "quarter": qlabel, "month_names": month_names,
            "synthesized": synthesized, "comparison": comparison}


def main():
    ap = argparse.ArgumentParser(
        description="Build quarterly Exports.xlsx / Imports.xlsx pivot "
                    "workbooks from raw KRA extracts.")
    ap.add_argument("--excel-dir", default=os.path.join(
        BASE_DIR, "EXPORT PERFORMANCE FOR Q2"),
        help="Folder holding the raw KRA extracts")
    ap.add_argument("--out-dir", default=os.path.join(BASE_DIR, "output",
                                                      "quarterly"),
        help="Folder to write Exports.xlsx / Imports.xlsx into")
    ap.add_argument("--top-markets", type=int, default=25)
    ap.add_argument("--top-products", type=int, default=20)
    ap.add_argument("--synthesize-prev-year", action="store_true",
                    help="Generate synthetic previous-year-quarter rows "
                         "(random variation, fixed seed) when the extracts "
                         "do not already contain them. For testing only.")
    ap.add_argument("--seed", type=int, default=42,
                    help="Random seed used with --synthesize-prev-year")
    args = ap.parse_args()

    print("[1/4] Reading KRA extracts from : %s"
          % os.path.abspath(args.excel_dir))
    meta = generate_quarterly_tables(
        args.excel_dir, args.out_dir,
        top_markets=args.top_markets, top_products=args.top_products,
        synthesize_prev=args.synthesize_prev_year, seed=args.seed)
    if meta["comparison"]:
        print("[2/4] Review period            : %s %d (vs %s %d)"
              % (meta["quarter"], meta["year"],
                 meta["quarter"], meta["year"] - 1))
    else:
        print("[2/4] Review period            : %s %d "
              "(single year - no comparison data found)"
              % (meta["quarter"], meta["year"]))
    if meta["synthesized"]:
        print("[3/4] Synthesized %s data for %d (seed %d)"
              % ("/".join(meta["synthesized"]), meta["year"] - 1, args.seed))
    print("[4/4] Workbooks written to     : %s" % os.path.abspath(args.out_dir))


if __name__ == "__main__":
    main()
