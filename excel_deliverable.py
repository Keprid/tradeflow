"""Build an editable Excel workbook deliverable from a report Analysis.

Produces one workbook per report with:
  * every report table as a plain, editable cell range (NO AutoFilter, so
    no filter-dropdown buttons appear when rows are copied out), styled with
    a thick outside border and a light interior grid;
  * native Excel charts (editable in Excel, not images) for the pie charts
    that would otherwise be raster figures.

Values are written as real numbers (so cells stay editable/recalculable) with
the same display formats as the printed report.  Nothing here parses source
workbooks or the docx: it reads the already-computed ``Analysis``/``cfg``.
"""

import os

import openpyxl
from openpyxl.chart import (PieChart, DoughnutChart, BarChart, Reference,
                            Series)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from country_names import short_product_name
from generate_report import num, pct, clean_label

# House palette (matches the printed report / make_tables.py).
NAVY = "1F3864"
GRID_COLOR = "D6DCE4"
THICK_COLOR = "1F3864"
HDR_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
HDR_FONT = Font(name="Century Gothic", size=11, bold=True, color="FFFFFF")
THIN = Side(style="thin", color=GRID_COLOR)
THICK = Side(style="thick", color=THICK_COLOR)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THEME = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
         "70AD47", "FF0000", "7030A0", "00B0F0", "F7A7A7"]


def _apply_borders(ws, min_row, min_col, max_row, max_col, n_cols):
    """Thick outside border + thin interior grid over the given range."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            b = cell.border
            cell.border = Border(
                left=b.left, right=b.right, top=b.top, bottom=b.bottom)
    # Overwrite edges: outer = thick, inner = thin.
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            is_top = r == min_row
            is_bottom = r == max_row
            is_left = c == min_col
            is_right = c == max_col
            cell.border = Border(
                left=THICK if is_left else THIN,
                right=THICK if is_right else THIN,
                top=THICK if is_top else THIN,
                bottom=THICK if is_bottom else THIN)


def _title_row(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = Font(name="Century Gothic", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row, i)
        cell.value = h
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = Border(top=THICK, bottom=THIN, left=THIN, right=THIN)
        ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1


def _iter_table_rows(table):
    """Yield display rows: headerless items/all-other/total in order."""
    return table["data"]


def _value_cell(cell, v, fmt):
    if v is None:
        cell.value = ""
        return
    cell.value = v
    cell.number_format = fmt


def _fit_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_rank_table(ws, table, years, title, header, widths, flow_label,
                      unit_row=False, total_header=None):
    """Write one parsed rank table (Table 1-6) as an editable range.

    ``header`` lists the plain header cells that precede any merged content;
    here we keep a flat, copy-friendly header: Rank, Code (if has_code),
    label/partner, one column per year, then Share.
    """
    has_code = bool(table.get("has_code"))
    n_years = len(years)
    col0 = 1  # rank
    col_code = 2 if has_code else None
    col_label = (col_code + 1) if has_code else 2
    col_share = col_label + n_years

    r = 1
    if title:
        span = col_share
        r = _title_row(ws, r, title, span)
    hstart = r
    hdr_cells = ["Rank", "Code", header]
    # one header cell per year
    r2 = r + 1

    # --- header ---
    _put = ws.cell
    # Rank merged across 2 rows
    ws.merge_cells(start_row=r, start_column=col0, end_row=r2, end_column=col0)
    _put(r, col0).value = "Rank in %s" % years[-1]
    if col_code is not None:
        ws.merge_cells(start_row=r, start_column=col_code, end_row=r2,
                       end_column=col_code)
        _put(r, col_code).value = "Code"
    ws.merge_cells(start_row=r, start_column=col_label,
                   end_row=r, end_column=col_label + len(years) - 1)
    _put(r, col_label).value = flow_label
    for i, y in enumerate(years):
        _put(r2, col_label + i).value = y
    ws.merge_cells(start_row=r, start_column=col_share, end_row=r2,
                   end_column=col_share)
    _put(r, col_share).value = "Share in %s %%" % years[-1]

    # style header
    for rr in (r, r2):
        for cc in range(col0, col_share + 1):
            cell = ws.cell(rr, cc)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = Border(top=THICK, bottom=THIN,
                                 left=THIN, right=THIN)

    # --- unit row (optional) ---
    rbody = r2 + 1
    body_start = None
    n_years = len(years)
    last_col = col_share

    rows = _iter_table_rows(table)
    first_body = None
    for d in rows:
        body_start = body_start if body_start is not None else rbody
        is_total = d["kind"] == "total"
        is_all = d["kind"] == "all_other"
        b = is_total or is_all
        rw = rbody
        _value_cell(_put(rw, col0), d["rank"], "0")
        if col_code is not None:
            _put(rw, col_code).value = d.get("code")
        nm = clean_label(d.get("label") or d.get("name"))
        name_cell = _put(rw, col_label)
        name_cell.value = nm or ""
        name_cell.alignment = LEFT
        for i, y in enumerate(years):
            v = d["years"][i] if i < len(d.get("years") or []) else None
            vc = _put(rw, col_label + i)
            _value_cell(vc, v if v is not None else "", "#,##0.0")
            vc.alignment = RIGHT if isinstance(v, (int, float)) else CENTER
        sc = _put(rw, col_share)
        if d.get("share") is not None:
            sc.value = round(d["share"], 6)
            sc.number_format = "0.0%"
        sc.alignment = CENTER
        for cc in range(col0, last_col + 1):
            cell = ws.cell(rw, cc)
            if cell.font is None or not cell.font.bold:
                if b:
                    cell.font = Font(bold=True)
        if first_body is None:
            first_body = rw
        rbody += 1

    body_end = rbody - 1 if first_body is not None else r2

    # column widths
    _lens = [len(clean_label(d.get("label") or d.get("name")))
             for d in _iter_table_rows(table)]
    label_w = 28
    if _lens:
        label_w = max(28, min(60, int(max(_lens) * 0.6) + 6))
    wlist = [8, 9] + [label_w] if has_code else [8, label_w]
    wlist += [10] * n_years
    wlist += [12]
    _fit_col_widths(ws, wlist)

    _apply_borders(ws, r, col0, body_end, last_col, n_cols=last_col)
    return {"first": r, "last_col": last_col, "body_end": body_end,
            "share_col": col_share}


def _make_table_sheet(wb, title, table, years, flow_label, unit_row=False):
    ws = wb.create_sheet(title)
    return _write_rank_table(ws, table, years, title, "", [], flow_label,
                             unit_row=unit_row)


def _pie_data_rows(items, n=10):
    """Short product labels + shares for the top ``n`` items."""
    out = []
    for d in (items or [])[:n]:
        if d.get("share") is None:
            continue
        lab = short_product_name(d.get("label") or d.get("name"),
                                 d.get("code"), maxlen=42)
        out.append((lab, round(d["share"], 6)))
    return out


def _pie_data_rows_series(items, n=10):
    """Pie rows derived from the latest non-empty year when 'share' is blank.

    Mirrors ``charts.series_shares``: falls back to the most recent year whose
    values are populated and normalises them to a fraction.
    """
    items = (items or [])[:n]
    if not items:
        return []
    parsed = [d.get("share") for d in items]
    if any(s is not None for s in parsed):
        return _pie_data_rows(items, n)
    rows = [d.get("years") or [] for d in items]
    ncol = max((len(r) for r in rows), default=0)
    for k in range(ncol - 1, -1, -1):
        vals = [r[k] if k < len(r) and r[k] else 0.0 for r in rows]
        total = sum(vals)
        if total > 0:
            out = []
            for d, v in zip(items, vals):
                lab = short_product_name(d.get("label") or d.get("name"),
                                         d.get("code"), maxlen=42)
                out.append((lab, round(v / total, 6)))
            return out
    return []


def _add_pie(ws, top_row, title, labels_values):
    # write data table
    start = top_row
    # header
    ws.cell(start, 1).value = "Category"
    ws.cell(start, 2).value = "Share"
    for cc in (1, 2):
        c = ws.cell(start, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    for i, (lab, v) in enumerate(labels_values, start=start + 1):
        ws.cell(i, 1).value = lab
        cv = ws.cell(i, 2)
        cv.value = v
        cv.number_format = "0.0%"
    last = start + len(labels_values)

    data = Reference(ws, min_col=2, min_row=start + 1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=last)
    pie = PieChart()
    pie.title = title
    pie.width = 14
    pie.height = 8
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(cats)
    pie.series[0].data_points = [
        DataPoint(idx=i) for i in range(len(labels_values))]
    for i in range(len(labels_values)):
        pie.series[0].data_points[i].graphicalProperties.solidFill = \
            THEME[i % len(THEME)]
    pie.legend = None
    ws.add_chart(pie, "D%d" % start)
    return pie


def _add_balance_bar(ws, top_row, a):
    years = a.balance["years"]
    exports = a.balance["exports"]
    imports = a.balance["imports"]
    balance = a.balance["balance"]
    r0 = top_row
    ws.cell(r0, 1).value = "Year"
    ws.cell(r0, 2).value = "Exports"
    ws.cell(r0, 3).value = "Imports"
    ws.cell(r0, 4).value = "Balance"
    for cc in range(1, 5):
        c = ws.cell(r0, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    for i, y in enumerate(years):
        rr = r0 + 1 + i
        ws.cell(rr, 1).value = y
        for cc, series in ((2, exports), (3, imports), (4, balance)):
            v = series[i] if i < len(series) else None
            c = ws.cell(rr, cc)
            c.value = v if v is not None else ""
            c.number_format = "#,##0.0"
            c.alignment = RIGHT

    n = len(years)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Kenya – %s Balance of Trade (USD Million)" % a.country
    chart.height = 9
    chart.width = 18
    data = Reference(ws, min_col=1, min_row=r0, max_col=4, max_row=r0 + n)
    cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r0 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A%d" % (r0 + n + 2))


def build_deliverable(a, cfg, out_path):
    """Write the editable Excel workbook for the report."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    c = cfg["country"]
    country = a.country
    years = a.years

    sheets = [
        ("Table 1 - Import Source Markets", a.table1, c["short"], False),
        ("Table 2 - Import Products from the World", a.table2, c["short"], True),
        ("Table 3 - Export Destination Markets", a.table3, c["short"], False),
        ("Table 4 - Export Products to the World", a.table4, c["short"], True),
        ("Table 5 - Kenya Exports to %s" % country, a.table5, country, True),
        ("Table 6 - Kenya Imports from %s" % country, a.table6, country, True),
    ]
    for title, table, hdr, has_code_flag in sheets:
        if not table:
            continue
        ws = wb.create_sheet(title[:31])
        flow = "Importing Countries from %s" % country if not has_code_flag \
            and "Import Source" in title else \
            ("Exports to %s" % hdr if "Export Dest" in title else hdr)
        _write_rank_table(
            ws, table, years, title,
            "Market" if not table.get("has_code") else "Product",
            [], flow,
            unit_row=not table.get("has_code"))

    if a.table7:
        _write_table7_sheet(wb, a)

    _write_charts_sheet(wb, a)

    wb.save(out_path)
    return out_path


def _write_table7_sheet(wb, a):
    ws = wb.create_sheet("Table 7 - Market Alignment")
    t7 = a.table7
    years = t7.get("years") or a.years
    n = len(years)
    rows = t7.get("data") or []
    r = _title_row(ws, 1, "Table 7 - Market Alignment", 3 + n)
    hr = r
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=2 + n)
    ws.cell(r, 3).value = "Kenya exports to %s (USD Million)" % a.country
    ws.cell(r, 1).value = "Rank"
    ws.cell(r, 2).value = "Country"
    ws.cell(r, 3 + n).value = "Share in %s %%" % years[-1]
    ws.merge_cells(start_row=r, start_column=2 + n, end_row=r,
                   end_column=3 + n)
    for cc in range(1, 4 + n):
        cell = ws.cell(hr, cc)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = Border(top=THICK, bottom=THIN, left=THIN, right=THIN)
    rr = hr + 1
    first = None
    for d in rows:
        if first is None:
            first = rr
        ws.cell(rr, 1).value = d.get("rank")
        ws.cell(rr, 2).value = d.get("name")
        for i, y in enumerate(years):
            v = d.get("years") or []
            c = ws.cell(rr, 3 + i)
            val = v[i] if i < len(v) else None
            c.value = val if val is not None else ""
            c.number_format = "#,##0.0"
        sc = ws.cell(rr, 3 + n)
        if d.get("share") is not None:
            sc.value = round(d["share"], 6)
            sc.number_format = "0.0%"
        for cc in range(1, 4 + n):
            ws.cell(rr, cc).alignment = CENTER
            ws.cell(rr, cc).border = Border(left=THIN, right=THIN,
                                            top=THIN, bottom=THIN)
        ws.cell(rr, 2).alignment = LEFT
        rr += 1
    last = rr - 1 if first is not None else hr
    _apply_borders(ws, hr, 1, last, 3 + n, 3 + n)
    _fit_col_widths(ws, [8, 30] + [11] * n + [12])


def _add_doughnut(ws, top_row, title, labels_values, colors=None,
                  hole_size=50, offset_anchor="C"):
    """Editable doughnut of ``labels_values`` (label, fraction)."""
    start = top_row
    ws.cell(start, 1).value = "Category"
    ws.cell(start, 2).value = "Share"
    for cc in (1, 2):
        c = ws.cell(start, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    for i, (lab, v) in enumerate(labels_values, start=start + 1):
        ws.cell(i, 1).value = lab
        cv = ws.cell(i, 2)
        cv.value = v
        cv.number_format = "0.0%"
    last = start + len(labels_values)

    data = Reference(ws, min_col=2, min_row=start + 1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=last)
    chart = DoughnutChart()
    chart.title = title
    chart.holeSize = hole_size
    chart.width = 14
    chart.height = 8
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.numFmt = "0.0%"
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    for i in range(len(labels_values)):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = (colors or THEME)[i % len(colors or THEME)]
        chart.series[0].data_points.append(dp)
    chart.legend = None
    ws.add_chart(chart, "%s%d" % (offset_anchor, start))
    return chart


def _add_bar_generic(ws, top_row, title, cats, series, colors=None,
                     offset_anchor="A", grouping="clustered", stack=False):
    """Editable column chart; ``series`` = [(name, [values, ...]), ...]."""
    r0 = top_row
    ws.cell(r0, 1).value = "Category"
    for i, (name, _vals) in enumerate(series, start=2):
        ws.cell(r0, i).value = name
    n = max((len(v) for _, v in series), default=0)
    for i, c in enumerate(cats, start=1):
        rr = r0 + i
        ws.cell(rr, 1).value = c
        for si, (_name, vals) in enumerate(series, start=2):
            v = vals[i - 1] if i - 1 < len(vals) else None
            cell = ws.cell(rr, si)
            cell.value = v if v is not None else ""
            cell.number_format = "#,##0.0"
    for cc in range(1, 2 + len(series)):
        c = ws.cell(r0, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER

    chart = BarChart()
    chart.type = "col"
    chart.grouping = grouping
    chart.title = title
    chart.height = 9
    chart.width = 18
    data = Reference(ws, min_col=1, min_row=r0, max_col=1 + len(series),
                     max_row=r0 + n)
    cats_ref = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r0 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    if colors:
        for i, s in enumerate(chart.series):
            s.graphicalProperties.solidFill = colors[i % len(colors)]
    ws.add_chart(chart, "%s%d" % (offset_anchor, r0 + n + 2))
    return chart


def _add_series_bar(ws, top_row, title, cats, values, offset_anchor="A",
                    colors=None):
    return _add_bar_generic(ws, top_row, title, cats,
                            [("Value", values or [])], colors=colors,
                            offset_anchor=offset_anchor)


def _region_bar_series(excel_path):
    """Grouped-bar data from the Table 7/8 regional workbooks."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        val = None
        try:
            val = float(row[3])
        except (TypeError, ValueError):
            val = None
        if val is None:
            continue
        rows.append((str(row[0]).strip(), str(row[2]).strip(), val))
    regions = sorted({r for r, _, _ in rows},
                     key=lambda r: -sum(v for rr, _, v in rows if rr == r))
    ranks = {}
    for r, _n, val in rows:
        ranks.setdefault(r, []).append(val)
    series = []
    for rank in range(3):
        vals = []
        for r in regions:
            entries = sorted(ranks.get(r, []), reverse=True)
            vals.append(round(entries[rank], 1) if rank < len(entries) else None)
        series.append(("Rank %d" % (rank + 1), vals))
    return regions, series


def _write_charts_sheet(wb, a):
    # One sheet per chart so data tables, charts and titles never overlap.
    if a.balance and a.balance.get("years"):
        ws = wb.create_sheet("Figure 1 - Balance of Trade")
        _add_balance_bar(ws, 1, a)

    exp = _pie_data_rows(a.table5.get("items") if a.table5 else [], 10)
    if exp:
        ws = wb.create_sheet("Figure 2 - Top Exports")
        _add_pie(ws, 1, "Share of Kenya's Top Exports to %s in %s"
                 % (a.country, a.years[-1]), exp)

    imp = _pie_data_rows(a.table6.get("items") if a.table6 else [], 10)
    if imp:
        ws = wb.create_sheet("Figure 3 - Top Imports")
        _add_pie(ws, 1, "Share of Kenya's Top Imports from %s in %s"
                 % (a.country, a.years[-1]), imp)


# ---------------------------------------------------------------------------
# Services report deliverable (shared goods-style table shape)
# ---------------------------------------------------------------------------
def build_services_deliverable(a, cfg, out_path, excel_dir=None):
    """Editable Excel workbook for the services trade-flow report (Tables 1-4
    + editable balance, category-share, regional, RCA, concentration,
    diversification and composition charts).  Reuses the goods table writers
    because services tables parse to the same shape."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    years = a.years or [a.year]
    specs = [
        ("Table 1 - Global Service Exporters", a.table1,
         "Service exporters (USD billion)", False),
        ("Table 2 - Global Service Importers", a.table2,
         "Service importers (USD billion)", False),
        ("Table 3 - Kenya Service Exports by Category", a.table3,
         "Service export category", True),
        ("Table 4 - Kenya Service Imports by Category", a.table4,
         "Service import category", True),
    ]
    for title, table, flow, is_prod in specs:
        if not table:
            continue
        ws = wb.create_sheet(title[:31])
        _write_rank_table(
            ws, table, years, title,
            "Category" if is_prod else "Country / Partner",
            [], flow, unit_row=not is_prod)

    if a.balance and a.balance.get("years"):
        ws = wb.create_sheet("Figure 3 - Services Balance of Trade")
        _add_balance_bar(ws, 1, a)

    exp = _pie_data_rows_series(a.table3.get("items") if a.table3 else [], 10)
    if exp:
        ws = wb.create_sheet("Figure 1 - Service Export Shares")
        _add_doughnut(ws, 1, "Share of Kenya's Service Exports by Category in %s"
                      % a.year, exp, colors=[c.lstrip("#") for c in THEME])
    imp = _pie_data_rows_series(a.table4.get("items") if a.table4 else [], 10)
    if imp:
        ws = wb.create_sheet("Figure 2 - Service Import Shares")
        _add_doughnut(ws, 1, "Share of Kenya's Service Imports by Category in %s"
                      % a.year, imp, colors=[c.lstrip("#") for c in THEME])

    if excel_dir and os.path.isdir(excel_dir):
        for sheet, fname, caption, title in (
                ("Figure 6 - Service Exporters by Region",
                 "Table 7 Service Exporters by Region.xlsx",
                 "Figure 6: Top Service Exporters by Region",
                 "Top Service Exporters by Region (USD Billion)"),
                ("Figure 7 - Service Importers by Region",
                 "Table 8 Service Importers by Region.xlsx",
                 "Figure 7: Top Service Importers by Region",
                 "Top Service Importers by Region (USD Billion)")):
            p = os.path.join(excel_dir, fname)
            if os.path.exists(p):
                cats, series = _region_bar_series(p)
                ws = wb.create_sheet(sheet)
                _add_bar_generic(ws, 1, title, cats, series,
                                 colors=["2E75B6", "ED7D31", "70AD47"])

    rca_items = [d for d in (a.rca or {}).get("items", [])
                 if d.get("rca") is not None]
    if rca_items:
        rca_items = sorted(rca_items, key=lambda d: d["rca"])
        ws = wb.create_sheet("Figure 8 - Kenya Services RCA")
        _add_series_bar(
            ws, 1, "Kenya's Service Export RCA by Category",
            [clean_label(d.get("category") or "") for d in rca_items],
            [d["rca"] for d in rca_items], colors=["2E75B6"])

    conc = a.concentration if hasattr(a, "concentration") else None
    conc = conc or {}
    if any(conc.get(k) is not None for k in ("kenya_hhi", "kenya_top3",
                                             "kenya_top5")):
        ws = wb.create_sheet("Figure 9 - Concentration Metrics")
        _add_bar_generic(
            ws, 1, "Kenya vs World: Export Concentration Metrics",
            ["HHI", "Top-3 Share", "Top-5 Share"],
            [("Kenya", [conc.get("kenya_hhi"), conc.get("kenya_top3"),
                        conc.get("kenya_top5")]),
             ("World", [conc.get("world_hhi"), conc.get("world_top3"),
                        conc.get("world_top5")])],
            colors=["2E75B6", "ED7D31"])

    div_items = [d for d in (a.diversification or {}).get("items", [])
                 if d.get("score") is not None and d.get("score", 0) > 0]
    if div_items:
        div_items = sorted(div_items, key=lambda d: d["score"])
        ws = wb.create_sheet("Figure 10 - Diversification Opportunities")
        _add_series_bar(
            ws, 1, "Kenya's Top Diversification Opportunities in Services",
            [clean_label(d.get("category") or "") for d in div_items],
            [round(d["score"], 4) for d in div_items], colors=["70AD47"])

    traj = getattr(a, "trajectory", {}) or {}
    shares_t = traj.get("shares") or {}
    years_t = traj.get("years") or []
    if years_t and any(shares_t.get(k) for k in ("high", "traditional", "other")):
        def _pct(vals):
            return [round((v or 0) * 100, 1) for v in vals]
        ws = wb.create_sheet("Figure 11 - Composition Trajectory")
        _add_bar_generic(
            ws, 1, "Kenya's Service Export Composition Over Time",
            [str(y) for y in years_t],
            [("High-Value Services", _pct(shares_t.get("high", []))),
             ("Traditional Services", _pct(shares_t.get("traditional", []))),
             ("Other Services", _pct(shares_t.get("other", [])))],
            colors=["2E75B6", "70AD47", "A5A5A5"], grouping="stacked")

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Quarterly report deliverable
# ---------------------------------------------------------------------------
def _q_value(cell, v, fmt):
    if v is None:
        cell.value = ""
        return
    cell.value = v
    cell.number_format = fmt


def _write_quarterly_table(ws, table, title,
                           flow_label="Value in Ksh. Billion"):
    """Write one parsed quarterly rank table (current vs previous-year quarter).

    Quarterly items carry per-month values for both years, a current/previous
    Total, a share, and (when comparing) an absolute change + % change.
    """
    months = table["months"]
    years = table["years"]
    comparison = bool(table.get("comparison"))
    y_prev = years[0] if years and comparison else None
    y_cur = years[1] if years and len(years) > 1 else None
    n_m = len(months)
    col_rank, col_name = 1, 2
    # column plan
    #  rank | name | [Y-1: m...m ] Total | [Y: m...m ] Total | Change | % | Share
    cur_start = col_name + 1
    prev_start = cur_start
    prev_ncols = 0
    if comparison:
        prev_start = cur_start
        prev_ncols = n_m + 1
        cur_start = prev_start + prev_ncols
    cur_ncols = n_m + 1
    chg_col = cur_start + cur_ncols          # absolute change
    pct_col = chg_col + 1 if comparison else None
    share_col = (pct_col + 1) if comparison else (cur_start + cur_ncols)
    last_col = share_col

    r = 1
    r = _title_row(ws, r, title, last_col)
    hdr = r
    # first header row: Rank, Name, "Q <year>" (merged month span + total),
    # then Change / % / Share (merged if present)
    for cc in range(1, last_col + 1):
        c = ws.cell(hdr, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    ws.cell(hdr, col_rank).value = "Rank"
    ws.cell(hdr, col_name).value = flow_label
    if comparison:
        ws.merge_cells(start_row=hdr, start_column=prev_start,
                       end_row=hdr, end_column=prev_start + prev_ncols - 1)
        ws.cell(hdr, prev_start).value = "%s" % (y_prev or "")
    ws.merge_cells(start_row=hdr, start_column=cur_start,
                   end_row=hdr, end_column=cur_start + cur_ncols - 1)
    ws.cell(hdr, cur_start).value = "%s" % (y_cur or "")
    if comparison:
        ws.merge_cells(start_row=hdr, start_column=chg_col,
                       end_row=hdr, end_column=pct_col)
        ws.cell(hdr, chg_col).value = "Change"
    ws.merge_cells(start_row=hdr, start_column=share_col,
                   end_row=hdr, end_column=share_col)
    ws.cell(hdr, share_col).value = "Share in %s %%" % (y_cur or "")

    # second header row: month names + Total (per year), then % under Change
    hdr2 = hdr + 1
    for cc in range(1, last_col + 1):
        c = ws.cell(hdr2, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    ws.cell(hdr2, col_rank).value = ""
    ws.cell(hdr2, col_name).value = ""
    if comparison:
        for k, m in enumerate(months):
            ws.cell(hdr2, prev_start + k).value = m
        ws.cell(hdr2, prev_start + n_m).value = "Total"
    for k, m in enumerate(months):
        ws.cell(hdr2, cur_start + k).value = m
    ws.cell(hdr2, cur_start + n_m).value = "Total"
    if comparison:
        ws.cell(hdr2, pct_col).value = "%"
    ws.cell(hdr2, share_col).value = ""

    body = hdr2 + 1
    first = None
    last = None
    for d in table.get("orows", table["items"]):
        if first is None:
            first = body
        b = d.get("kind") in ("total", "all_other")
        ws.cell(body, col_rank).value = d.get("rank")
        nc = ws.cell(body, col_name)
        nc.value = clean_label(d.get("label") or d.get("name"))
        nc.alignment = LEFT
        mvals = d.get("months") or [[], []]
        prev_vals = mvals[0] if len(mvals) > 0 else []
        cur_vals = mvals[1] if len(mvals) > 1 else []
        if comparison:
            for k in range(n_m):
                _q_value(ws.cell(body, prev_start + k),
                         prev_vals[k] if k < len(prev_vals) else None, "#,##0.0")
            _q_value(ws.cell(body, prev_start + n_m),
                     (d.get("totals") or [None, None])[0], "#,##0.0")
        for k in range(n_m):
            _q_value(ws.cell(body, cur_start + k),
                     cur_vals[k] if k < len(cur_vals) else None, "#,##0.0")
        _q_value(ws.cell(body, cur_start + n_m),
                 (d.get("totals") or [None, None])[1], "#,##0.0")
        if comparison:
            chg = d.get("change")
            if chg is not None:
                chg_cell = ws.cell(body, chg_col)
                chg_cell.value = round(chg, 2)
                chg_cell.number_format = "+#,##0.0;-#,##0.0;0.0"
            pctv = d.get("pct")
            p = ws.cell(body, pct_col)
            if pctv is not None:
                p.value = round(pctv, 4)
                p.number_format = "0.0%"
            p.alignment = CENTER
        sh = d.get("share")
        sc = ws.cell(body, share_col)
        if sh is not None:
            sc.value = round(sh / 100.0, 6) if sh > 1 else round(sh, 6)
            sc.number_format = "0.0%"
        sc.alignment = CENTER
        for cc in range(1, last_col + 1):
            cell = ws.cell(body, cc)
            if b:
                cell.font = Font(bold=True)
        last = body
        body += 1

    body_end = last if last is not None else hdr2
    _apply_borders(ws, hdr, 1, body_end, last_col, last_col)
    _fit_col_widths(ws, [7, 34] + [11] * ((2 * (n_m + 1)) if comparison
                                          else (n_m + 1))
                    + ([8, 8] if comparison else []) + [12])
    return {"first": hdr, "last_col": last_col, "body_end": body_end}


def build_quarterly_deliverable(a, cfg, out_path):
    """Editable Excel workbook for the quarterly export-performance report."""
    flat = cfg.get("report", {}).get("title_line2", "")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    specs = [
        ("Table 1 - Top Destination Markets", a.t1, "Partner"),
        ("Table 2 - Top Export Products", a.t2, "Product"),
        ("Annex 1 - Imports by Supplying Country", a.a1, "Country"),
        ("Annex 2 - Imports by Product", a.a2, "Product"),
    ]
    for title, table, flow in specs:
        if not table:
            continue
        ws = wb.create_sheet(title[:31])
        _write_quarterly_table(ws, table, title, flow)

    _add_quarterly_balance(wb, a)
    _add_quarterly_shares(wb, a)

    wb.save(out_path)
    return out_path


def _quarterly_slice_rows(a, table, year="cur", other_label="Other",
                          max_slices=8, min_pct=2.0, n=12):
    """(label, value) pairs consolidated like the reference doughnut PNG."""
    labels, values = [], []
    for name, v, _s in a.shares(table, n=n, year=year):
        if v:
            labels.append(name)
            values.append(v)
    return _consolidate_labels(labels, values, max_slices=max_slices,
                               min_pct=min_pct, other_label=other_label)


def _consolidate_labels(labels, values, max_slices=8, min_pct=2.0,
                        other_label="Other"):
    total = sum(v for v in values if v is not None) or 0.0
    if total <= 0:
        return [(l, v or 0.0) for l, v in zip(labels, values)]
    keep, other_v = [], 0.0
    for label, v in zip(labels, values):
        v = v or 0.0
        if len(keep) < max_slices and v / total * 100 >= min_pct and v > 0:
            keep.append((label, v))
        else:
            other_v += v
    if other_v > 0:
        keep.append((other_label, other_v))
    return keep


def _add_quarterly_shares(wb, a):
    """Editable doughnuts for Figure 1 (markets) and Figure 2 (products)."""
    for sheet, table, other, fig in (
            ("Figure 1 - Market Shares", a.t1, "Other markets", 1),
            ("Figure 2 - Product Shares", a.t2, "All others products", 2)):
        if not table or not table.get("grand", {}).get("totals"):
            continue
        ws = wb.create_sheet(sheet)
        top = 1
        for yr, year in ((a.year, "cur"), (a.year_prev, "prev")):
            if yr is None or not a.has_comparison and year == "prev":
                continue
            rows = _quarterly_slice_rows(
                a, table, year=year, other_label=other)
            if not rows:
                continue
            _add_doughnut(ws, top, "Share in %d" % yr, rows,
                          colors=[c.lstrip("#") for c in THEME],
                          offset_anchor="A")
            top += len(rows) + 12


def _add_quarterly_balance(wb, a):
    """Monthly exports / imports / balance bar (from t1 & a1 grand totals)."""
    t1, a1 = a.t1, a.a1
    if not t1 or not a1:
        return
    months = t1["months"]
    exp = list(t1["grand"]["months"][1])
    imp = list(a1["grand"]["months"][1])
    bal = [e - i for e, i in zip(exp, imp)]
    ws = wb.create_sheet("Balance by Month")
    r0 = 1
    ws.cell(r0, 1).value = "Month"
    ws.cell(r0, 2).value = "Exports"
    ws.cell(r0, 3).value = "Imports"
    ws.cell(r0, 4).value = "Balance"
    for cc in range(1, 5):
        c = ws.cell(r0, cc)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    for i, m in enumerate(months):
        rr = r0 + 1 + i
        ws.cell(rr, 1).value = m
        for cc, series in ((2, exp), (3, imp), (4, bal)):
            v = series[i] if i < len(series) else None
            c = ws.cell(rr, cc)
            _q_value(c, v if v is not None else "", "#,##0.0")
            c.alignment = RIGHT
    n = len(months)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Kenya %s %s Balance of Trade (Ksh. Billion)" % (
        a.quarter.title(), str(a.year))
    chart.height = 9
    chart.width = 18
    data = Reference(ws, min_col=1, min_row=r0, max_col=4, max_row=r0 + n)
    cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r0 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A%d" % (r0 + n + 2))
