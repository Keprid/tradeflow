#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
webapp/main.py
==============

FastAPI web app for the Trade Flow Report Generator.

One click, end to end:
  1. upload the six raw ITC downloads (or a zip of them), or the seven
     ready-made tables;
  2. pick a country config;
  3. the server builds Table 1-6 + Figure 1 (make_tables) and the Word report
     (generate_report), then serves the .docx for download.

Run (from the project root):
    python -m uvicorn webapp.main:app --host 127.0.0.1 --port 8000
"""

import csv
import html
import html.parser
import io
import json
import logging
import os
import re
import shutil
import sys
import time
import traceback
import uuid
import zipfile
from contextlib import asynccontextmanager
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
LOG = logging.getLogger("tradeflow.webapp")

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

import make_tables                       # noqa: E402
import generate_report as gr             # noqa: E402
import make_config                       # noqa: E402
import make_services_tables              # noqa: E402
import generate_services_report as gsr   # noqa: E402

WEBAPP_DIR = Path(__file__).resolve().parent
STATIC_DIR = WEBAPP_DIR / "static"
JOBS_DIR = WEBAPP_DIR / "jobs"
CONFIG_DIR = BASE_DIR / "config"
JOB_TTL_SECONDS = 24 * 3600              # old jobs are cleaned up after 1 day

ALLOWED_EXT = (".xlsx", ".xlsm", ".xls", ".xlsb", ".csv")
RAW_KEYWORDS = ("imports-from-world", "exports-to-world",
                "kenyas-imports-from", "kenyas-exports-to")
SERVICE_RAW_KEYWORDS = ("exported_services_for", "imported_services_for",
                        "services_exported_by", "services_imported_by",
                        "services_commercialized", "list_of_exporters_for",
                        "list_of_importers_for")

@asynccontextmanager
async def _lifespan(app: FastAPI):
    _cleanup_old_jobs()
    yield

app = FastAPI(title="Trade Flow Report Generator", lifespan=_lifespan)


# ---------------------------------------------------------------------------
# Job storage helpers
# ---------------------------------------------------------------------------
def _new_job_dir():
    jd = JOBS_DIR / uuid.uuid4().hex
    (jd / "uploads").mkdir(parents=True)
    return jd


def _cleanup_old_jobs():
    now = time.time()
    if not JOBS_DIR.exists():
        return
    for d in JOBS_DIR.iterdir():
        if d.is_dir():
            try:
                if now - d.stat().st_mtime > JOB_TTL_SECONDS:
                    shutil.rmtree(d, ignore_errors=True)
            except OSError:
                pass


def _save_upload(uploads_dir, upload: UploadFile):
    """Persist one uploaded file, keeping only its basename."""
    name = os.path.basename(upload.filename or "")
    if not name.lower().endswith(ALLOWED_EXT):
        raise HTTPException(400, f"Unsupported file type: {upload.filename}")
    dst = uploads_dir / name
    if dst.exists():
        base, ext = os.path.splitext(name)
        i = 1
        while dst.exists():
            dst = uploads_dir / f"{base}__{i}{ext}"
            i += 1
    with dst.open("wb") as f:
        shutil.copyfileobj(upload.file, f)
    return _convert_to_xlsx(dst)


def _extract_zip(zpath, dest):
    """Extract a zip safely (no path traversal), then flatten spreadsheets."""
    try:
        with zipfile.ZipFile(zpath) as z:
            for member in z.namelist():
                if member.endswith("/"):
                    continue
                safe = os.path.normpath(member)
                if safe.startswith("..") or os.path.isabs(safe):
                    raise HTTPException(400, f"Unsafe path inside zip: {member}")
                out = dest / safe
                out.parent.mkdir(parents=True, exist_ok=True)
                with z.open(member) as src, out.open("wb") as dst:
                    shutil.copyfileobj(src, dst)
    except zipfile.BadZipFile:
        raise HTTPException(400, f"{zpath.name} is not a valid zip file")
    _flatten_xlsx(dest)
    os.remove(zpath)


def _flatten_xlsx(dest):
    """Move every supported spreadsheet from nested folders up to `dest` (zip case)."""
    for root, _, files in os.walk(dest):
        for f in files:
            if not f.lower().endswith(ALLOWED_EXT):
                continue
            src = os.path.join(root, f)
            if os.path.dirname(src) == str(dest):
                continue
            base, ext = os.path.splitext(f)
            dst = os.path.join(dest, f)
            i = 1
            while os.path.exists(dst):
                dst = os.path.join(dest, f"{base}__{i}{ext}")
                i += 1
            shutil.move(src, dst)
    for root, dirs, _ in os.walk(dest, topdown=False):
        for d in dirs:
            p = os.path.join(root, d)
            if not os.listdir(p):
                os.rmdir(p)
    for f in list(dest.iterdir()):
        if f.suffix.lower() not in (".xlsx", ".xlsm"):
            _convert_to_xlsx(f)


# ---------------------------------------------------------------------------
# Spreadsheet normalisation: anything -> .xlsx (openpyxl can only read .xlsx)
# ---------------------------------------------------------------------------
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name, fallback):
    """A valid openpyxl sheet title (non-empty, <= 31 chars, no bad chars)."""
    n = _INVALID_SHEET_CHARS.sub("_", (name or "").strip())
    return n[:31] or fallback


def _read_text_any(path):
    """Read a text file, trying common encodings in turn."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="latin-1", errors="replace")


def _sniff_spreadsheet_kind(path):
    """Classify a non-.xlsx upload: 'xls' | 'html' | 'csv'.

    ITC / Trade Map downloads are often HTML tables (or comma/tab-delimited
    text) saved with a ``.xls`` extension. xlrd only understands the real
    binary format, so sniff the first bytes to pick the right parser.
    """
    with path.open("rb") as f:
        head = f.read(8192)
    low = head.lower()
    if b"<html" in low or b"<table" in low or b"<!doctype" in low \
            or b"<head" in low or b"<meta" in low:
        return "html"
    if b"\x00" in head:
        return "xls"
    try:
        text = head.decode("utf-8", "ignore")
    except Exception:
        return "xls"
    if any(ch in text for ch in ("\t", ",", ";", "|")):
        return "csv"
    return "xls"


class _HTMLTableParser(html.parser.HTMLParser):
    """Extract the largest <table> in an HTML file as a list of row lists."""

    def __init__(self):
        super().__init__()
        self.tables = []
        self._cur = None
        self._row = None
        self._data = []
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._cur = []
        elif tag == "tr" and self._cur is not None:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._data = []
            self._in_cell = True

    def handle_endtag(self, tag):
        if tag == "table" and self._cur is not None:
            if self._cur:
                self.tables.append(self._cur)
            self._cur = None
        elif tag == "tr" and self._row is not None:
            if self._row:
                self._cur.append(self._row)
            self._row = None
        elif tag in ("td", "th") and self._in_cell:
            self._row.append("".join(self._data).replace("\xa0", " ").strip())
            self._data = []
            self._in_cell = False

    def handle_data(self, data):
        if self._in_cell:
            self._data.append(data)


def _write_html_as_xlsx(src, out):
    """Parse an HTML file (often a .xls download) into the first .xlsx sheet."""
    parser = _HTMLTableParser()
    parser.feed(_read_text_any(src))
    if not parser.tables:
        raise ValueError("No HTML table found in file")
    table = max(parser.tables, key=len)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, row in enumerate(table, start=1):
        for c, v in enumerate(row, start=1):
            if v != "":
                ws.cell(row=r, column=c, value=v)
    wb.save(str(out))


def _write_xls_as_xlsx(src, out):
    import xlrd
    kind = _sniff_spreadsheet_kind(src)
    if kind == "html":
        _write_html_as_xlsx(src, out)
        return
    if kind == "csv":
        _write_csv_as_xlsx(src, out)
        return
    book = xlrd.open_workbook(str(src), formatting_info=False)
    wb = openpyxl.Workbook()
    for idx in range(book.nsheets):
        sheet = book.sheet_by_index(idx)
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = _safe_sheet_name(sheet.name, f"Sheet{idx + 1}")
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        v = xlrd.xldate_as_datetime(v, book.datemode)
                    except (ValueError, OverflowError):
                        pass
                elif cell.ctype == xlrd.XL_CELL_NUMBER \
                        and isinstance(v, float) and v.is_integer():
                    v = int(v)
                if v is not None:
                    ws.cell(row=r + 1, column=c + 1, value=v)
    wb.save(str(out))


def _write_xlsb_as_xlsx(src, out):
    from pyxlsb import open_workbook
    wb = openpyxl.Workbook()
    first = True
    with open_workbook(str(src)) as book:
        for name in book.sheets:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = _safe_sheet_name(name, f"Sheet{len(wb.sheetnames) + 1}")
            with book.get_sheet(name) as sheet:
                r = 1
                for row in sheet.rows():
                    for c, cell in enumerate(row):
                        v = cell.v
                        if isinstance(v, float) and v.is_integer():
                            v = int(v)
                        if v is not None:
                            ws.cell(row=r, column=c + 1, value=v)
                    r += 1
    wb.save(str(out))


def _write_csv_as_xlsx(src, out):
    text = _read_text_any(src)
    lines = text.splitlines()
    first = lines[0] if lines else ""
    counts = {d: first.count(d) for d in ("\t", ";", ",", "|")}
    delim = max(counts, key=counts.get) if any(counts.values()) else ","
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r, row in enumerate(csv.reader(io.StringIO(text), delimiter=delim), start=1):
        for c, v in enumerate(row, start=1):
            if v != "":
                ws.cell(row=r, column=c, value=v)
    wb.save(str(out))


def _convert_to_xlsx(src):
    """Convert an .xls/.xlsb/.csv upload into a .xlsx openpyxl can read.

    Returns the new .xlsx path (the original file is deleted). Files that are
    already .xlsx/.xlsm are returned untouched.
    """
    if src.suffix.lower() in (".xlsx", ".xlsm"):
        return src
    ext = src.suffix.lower()
    if ext not in (".xls", ".xlsb", ".csv"):
        return src
    out = src.with_suffix(".xlsx")
    i = 1
    while out.exists():
        out = src.with_name(f"{src.stem}__{i}.xlsx")
        i += 1
    try:
        if ext == ".xls":
            _write_xls_as_xlsx(src, out)
        elif ext == ".xlsb":
            _write_xlsb_as_xlsx(src, out)
        else:
            _write_csv_as_xlsx(src, out)
    except Exception as e:
        if out.exists():
            out.unlink()
        LOG.warning("Spreadsheet conversion failed for %s: %s", src.name, e)
        raise HTTPException(400, f"Could not read spreadsheet file {src.name} ({ext})")
    src.unlink()
    return out


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------
@app.on_event("startup")
def _startup():
    _cleanup_old_jobs()


@app.get("/", include_in_schema=False)
def index():
    return FileResponse(STATIC_DIR / "index.html")


app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/api/configs")
def api_configs():
    out = [{
        "id": "__auto__",
        "name": "Auto-detect country from data",
        "title": "",
        "year": 0,
    }]
    for p in sorted(CONFIG_DIR.glob("*.json")):
        if p.name == "config_template.json":
            continue
        try:
            cfg = gr.load_config(str(p))
        except Exception:
            continue
        out.append({
            "id": p.stem,
            "name": cfg["country"]["name"],
            "title": cfg["country"]["title"],
            "year": cfg["report"].get("year", 0),
        })
    return out


def _detect_mode(uploads_dir, report_type="goods"):
    names = [f.lower() for f in os.listdir(uploads_dir) if f.lower().endswith(ALLOWED_EXT)]
    if not names:
        return None, "No Excel files were uploaded."
    if report_type == "services":
        if any(any(k in n for k in SERVICE_RAW_KEYWORDS) for n in names):
            return "services_raw", ""
        uploaded = ", ".join(sorted(names))
        kw_list = ", ".join(SERVICE_RAW_KEYWORDS)
        return None, (
            f"Could not recognise the services upload set.\n"
            f"Uploaded files: {uploaded}\n"
            f"Expected filenames containing any of: {kw_list}")
    if any(any(k in n for k in RAW_KEYWORDS) for n in names):
        return "raw", ""
    if sum(1 for n in names if re.search(r"table\s*\d", n)) >= 6:
        return "ready", ""
    uploaded = ", ".join(sorted(names))
    raw_kw = ", ".join(RAW_KEYWORDS)
    return None, (
        f"Could not recognise the upload set.\n"
        f"Uploaded files: {uploaded}\n"
        f"Expected raw filenames containing any of: {raw_kw}\n"
        f"Or ready-made files named 'Table 1' through 'Table 6' plus "
        f"'Figure 1 Trade Balance'.")


def _detect_reporter(excel_dir):
    """Return the country the data actually describes, from Table 1's title."""
    return make_config.read_table1_meta(excel_dir)[0]


def _norm(s):
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _check_reporter_matches_config(cfg, excel_dir):
    detected = _detect_reporter(excel_dir)
    if not detected:
        return
    expected = cfg["country"]["name"]
    if _norm(detected) != _norm(expected):
        raise HTTPException(
            400,
            f"Country mismatch: the uploaded data is for '{detected}' but you "
            f"selected the '{expected}' configuration. Choose the matching "
            f"country in the dropdown (or pick 'Auto-detect country from "
            f"data') and try again.")


def _resolve_cfg(excel_dir, cfg_id, logs):
    """Resolve the config to use.

    ``cfg_id`` is normally a config/*.json stem. When it is empty or
    ``"__auto__"`` the country is detected from the data: an existing config
    for that country is reused, otherwise ``config/<country>.json`` is
    generated automatically from the data.
    """
    if cfg_id and cfg_id != "__auto__":
        cfg_path = CONFIG_DIR / f"{cfg_id}.json"
        if not cfg_path.exists():
            raise HTTPException(404, f"Config '{cfg_id}' not found")
        gr.cfg_path = os.path.abspath(str(cfg_path))
        cfg = gr.load_config(gr.cfg_path)
        _check_reporter_matches_config(cfg, excel_dir)
        return cfg, cfg_id

    detected = _detect_reporter(excel_dir)
    if not detected:
        raise HTTPException(
            400,
            "Could not detect the country from the uploaded data. Pick a "
            "country config manually (or upload the standard Table 1 with a "
            "country title).")
    for p in sorted(CONFIG_DIR.glob("*.json")):
        if p.name == "config_template.json":
            continue
        try:
            existing = gr.load_config(str(p))
        except Exception:
            continue
        if _norm(existing["country"]["name"]) == _norm(detected):
            gr.cfg_path = os.path.abspath(str(p))
            logs.append(f"Country detected from data: {detected} -> using existing config '{p.stem}'")
            return gr.load_config(gr.cfg_path), p.stem
    new_id = make_config.create_config_file(excel_dir, str(CONFIG_DIR))[0]
    cfg_path = CONFIG_DIR / f"{new_id}.json"
    gr.cfg_path = os.path.abspath(str(cfg_path))
    logs.append(f"Country detected from data: {detected} -> generated config/{new_id}.json automatically")
    return gr.load_config(gr.cfg_path), new_id


def _run_pipeline(job_dir, cfg_id, top_n, mode, logs):
    uploads = job_dir / "uploads"
    tables = job_dir / "tables"
    charts = job_dir / "charts"

    if mode == "raw":
        logs.append("Mode: raw ITC downloads -> building Table 1-6 + Figure 1 ...")
        try:
            make_tables.generate_tables(str(uploads), str(tables), top_n)
        except SystemExit as e:
            raise HTTPException(400, str(e))
        excel_dir = str(tables)
        logs.append(f"Tables written to {tables}")
    else:
        excel_dir = str(uploads)
        logs.append("Mode: ready-made tables -> skipping make_tables.py")

    cfg, cfg_id = _resolve_cfg(excel_dir, cfg_id, logs)

    os.makedirs(charts, exist_ok=True)
    report_name = f"KENYA-{gr.title_partner(cfg['country']['name'])} TRADE FLOW.docx"
    report_path = job_dir / report_name
    logs.append(f"Building report for {cfg['country']['name']} (charts -> {charts})")
    try:
        gr.build_report(cfg, excel_dir, str(report_path), str(charts))
    except SystemExit as e:
        raise HTTPException(400, str(e))
    logs.append(f"Report saved as {report_name}")

    all_tables = tables / "All Tables.xlsx"
    chart_png = charts / "chart_balance.png"
    if all_tables.exists() and chart_png.exists():
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            wb = openpyxl.load_workbook(str(all_tables))
            if "Figure 1" in wb.sheetnames:
                ws_fig = wb["Figure 1"]
                ws_fig._charts = []
                img = XLImage(str(chart_png))
                year_cols = sum(1 for cell in ws_fig[3] if cell.value is not None) - 1
                anchor_row = 8 if year_cols <= 5 else 9
                img.anchor = f"A{anchor_row}"
                ws_fig.add_image(img)
                wb.save(str(all_tables))
                logs.append("Replaced Figure 1 chart in All Tables.xlsx with correct PNG")
        except Exception as exc:
            logs.append(f"Warning: could not replace Figure 1 chart: {exc}")

    manifest = {
        "mode": mode,
        "report_name": report_name,
        "excel_dir": os.path.relpath(excel_dir, job_dir),
        "config": cfg_id,
    }
    (job_dir / "manifest.json").write_text(
        json.dumps(manifest), encoding="utf-8")
    return report_path, manifest


def _run_services_pipeline(job_dir, cfg_id, top_n, logs):
    uploads = job_dir / "uploads"
    tables = job_dir / "tables"
    charts = job_dir / "charts"

    logs.append("Mode: raw ITC services downloads -> building service tables ...")
    try:
        make_services_tables.generate_service_tables(
            str(uploads), str(tables), top_n)
    except SystemExit as e:
        raise HTTPException(400, str(e))
    excel_dir = str(tables)
    logs.append(f"Service tables written to {tables}")

    if not cfg_id or cfg_id == "__auto__":
        # Auto-select services_world config when report_type is services
        cfg_id = "services_world"

    cfg_path = CONFIG_DIR / f"{cfg_id}.json"
    if not cfg_path.exists():
        raise HTTPException(404, f"Config '{cfg_id}' not found")
    gr.cfg_path = os.path.abspath(str(cfg_path))
    cfg = gr.load_config(gr.cfg_path)

    os.makedirs(charts, exist_ok=True)
    report_name = f"KENYA-{gr.title_partner(cfg['country']['name'])} SERVICES TRADE FLOW.docx"
    report_path = job_dir / report_name
    logs.append(f"Building services report for {cfg['country']['name']} (charts -> {charts})")
    try:
        gsr.build_services_report(cfg, excel_dir, str(report_path), str(charts))
    except SystemExit as e:
        raise HTTPException(400, str(e))
    logs.append(f"Report saved as {report_name}")

    all_tables = tables / "All Service Tables.xlsx"
    chart_png = charts / "chart_balance.png"
    if all_tables.exists() and chart_png.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            wb = openpyxl.load_workbook(str(all_tables))
            if "Figure 1" in wb.sheetnames:
                ws_fig = wb["Figure 1"]
                ws_fig._charts = []
                img = XLImage(str(chart_png))
                img.anchor = "A8"
                ws_fig.add_image(img)
                wb.save(str(all_tables))
                logs.append("Replaced Figure 1 chart in All Service Tables.xlsx with correct PNG")
        except Exception as exc:
            logs.append(f"Warning: could not replace Figure 1 chart: {exc}")

    manifest = {
        "mode": "services_raw",
        "report_type": "services",
        "report_name": report_name,
        "excel_dir": os.path.relpath(excel_dir, job_dir),
        "config": cfg_id,
    }
    (job_dir / "manifest.json").write_text(
        json.dumps(manifest), encoding="utf-8")
    return report_path, manifest


@app.post("/api/run")
async def api_run(config: str = Form("__auto__"), top: int = Form(20),
                  report_type: str = Form("goods"),
                  files: list[UploadFile] = File(default=None),
                  zipfile: UploadFile | None = File(default=None)):
    files = files or []
    if not files and zipfile is None:
        raise HTTPException(400, "Please select files (or a zip) to upload.")
    if top < 1:
        raise HTTPException(400, "top must be >= 1")
    if report_type not in ("goods", "services"):
        raise HTTPException(400, "report_type must be 'goods' or 'services'")

    job_dir = _new_job_dir()
    try:
        for up in files:
            _save_upload(job_dir / "uploads", up)
        if zipfile is not None:
            zname = os.path.basename(zipfile.filename or "upload.zip")
            zpath = job_dir / "uploads" / zname
            with zpath.open("wb") as f:
                shutil.copyfileobj(zipfile.file, f)
            _extract_zip(zpath, job_dir / "uploads")

        mode, err = _detect_mode(job_dir / "uploads", report_type)
        if mode is None:
            raise HTTPException(400, err)

        LOG.info("POST /api/run config=%s top=%s mode=%s report_type=%s job=%s",
                 config, top, mode, report_type, job_dir.name)

        logs = []
        if report_type == "services":
            report_path, manifest = _run_services_pipeline(
                job_dir, config, top, logs)
        else:
            report_path, manifest = _run_pipeline(
                job_dir, config, top, mode, logs)
        return {
            "job_id": job_dir.name,
            "report_name": manifest["report_name"],
            "report_url": f"/api/download/{job_dir.name}",
            "tables_url": f"/api/tables/{job_dir.name}",
            "mode": mode,
            "log": logs,
        }
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except SystemExit as e:
        LOG.error("Pipeline aborted (SystemExit) config=%s job=%s: %s",
                  config, job_dir.name, e)
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(400, f"Pipeline aborted: {e}")
    except Exception as e:
        LOG.error("Pipeline failed for config=%s top=%s job=%s",
                  config, top, job_dir.name)
        LOG.error("%s", traceback.format_exc())
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(500, f"Pipeline failed: {e}")


@app.get("/api/download/{job_id}")
def api_download(job_id: str):
    job_dir = JOBS_DIR / job_id
    manifest_path = job_dir / "manifest.json"
    if not manifest_path.exists():
        raise HTTPException(404, "Job not found (or expired).")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    report = job_dir / manifest["report_name"]
    if not report.exists():
        raise HTTPException(404, "Report not found.")
    return FileResponse(str(report), filename=report.name,
                        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@app.get("/api/tables/{job_id}")
def api_tables(job_id: str):
    job_dir = JOBS_DIR / job_id
    manifest_path = job_dir / "manifest.json"
    if not manifest_path.exists():
        raise HTTPException(404, "Job not found (or expired).")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    src = job_dir / manifest["excel_dir"]
    if not src.is_dir():
        raise HTTPException(404, "Tables not found.")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(src.glob("*.xlsx")):
            z.write(str(f), arcname=f.name)
    buf.seek(0)
    return Response(content=buf.getvalue(),
                    media_type="application/zip",
                    headers={"Content-Disposition": f"attachment; filename=tables-{job_id}.zip"})


# ---------------------------------------------------------------------------
# Convenience: run with `python webapp/main.py`
# ---------------------------------------------------------------------------
def main():
    import uvicorn
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    print(f"Trade Flow Report Generator -> http://{host}:{port}")
    print("Upload the six raw ITC files (or a zip), or the seven ready tables.")
    uvicorn.run(app, host=host, port=port, log_level="info")


if __name__ == "__main__":
    main()
