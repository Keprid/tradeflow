#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_config.py
==============

Auto-generate config/<country>.json from the trade data, so a new country
can be added without hand-writing a config file.

Given a folder of ready-made tables (Table 1-6 + Figure 1) or the raw ITC
downloads, this module:
  * detects the country from the Table 1 title;
  * reads the report years from the table columns;
  * builds a config with the derived fields (country name / title /
    possessive, report year, month-year) and placeholders ("...") for the
    facts that must be researched (capital, quick facts, map image, ...).

CLI usage:
    python make_config.py <excel-folder> [--output config/<country>.json]
"""

import argparse
import datetime
import json
import os
import re
import shutil
import sys
import tempfile

import openpyxl

import make_tables

QUICK_FACTS = [
    ("Official name", "{name}"),
    ("Head of state", "..."),
    ("Head of government", "..."),
    ("Capital", "..."),
    ("Form of government", "..."),
    ("Official language", "..."),
    ("Total area (sq km)", "..."),
    ("Population", "..."),
    ("Population rank", "..."),
    ("Population projection 2030", "..."),
    ("Population density", "..."),
    ("Urban / rural population", "..."),
    ("Life expectancy at birth", "..."),
    ("Literacy (age 15 and over)", "..."),
    ("Monetary unit", "..."),
    ("Currency exchange rate", "..."),
    ("GNI", "..."),
    ("GNI per capita", "..."),
]


def _slug(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _possessive(name):
    return name + "'" if name.rstrip().lower().endswith("s") else name + "'s"


def read_table1_meta(dirpath):
    """Return (reporter, years) read from the Table 1 xlsx in ``dirpath``.

    Works for the ready-made tables produced by this pipeline (and the
    sample format it mirrors). Returns ``(None, [])`` when the Table 1 file
    cannot be found or parsed.
    """
    t1 = None
    for f in sorted(os.listdir(dirpath)):
        low = f.lower()
        if "table 1" in low and low.endswith((".xlsx", ".xlsm")):
            t1 = os.path.join(dirpath, f)
            break
    if not t1:
        return None, []

    try:
        wb = openpyxl.load_workbook(t1, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c.value for c in r] for r in ws.iter_rows(max_row=8)]
        wb.close()
    except Exception:
        return None, []

    reporter = None
    for r in rows[:3]:
        for cell in r[:4]:
            t = str(cell or "")
            if "List of Exporters to " in t:
                reporter = t.split("List of Exporters to ")[1].split("\n")[0].strip()
                break
            if "Import Source Markets" in t:
                reporter = t.replace("Import Source Markets", "").strip()
                break
        if reporter:
            break

    years = []
    for r in rows[1:5]:
        run = []
        for cell in r:
            try:
                v = int(str(cell).strip())
            except (TypeError, ValueError):
                v = None
            if v is not None and 1950 <= v <= 2100 and (not run or v == run[-1] + 1):
                run.append(v)
            elif run and len(run) >= 2:
                years = run
                break
            else:
                run = []
        if years:
            break
        if len(run) >= 2:
            years = run
            break
    return reporter, years


def detect(excel_dir):
    """Return (reporter, years) for a ready-tables or raw-files folder.

    Raw ITC downloads are first turned into Table 1-6 via make_tables
    (in a temporary folder) before the country and years are read.
    """
    reporter, years = read_table1_meta(excel_dir)
    if reporter:
        return reporter, years
    tmp = tempfile.mkdtemp(prefix="tradeflow_cfg_")
    try:
        make_tables.generate_tables(excel_dir, tmp, 20)
        return read_table1_meta(tmp)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def build_config(excel_dir):
    """Return (config_dict, slug) derived from the data in ``excel_dir``."""
    reporter, years = detect(excel_dir)
    if not reporter:
        raise ValueError(
            f"Could not detect the country from the data in '{excel_dir}'. "
            f"Is there a Table 1 file (or the six raw ITC downloads)?")
    name = reporter.strip()
    slug = _slug(name)
    year = years[-1] if years else datetime.date.today().year

    cfg = {
        "country": {
            "name": name,
            "short": name,
            "title": name.upper(),
            "adjective": name,
            "possessive": _possessive(name),
            "capital": "...",
        },
        "report": {
            "month_year": datetime.date.today().strftime("%B %Y").upper(),
            "year": year,
            "title_line1": f"KENYA- {name.upper()} TRADE FLOW",
            "title_line2": "ANALYSIS REPORT",
        },
        "world_trade": {
            "world_exports_usd_billion": 25800,
            "world_imports_usd_billion": 25200,
        },
        "export_potential": {
            "image": f"../assets/export_potential_{slug}.png",
            "paragraphs": [
                f"The products with greatest export potential from Kenya to {name} are ...",
                "Kenya has the highest supply capacity in ...",
            ],
        },
        "map": {
            "image": f"../assets/map_{slug}.png",
            "source": "Google map",
        },
        "quick_facts": [[k, v.format(name=name)] for k, v in QUICK_FACTS],
        "references": [
            "Google Maps",
            "World Bank (World Development Indicators; Macro Poverty Outlook), IMF (World Economic Outlook)",
            "WTO (member information, tariff and trade data)",
            "National statistics office of the country, if applicable",
        ],
    }
    return cfg, slug


def create_config_file(excel_dir, config_dir):
    """Build the config and write config/<slug>.json. Returns (slug, path)."""
    cfg, slug = build_config(excel_dir)
    path = os.path.join(config_dir, f"{slug}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return slug, path


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Auto-generate config/<country>.json from the trade data.")
    ap.add_argument("excel_dir", help="folder with ready-made tables or raw ITC downloads")
    ap.add_argument("--output", "-o", default=None,
                    help="config path (default config/<country-slug>.json)")
    args = ap.parse_args(argv)

    cfg, slug = build_config(args.excel_dir)
    out = args.output or os.path.join("config", f"{slug}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    print(f"[1/1] Wrote {out} (country: {cfg['country']['name']}, "
          f"report year: {cfg['report']['year']})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
